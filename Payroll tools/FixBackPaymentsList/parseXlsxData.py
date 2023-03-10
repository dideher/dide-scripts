from openpyxl import load_workbook


def parseXlsxData(fileName):
    workbook = load_workbook(filename=fileName)
    sheet = workbook.active

    data = list()

    for r, row in enumerate(sheet.iter_rows()):
        entry = list()
        for c, cell in enumerate(row):
            if c > 14:
                break

            if cell.value is None or cell.value == '':
                entry.append("")
            else:
                text = str(cell.value)
                if c == 0 and len(text) == 8:
                    text = '0' + text
                if c < 7 or c > 12 or r == 0:
                    entry.append(text)
                else:
                    entry.append(float(text.replace(',', '.')))

        data.append(entry)

    return data
