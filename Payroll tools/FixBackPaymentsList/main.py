from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from openpyxl import *
from openpyxl.utils import get_column_letter
from parseXlsxData import *
import xlrd


class GUI():
    def __init__(self):
        self.window = Tk()

        self.window.title("Διόρθωση λίστας αναδρομικών αποδοχών")
        self.window.resizable(False, False)
        self.create_widgets()

    def getInputFilename(self):
        fName = filedialog.askopenfilename(initialdir=".", title="Επιλέξτε το αρχείο xls",
                                           filetypes=(("xls files", "*.xls"), ("all files", "*.*")))

        if fName == "":
            return

        self.btnOpenInputFile.configure(state='disabled')
        self.inputFilename.set(fName)
        self.convert_xls(fName)
        self.data = parseXlsxData(fName + 'x')
        self.btnRun.configure(state='normal')

    def setColsWidth(self, ws):
        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

    def saveFile(self, data, outputFile):
        wb = Workbook()
        ws = wb.active

        for entry in data:
            ws.append(entry)

        self.setColsWidth(ws)

        notSaved = True

        while notSaved:
            try:
                wb.save(outputFile)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message=f"Παρακαλώ κλείστε το αρχείο '{outputFile}' ώστε να ολοκληρωθεί η αποθήκευση.")
            else:
                notSaved = False

    def convert_xls(self, fname):
        book_xls = xlrd.open_workbook(fname)
        book_xlsx = Workbook()

        sheet_names = book_xls.sheet_names()
        for sheet_index, sheet_name in enumerate(sheet_names):
            sheet_xls = book_xls.sheet_by_name(sheet_name)
            if sheet_index == 0:
                sheet_xlsx = book_xlsx.active
                sheet_xlsx.title = sheet_name
            else:
                sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

            for row in range(0, sheet_xls.nrows):
                for col in range(0, sheet_xls.ncols):
                    sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)

            self.setColsWidth(sheet_xlsx)

        book_xlsx.save(fname + 'x')

    def get_types(self, afm):
        teacher_types = ''
        for entry in self.teachers[afm]:
            if teacher_types == '':
                teacher_types += entry[6]
            else:
                teacher_types += '-' + entry[6]

        return teacher_types

    def fix_data(self):
        self.teachers = dict()

        for entry in self.data[1:]:
            afm = entry[0]

            if afm not in self.teachers:
                self.teachers[afm] = list()

            self.teachers[afm].append(entry)

        self.fixed_data = list()
        self.fixed_data.append(self.data[0])

        count_1_2_20 = 0
        count_1_20 = 0
        count_2_20 = 0
        count_20 = 0

        tts = set()
        ttd = dict()

        for afm in self.teachers:
            teacher_types = self.get_types(afm)
            tts.add(teacher_types)

            if teacher_types not in ttd:
                ttd[teacher_types] = list()

            ttd[teacher_types].append(self.teachers[afm][0][0:5])

            if teacher_types == '1-2-20':
                count_1_2_20 += 1
                for i in range(7, 10):
                    self.teachers[afm][0][i] = f'{self.teachers[afm][0][i]:.2f}'.replace('.', ',')
                for i in range(11, 13):
                    self.teachers[afm][0][i] = f'{self.teachers[afm][0][i]:.2f}'.replace('.', ',')

                self.fixed_data.append(self.teachers[afm][0])

                sum_entry = list()
                sum_entry += self.teachers[afm][0][:6]
                sum_entry.append('2')
                sum_entry.append(f'{self.teachers[afm][1][7] + self.teachers[afm][2][7]:.2f}'.replace('.', ','))
                sum_entry.append(f'{self.teachers[afm][1][8] + self.teachers[afm][2][8]:.2f}'.replace('.', ','))
                sum_entry.append(f'{self.teachers[afm][1][9] + self.teachers[afm][2][9]:.2f}'.replace('.', ','))
                sum_entry.append('')
                sum_entry.append(f'{self.teachers[afm][1][11] + self.teachers[afm][2][11]:.2f}'.replace('.', ','))
                sum_entry.append(f'{self.teachers[afm][1][12] + self.teachers[afm][2][12]:.2f}'.replace('.', ','))
                sum_entry += ['0', '0']

                self.fixed_data.append(sum_entry)
            elif teacher_types == '1-20':
                count_1_20 += 1
                self.teachers[afm][1][6] = '2'
                for i in range(7, 10):
                    self.teachers[afm][0][i] = f'{self.teachers[afm][0][i]:.2f}'.replace('.', ',')
                    self.teachers[afm][1][i] = f'{self.teachers[afm][1][i]:.2f}'.replace('.', ',')
                for i in range(11, 13):
                    self.teachers[afm][0][i] = f'{self.teachers[afm][0][i]:.2f}'.replace('.', ',')
                    self.teachers[afm][1][i] = f'{self.teachers[afm][1][i]:.2f}'.replace('.', ',')

                self.fixed_data.append(self.teachers[afm][0])
                self.fixed_data.append(self.teachers[afm][1])
            elif teacher_types == '2-20':
                count_2_20 += 1
                sum_entry = list()
                sum_entry += self.teachers[afm][0][:6]
                sum_entry.append('2')
                sum_entry.append(f'{self.teachers[afm][0][7] + self.teachers[afm][1][7]:.2f}'.replace('.', ','))
                sum_entry.append(f'{self.teachers[afm][0][8] + self.teachers[afm][1][8]:.2f}'.replace('.', ','))
                sum_entry.append(f'{self.teachers[afm][0][9] + self.teachers[afm][1][9]:.2f}'.replace('.', ','))
                sum_entry.append('')
                sum_entry.append(f'{self.teachers[afm][0][11] + self.teachers[afm][1][11]:.2f}'.replace('.', ','))
                sum_entry.append(f'{self.teachers[afm][0][12] + self.teachers[afm][1][12]:.2f}'.replace('.', ','))
                sum_entry += ['0', '0']

                self.fixed_data.append(sum_entry)
            elif teacher_types == '20':
                count_20 += 1
                self.teachers[afm][0][6] = '2'
                for i in range(7, 10):
                    self.teachers[afm][0][i] = f'{self.teachers[afm][0][i]:.2f}'.replace('.', ',')
                for i in range(11, 13):
                    self.teachers[afm][0][i] = f'{self.teachers[afm][0][i]:.2f}'.replace('.', ',')

                self.fixed_data.append(self.teachers[afm][0])
            else:
                for entry in self.teachers[afm]:
                    for i in range(7, 10):
                        entry[i] = f'{entry[i]:.2f}'.replace('.', ',')
                    for i in range(11, 13):
                        entry[i] = f'{entry[i]:.2f}'.replace('.', ',')

                    self.fixed_data.append(entry)

        print(20 * '-', 'Συνδυασμοί αποδοχών', 20 * '-')
        for tt in ttd:
            print(tt)

        print(20 * '-', 'Εκπαιδευτικοί ανά Συνδυασμό αποδοχών', 20 * '-')
        for tt in ttd:
            print(20 * '-', tt, 20 * '-')

            for entry in ttd[tt]:
                print(entry)

        err_msg = ''
        for entry in tts:
            if '20' in entry and entry not in ['1-2-20', '1-20', '2-20', '20']:
                err_msg += f'{entry}\n'

        if err_msg != '':
            showwarning(title='Προσοχή',
                        message=f'Δεν υπάρχει πρόβλεψη για τους παρακάτω συνδυασμούς τύπων αποδοχών:\n'
                                f' {err_msg} \nΕνημερώστε τον Χάρη :-)')

        showinfo(title='Στατιστικά',
                 message=f'- Αρχικές εγγραφές: {len(self.data[1:])}\n'
                         f'- Εγγραφές 1-2-20: {count_1_2_20}\n'
                         f'- Εγγραφές 1-20: {count_1_20}\n'
                         f'- Εγγραφές 2-20: {count_2_20}\n'
                         f'- Εγγραφές 20: {count_20}\n'
                         f'- Πλήθος συγχωνεύσεων: {count_1_2_20 + count_2_20}\n'
                         f'- Εγγραφές μετά τη διόρθωση: {len(self.fixed_data[1:])}\n'
                         f'- Διαφορά πλήθους εγγραφών: {len(self.data[1:]) - len(self.fixed_data[1:])}\n')

    def run(self):
        self.btnRun.configure(state='disabled')

        inputFile = self.inputFilename.get()
        outputFile = inputFile.replace(".xls", "_fix.xlsx")

        self.fix_data()

        self.saveFile(self.fixed_data, outputFile)

        showinfo(title='Ολοκλήρωση Εκτέλεσης',
                 message=f'Η διόρθωση της λίστας των αναδρομικών αποδοχών ολοκληρώθηκε.')
        self.window.destroy()

    def create_widgets(self):
        self.fData = Frame(self.window)

        self.lInputFile = Label(self.fData, text="Αρχείο xls:")
        self.lInputFile.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.inputFilename = StringVar()
        self.inputFilename.set('')
        self.ntrInputFilename = Entry(self.fData, width=128, state='readonly', textvariable=self.inputFilename)
        self.ntrInputFilename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenInputFile = Button(self.fData, text="Επιλέξτε αρχείο...", command=self.getInputFilename)
        self.btnOpenInputFile.grid(column=2, row=0, padx=10, pady=10)

        self.btnRun = Button(self.fData, text="Εκτέλεση", command=self.run, state='disabled')
        self.btnRun.grid(column=1, row=10, padx=10, pady=10)

        self.fData.pack()


gui = GUI()
gui.window.mainloop()
