import xml.etree.ElementTree as ET
from tkinter import *
from tkinter import filedialog
from tkinter.ttk import *
from tkinter.messagebox import showinfo, showwarning
from openpyxl import *
from openpyxl.utils import get_column_letter


class GUI:
    def __init__(self):
        self.window = Tk()

        self.window.title("Δημιουργία αρχείου xlsx από το xml για τις απολύσεις στην Εργάνη")
        self.window.resizable(False, False)
        self.create_widgets()

    def parse_xml_data(self, filename):
        tree = ET.parse(filename)
        root = tree.getroot()

        fields = ['f_afm',
                  'f_eponymo',
                  'f_onoma',
                  'f_onoma_patros',
                  'f_onoma_mitros',
                  'f_sex',
                  'f_birthdate',
                  'f_marital_status',
                  'f_arithmos_teknon',
                  'f_amka',
                  'f_ar_taytothtas',
                  'ΚΛΑΔΟΣ',
                  'f_eidikothta',
                  'f_proslipsidate',
                  'f_lixisymbashdate',
                  'f_apodoxes',
                  'f_epipedo_morfosis',
                  'f_kathestosapasxolisis',
                  'f_apolysisdate']

        self.data = list()
        for child in root:
            entry = list()
            for field in fields:
                if field == 'ΚΛΑΔΟΣ':
                    value = ''
                else:
                    value = child.find(field).text

                if field == 'f_sex':
                    # (0) ΑΝΤΡΑΣ - (1) ΓΥΝΑΙΚΑ
                    if value == '0':
                        value = 'ΑΝΤΡΑΣ'
                    else:
                        value = 'ΓΥΝΑΙΚΑ'
                elif field == 'f_marital_status':
                    # (0) ΑΓΑΜΟΣ - (1) ΕΓΓΑΜΟΣ - (2) ΔΙΑΖΕΥΓΜΕΝΟΣ - (3) ΧΗΡΟΣ
                    if value == '0':
                        value = 'ΑΓΑΜΟΣ'
                    elif value == '1':
                        value = 'ΕΓΓΑΜΟΣ'
                    elif value == '2':
                        value = 'ΔΙΑΖΕΥΓΜΕΝΟΣ'
                    else:
                        value = 'ΧΗΡΟΣ'
                elif field == 'f_kathestosapasxolisis':
                    # (0) ΠΛΗΡΗΣ (1) ΜΕΡΙΚΗ (2) ΕΚ ΠΕΡΙΤΡΟΠΗΣ
                    if value == '0':
                        value = 'ΠΛΗΡΗΣ'
                    elif value == '1':
                        value = 'ΜΕΡΙΚΗ'
                    else:
                        value = 'ΕΚ ΠΕΡΙΤΡΟΠΗΣ'
                elif field == 'f_epipedo_morfosis':
                    if value == '11':
                        value = 'ΑΕΙ'

                entry.append(value)

            self.data.append(entry)

    def export_data(self):
        wb = Workbook()
        ws = wb.active
        header = ['f_afm',
                  'f_eponymo',
                  'f_onoma',
                  'f_onoma_patros',
                  'f_onoma_mitros',
                  'f_sex',
                  'f_birthdate',
                  'f_marital_status',
                  'f_arithmos_teknon',
                  'f_amka',
                  'f_ar_taytothtas',
                  'ΚΛΑΔΟΣ',
                  'f_eidikothta',
                  'f_proslipsidate',
                  'f_lixisymbashdate',
                  'f_apodoxes',
                  'f_epipedo_morfosis',
                  'f_kathestosapasxolisis',
                  'f_apolysisdate']

        ws.append(header)

        for entry in self.data:
            ws.append(entry)

        self.set_cols_width(ws)

        output_file = "ergani_from_xml.xlsx"
        self.safe_save(wb, output_file)

    def set_cols_width(self, ws):
        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

    def safe_save(self, wb, output_file):
        not_saved = True

        while not_saved:
            try:
                wb.save(output_file)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message=f"Παρακαλώ κλείστε το αρχείο '{output_file}' ώστε να ολοκληρωθεί η αποθήκευση.")
            else:
                not_saved = False

    def get_data(self):
        f_name = filedialog.askopenfilename(initialdir="./data/",
                                            title="Επιλέξτε το αρχείο xml",
                                            filetypes=(("xml files", "*.xml"), ("all files", "*.*")))

        if f_name == "":
            return

        self.data_filename.set(f_name)
        self.parse_xml_data(f_name)
        self.btn_run.configure(state='normal')

    def run(self):
        self.export_data()
        showinfo(title="Ολοκλήρωση εκτέλεσης",
                 message="Η δημιουργία του αρχείου xlsx για τις απολύσεις στην Εργάνη ολοκληρώθηκε.")
        self.window.destroy()

    def create_widgets(self):
        self.f_data = Frame(self.window)

        self.l_data = Label(self.f_data, text="Αρχείο xml:")
        self.l_data.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.data_filename = StringVar()
        self.ntr_data_filename = Entry(self.f_data, width=128, state='readonly', textvariable=self.data_filename)
        self.ntr_data_filename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btn_get_data = Button(self.f_data, text="Επιλέξτε αρχείο...", command=self.get_data)
        self.btn_get_data.grid(column=2, row=0, padx=10, pady=10)

        self.btn_run = Button(self.f_data, text="Εκτέλεση", command=self.run, state='disabled')
        self.btn_run.grid(column=0, columnspan=3, row=10, padx=10, pady=10)

        self.f_data.pack()


gui = GUI()
gui.window.mainloop()
