from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from openpyxl import load_workbook
import os
from shutil import copyfile


class GUI():
    def __init__(self):
        self.window = Tk()

        self.window.title("Μετονομασία αρχείων με ΑΦΜ-Ονοματεπώνυμο")
        self.window.resizable(False, False)
        self.create_widgets()

    def verify_afm(self, afm):
        if len(afm) != 9:
            # print(f'Το ΑΦΜ: "{afm}" δεν είναι έγκυρο.\nΠρέπει να έχει εννέα ψηφία.')
            return False

        if not afm.isdigit():
            # print(f'Το ΑΦΜ: "{afm}" δεν είναι έγκυρο.\nΠρέπει να αποτελείται από ψηφία.')
            return False

        chcknumbers = [0, 2, 4, 8, 16, 32, 64, 128, 256]
        lchcknumbers = len(chcknumbers) - 1
        sum = 0

        for i in range(9):
            sum += (int(afm[i]) * chcknumbers[lchcknumbers - i])

        ch_digit = int(afm[8])

        ypoloipo = sum % 11

        if ypoloipo == 10:
            ypoloipo = 0

        return (ypoloipo == ch_digit)

    def getXlsxFilename(self):
        fName = filedialog.askopenfilename(initialdir=".", title="Επιλέξτε το αρχείο xlsx",
                                           filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if fName == "":
            return

        self.xlsxFilename.set(fName)
        self.xlsxData = self.getUnique(self.parseXlsxData(fName))

        status = self.checkAFMnamesList()
        if not status:
            showwarning(title='Λάθη στο αρχείο αντιστοίχισης',
                        message='Υπήρξαν λάθη στο αρχείο αντιστοίχισης ΑΦΜ σε Ονοματεπώνυμο.')
            return

        self.cbAction.configure(state='readonly')

    def getUnique(self, l):
        cleanList = list()

        for entry in l:
            if entry not in cleanList:
                cleanList.append(entry)

        return cleanList

    def parseXlsxData(self, fileName):
        workbook = load_workbook(filename=fileName)
        sheet = workbook.active

        data = list()

        for row in sheet.iter_rows():
            entry = list()
            for cell in row:
                if cell.value is None:
                    entry.append("")
                else:
                    text1 = (
                        str(cell.value).upper().replace(".", ". ").replace(" .", ". ").replace("Ά", "Α")
                            .replace("Έ", "Ε").replace("Ή", "Η").replace("Ί", "Ι").replace("Ϊ́", "Ϊ").replace("Ύ", "Υ")
                            .replace("Ϋ́", "Ϋ").replace("Ό", "Ο").replace("Ώ", "Ω").strip())

                    text2 = re.sub(r'([ ]+)', r' ', text1)
                    entry.append(re.sub(r'([0-9]+)Ο', r'\1ο', text2))

            data.append(entry)

        return self.getUnique(data[1:])

    def cbActionSelect(self, eventObject):
        if self.cbAction.current() != -1:
            self.ntrInputDir.configure(state='readonly')
            self.btnOpenInputDir.configure(state='normal')

    def getInputDir(self):
        dName = filedialog.askdirectory(initialdir="./data/", title="Επιλέξτε τον φάκελο με τα αρχεία για μετονομασία")

        if dName == "":
            return

        self.inputDir.set(dName)
        self.ntrOutputDir.configure(state='readonly')
        self.btnOpenOutputDir.configure(state='normal')

    def getOutputDir(self):
        dName = filedialog.askdirectory(initialdir="./data/", title="Επιλέξτε τον φάκελο που θα αποθηκευτούν τα αρχεία")

        if dName == "":
            return

        self.outputDir.set(dName)
        self.btnRun.configure(state='normal')

    def checkAFMnamesList(self):
        self.afmToName = dict()
        self.nameToAFM = dict()
        errors = ''

        errCount = 0
        for entry in self.xlsxData:
            afm = entry[0]
            name = entry[1]

            if not self.verify_afm(afm):
                errCount += 1
                errors += f'{errCount:3}) To ΑΦΜ "{afm}" δεν είναι σωστό.\n'

            if afm not in self.afmToName:
                self.afmToName[afm] = name
            else:
                errCount += 1
                errors += f'{errCount:3}) To ΑΦΜ "{afm}" εμφανίζεται και σε άλλο άτομο.\n'

            if name not in self.nameToAFM:
                self.nameToAFM[name] = afm
            else:
                errCount += 1
                errors += f'{errCount:3}) To Ονοματεπώνυμο "{name}" εμφανίζεται και σε άλλο άτομο.\n'

        if errors != '':
            print(10 * '-', ' Λάθη στο αρχείο αντιστοίχισης ΑΦΜ σε Ονοματεπώνυμο', 10 * '-')
            print(errors)
            print(72 * '-')
            return False

        return True

    def remove_xtra_spaces(self, text):
        while "  " in text:
            text = text.replace("  ", " ")

        return text.strip()

    def renameAFMtoName(self, only_name=True, starts_with_afm=False):
        inputDir = self.inputDir.get()
        outputDir = self.outputDir.get()
        errors = ''

        filesCounter = 0
        errorsCounter = 0
        for path, dirs, files in os.walk(inputDir):
            for file in files:
                if '.' in file:
                    dotPos = file.rindex('.')
                    afm = file[:dotPos]
                    ext = file[dotPos:]

                    if not self.verify_afm(afm):
                        errorsCounter += 1
                        errors += f'{errorsCounter:3}) Το αρχείο "{file}" δεν έχει σωστό ΑΦΜ... Θα αγνοηθεί.\n'
                        continue

                    if afm in self.afmToName:
                        filesCounter += 1
                        subdir = path.replace(inputDir, "").lstrip("\\")
                        destinationDir = os.path.join(outputDir, subdir)

                        if not os.path.exists(destinationDir):
                            os.makedirs(destinationDir)

                        if only_name:
                            target_name = self.afmToName[afm] + ext
                        else:
                            if starts_with_afm:
                                target_name = f'[{afm}] ' + self.afmToName[afm] + ext
                            else:
                                target_name = self.afmToName[afm] + f' [{afm}]' + ext
                        print(f'{filesCounter:3}: Μετονομασία του "{file}" σε "{target_name}"')
                        copyfile(os.path.join(path, file), os.path.join(destinationDir, target_name))
                    else:
                        errorsCounter += 1
                        errors += f'{errorsCounter:3}) Δεν βρέθηκε το ΑΦΜ "{afm}" στο αρχείο αντιστοίχισης.".\n'

        if errors != "":
            print("\n" + "-" * 30 + " Λάθη Μετονομασίας " + "-" * 30)
            print(errors)
            print(80 * '-')

    def renameNameToAFM(self, only_afm=True, starts_with_afm=False):
        inputDir = self.inputDir.get()
        outputDir = self.outputDir.get()
        errors = ''

        filesCounter = 0
        errorsCounter = 0
        for path, dirs, files in os.walk(inputDir):
            for file in files:
                if '.' in file:
                    dotPos = file.rindex('.')
                    name = file[:dotPos]
                    ext = file[dotPos:]

                    cleanName = self.remove_xtra_spaces(name)

                    if name != cleanName:
                        errorsCounter += 1
                        errors += f'{errorsCounter:3}) Το αρχείο "{file}" περιέχει περισσότερους κενούς χαρακτήρες από ότι χρειάζεται... Θα αναζητηθεί ως "{cleanName}"'
                        name = cleanName

                    if name in self.nameToAFM:
                        filesCounter += 1
                        subdir = path.replace(inputDir, "").lstrip("\\")
                        destinationDir = os.path.join(outputDir, subdir)

                        if not os.path.exists(destinationDir):
                            os.makedirs(destinationDir)

                        if only_afm:
                            target_name = self.nameToAFM[name] + ext
                        else:
                            if starts_with_afm:
                                target_name = f'[{self.nameToAFM[name]}] ' + name + ext
                            else:
                                target_name = name + f' [{self.nameToAFM[name]}]' + ext

                        print(f'{filesCounter:3}: Μετονομασία του "{file}" σε "{target_name}"')
                        copyfile(os.path.join(path, file), os.path.join(destinationDir, target_name))
                    else:
                        errorsCounter += 1
                        errors += f'{errorsCounter:3}) Δεν βρέθηκε το Ονοματεπώνυμο "{name}" στο αρχείο αντιστοίχισης.\n'

        if errors != "":
            print("\n" + "-" * 30 + " Λάθη Μετονομασίας " + "-" * 30)
            print(errors)
            print(80 * '-')

    def run(self):
        self.btnOpenXlsxFile.configure(state='disabled')
        self.cbAction.configure(state='disabled')
        self.btnOpenInputDir.configure(state='disabled')
        self.btnOpenOutputDir.configure(state='disabled')
        self.btnRun.configure(state='disabled')

        if self.cbAction.current() == 0:
            self.renameAFMtoName()
        elif self.cbAction.current() == 1:
            self.renameNameToAFM()
        elif self.cbAction.current() == 2:
            self.renameAFMtoName(only_name=False, starts_with_afm=True)
        elif self.cbAction.current() == 3:
            self.renameAFMtoName(only_name=False, starts_with_afm=False)
        elif self.cbAction.current() == 4:
            self.renameNameToAFM(only_afm=False, starts_with_afm=True)
        else:
            self.renameNameToAFM(only_afm=False, starts_with_afm=False)

        showinfo(title='Ολοκλήρωση Εκτέλεσης',
                 message='Η μετονομασία των αρχείων ολοκληρώθηκε.')

    def create_widgets(self):
        self.fData = Frame(self.window)

        self.lXlsxFile = Label(self.fData, text="Αρχείο αντιστοίχισης (xlsx)\nΑΦΜ σε Ονοματεπώνυμο:")
        self.lXlsxFile.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.xlsxFilename = StringVar()
        self.xlsxFilename.set('')
        self.ntrXlsxFilename = Entry(self.fData, width=128, state='readonly', textvariable=self.xlsxFilename)
        self.ntrXlsxFilename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenXlsxFile = Button(self.fData, text="Επιλέξτε αρχείο...", command=self.getXlsxFilename)
        self.btnOpenXlsxFile.grid(column=2, row=0, padx=10, pady=10)

        self.lSelection = Label(self.fData, text="Επιλογή:")
        self.lSelection.grid(column=0, row=1, padx=10, pady=10, sticky=E)

        self.action = StringVar()
        self.cbAction = Combobox(self.fData, width=125, textvariable=self.action, state='disabled')
        self.cbAction['values'] = ['ΑΦΜ --> Ονοματεπώνυμο',
                                   'Ονοματεπώνυμο --> ΑΦΜ',
                                   'ΑΦΜ --> [ΑΦΜ] Ονοματεπώνυμο',
                                   'ΑΦΜ --> Ονοματεπώνυμο [ΑΦΜ]',
                                   'Ονοματεπώνυμο --> [ΑΦΜ] Ονοματεπώνυμο',
                                   'Ονοματεπώνυμο --> Ονοματεπώνυμο [ΑΦΜ]'
                                   ]
        self.cbAction.bind("<<ComboboxSelected>>", self.cbActionSelect)
        self.cbAction.grid(column=1, row=1, padx=10, pady=10, sticky='NSEW')

        self.lInputDir = Label(self.fData, text="Φάκελος εισόδου:")
        self.lInputDir.grid(column=0, row=2, padx=10, pady=10, sticky=E)

        self.inputDir = StringVar()
        self.inputDir.set('')
        self.ntrInputDir = Entry(self.fData, width=128, state='disabled', textvariable=self.inputDir)
        self.ntrInputDir.grid(column=1, row=2, padx=10, pady=10, sticky=W)

        self.btnOpenInputDir = Button(self.fData, text="Επιλέξτε φάκελο...", state='disabled', command=self.getInputDir)
        self.btnOpenInputDir.grid(column=2, row=2, padx=10, pady=10)

        self.lOutputDir = Label(self.fData, text="Φάκελος εξόδου:")
        self.lOutputDir.grid(column=0, row=3, padx=10, pady=10, sticky=E)

        self.outputDir = StringVar()
        self.outputDir.set('')
        self.ntrOutputDir = Entry(self.fData, width=128, state='disabled', textvariable=self.outputDir)
        self.ntrOutputDir.grid(column=1, row=3, padx=10, pady=10, sticky=W)

        self.btnOpenOutputDir = Button(self.fData, text="Επιλέξτε φάκελο...", state='disabled',
                                       command=self.getOutputDir)
        self.btnOpenOutputDir.grid(column=2, row=3, padx=10, pady=10)

        self.btnRun = Button(self.fData, text="Εκτέλεση", command=self.run, state='disabled')
        self.btnRun.grid(column=1, row=10, padx=10, pady=10)

        self.fData.pack()


gui = GUI()
gui.window.mainloop()
