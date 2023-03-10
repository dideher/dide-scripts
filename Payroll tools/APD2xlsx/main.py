from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class GUI():
    def __init__(self):
        self.window = Tk()

        self.window.title("Μετατροπή ενός αρχείου ΑΠΔ σε xlsx")
        self.window.resizable(False, False)
        self.create_widgets()

    def getInputFilename(self):
        fName = filedialog.askopenfilename(initialdir=".", title="Επιλέξτε το αρχείο txt",
                                           filetypes=(("txt files", "*.txt"), ("all files", "*.*")))

        if fName == "":
            return

        self.btnOpenInputFile.configure(state='disabled')
        self.inputFilename.set(fName)

        self.btnRun.configure(state='normal')

    def saveFile(self, data, outputFile):
        wb = Workbook()
        ws = wb.active

        for entry in data:
            ws.append(entry)

        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

        notSaved = True

        while notSaved:
            try:
                wb.save(outputFile)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message="Παρακαλώ κλείστε το αρχείο '{}' ώστε να ολοκληρωθεί η αποθήκευση.".format(
                                outputFile))
            else:
                notSaved = False

    def convert_apd_to_excel(self, file_name):
        apd_file = open(file_name, "r", encoding='cp1253')

        entry_type_1_lengths = (
            1, 2, 2, 8, 2, 2, 3, 50, 80, 30, 30, 10, 9, 50, 10, 5, 30, 2, 4, 2, 4, 8, 12, 12, 8, 8, 30)
        entry_type_1_labels = ["Τύπος Εγγραφής", "Πλήθος μέσων που προσκομίζονται", "Α/Α μέσου", "Όνομα Αρχείου",
                               "Έκδοση",
                               "Τύπος Δήλωσης", "Υποκατάστημα ΙΚΑ Υποβολής",
                               "Ονομασία Υποκαταστήματος ΙΚΑ", "Επωνυμία / Επώνυμο", "Όνομα", "Όνομα Πατρός", "Α.Μ.Ε.",
                               "Α.Φ.Μ.", "Οδός", "Αριθμός", "Ταχυδρομικός Κωδικός", "Πόλη", "Από μήνα",
                               "Από έτος", "Έως μήνα", "Έως έτος", "Σύνολο Ημερών Ασφάλισης", "Σύνολο Αποδοχών",
                               "Σύνολο Καταβλητέων Εισφορών", "Ημερομηνία υποβολής", "Ημερομηνία παύσης εργασιών",
                               "Κενά"]

        entry_type_2_lengths = (1, 9, 11, 50, 30, 30, 30, 8, 9)
        entry_type_2_labels = ["Τύπος Εγγραφής", "Αριθμός Μητρώου Ασφαλισμένου", "Α.Μ.Κ.Α.", "Επώνυμο Ασφαλισμένου",
                               " Όνομα Ασφαλισμένου", " Όνομα Πατρός Ασφαλισμένου", "Όνομα Μητρός Ασφαλισμένου",
                               "Ημερομηνία Γέννησης", "Α.Φ.Μ."]

        entry_type_3_lengths = (1, 4, 4, 1, 1, 1, 6, 2, 4, 2, 4, 8, 8, 3, 3, 10, 10, 10, 10, 11, 10, 5, 10, 11)
        entry_type_3_labels = ["Τύπος Εγγραφής", "Αριθμός Παραρτήματος", "Κ.Α.Δ.", "Πλήρες Ωράριο", "Όλες εργάσιμες",
                               "Κυριακές", "Κωδικός Ειδικότητας", "Ειδικές περιπτώσεις ασφάλισης", "Πακέτο Κάλυψης",
                               "Μισθολογική περίοδος - μήνας", "Μισθολογική περίοδος - έτος",
                               "Από Ημερομηνία απασχόλησης",
                               "Έως Ημερομηνία απασχόλησης", "Τύπος αποδοχών", "Ημέρες Ασφάλισης", "Ημερομίσθιο",
                               "Αποδοχές", "Εισφορές Ασφαλισμένου", "Εισφορές Εργοδότη", "Συνολικές Εισφορές",
                               "Επιδότηση ασφαλισμένου (ποσό)", "Επιδότηση εργοδότη (%)", "Επιδότηση εργοδότη (ποσό)",
                               "Καταβλητέες εισφορές"]

        dataframe_columns = entry_type_2_labels + entry_type_3_labels
        dataframe_data = []
        dataframe_row = []

        multiple_payments = False
        for entry in apd_file:
            if entry[0] == "1":
                field_start = 0
                dde_data = []
                for i in range(len(entry_type_1_lengths)):
                    field_end = field_start + entry_type_1_lengths[i]
                    dde_data.append(entry[field_start:field_end])
                    field_start = field_end
            elif entry[0] == "2":
                multiple_payments = False
                field_start = 0
                employee_data = []
                for i in range(len(entry_type_2_lengths)):
                    field_end = field_start + entry_type_2_lengths[i]
                    if entry_type_2_labels[i] == "Ημερομηνία Γέννησης":
                        date_str_1 = entry[field_start:field_end]
                        date_str_2 = f"{date_str_1[0:2]}/{date_str_1[2:4]}/{date_str_1[4:8]}"
                        employee_data.append(date_str_2)
                    else:
                        employee_data.append(entry[field_start:field_end].strip())
                    field_start = field_end

                dataframe_row = dataframe_row + employee_data
            elif entry[0] == "3":
                field_start = 0

                if multiple_payments == True:
                    dataframe_row = [] + employee_data

                payment_data = []
                for i in range(len(entry_type_3_lengths)):
                    field_end = field_start + entry_type_3_lengths[i]
                    if entry_type_3_labels[i] == "Αποδοχές" or entry_type_3_labels[i] == "Εισφορές Ασφαλισμένου" or \
                            entry_type_3_labels[i] == "Εισφορές Εργοδότη" or entry_type_3_labels[
                        i] == "Συνολικές Εισφορές" or entry_type_3_labels[i] == "Επιδότηση ασφαλισμένου (ποσό)" or \
                            entry_type_3_labels[i] == "Επιδότηση εργοδότη (%)" or entry_type_3_labels[
                        i] == "Επιδότηση εργοδότη (ποσό)" or entry_type_3_labels[i] == "Καταβλητέες εισφορές":
                        income_str_1 = entry[field_start:field_end]
                        income_str_2 = income_str_1[:-2] + "." + income_str_1[-2:]
                        income = float(income_str_2)
                        payment_data.append(income)
                    elif entry_type_3_labels[i] == "Από Ημερομηνία απασχόλησης" or entry_type_3_labels[
                        i] == "Έως Ημερομηνία απασχόλησης":
                        date_str_1 = entry[field_start:field_end]
                        if date_str_1[4:8]:
                            date_str_2 = f"{date_str_1[0:2]}/{date_str_1[2:4]}/{date_str_1[4:8]}"
                            payment_data.append(date_str_2)
                        else:
                            payment_data.append(entry[field_start:field_end])
                    else:
                        payment_data.append(entry[field_start:field_end])
                    field_start = field_end

                dataframe_row = dataframe_row + payment_data
                dataframe_data.append(dataframe_row)
                multiple_payments = True

                dataframe_row = []

        self.data = []
        self.data.append(dataframe_columns)
        self.data += dataframe_data

    def run(self):
        self.btnRun.configure(state='disabled')

        inputFile = self.inputFilename.get()
        outputFile = inputFile.replace(".txt", ".xlsx")

        self.convert_apd_to_excel(inputFile)
        self.saveFile(self.data, outputFile)

        showinfo(title='Ολοκλήρωση Εκτέλεσης',
                 message=f'Η μετατροπή του αρχείου ΑΠΔ σε xlsx ολοκληρώθηκε.')

    def create_widgets(self):
        self.fData = Frame(self.window)

        self.lInputFile = Label(self.fData, text="Αρχείο ΑΠΔ (txt):")
        self.lInputFile.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.inputFilename = StringVar()
        self.inputFilename.set('')
        self.ntrInputFilename = Entry(self.fData, width=128, state='readonly', textvariable=self.inputFilename)
        self.ntrInputFilename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenInputFile = Button(self.fData, text="Επιλέξτε αρχείο...", command=self.getInputFilename)
        self.btnOpenInputFile.grid(column=2, row=0, padx=10, pady=10)

        self.btnRun = Button(self.fData, text="Εκτέλεση", command=self.run, state='disabled')
        self.btnRun.grid(column=1, row=10, padx=10, pady=10)

        self.fData.pack()


gui = GUI()
gui.window.mainloop()
