from tkinter import *
from tkinter import filedialog
from tkinter.ttk import *
from tkinter.messagebox import showwarning, showinfo
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import csv
from datetime import date


def conform_data(data):
    for item in data:
        for i in range(18):
            if i == 6:
                if item[i] == '':
                    continue

                d1, m1, y1 = item[i].split("/")
                if len(d1) == 1:
                    d1 = '0' + d1
                if len(m1) == 1:
                    m1 = '0' + m1
                if len(y1) == 2:
                    y1 = '20' + y1

                item[i] = f"{d1}/{m1}/{y1}"
            else:
                item[i] = item[i].replace('=', '')
                item[i] = item[i].replace('"', '')


def set_cols_width(ws):
    column_widths = []
    for row in ws.iter_rows():
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], len(str(cell.value)))
            except IndexError:
                column_widths.append(len(str(cell.value)))

    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23


def parse_csv_data(input_file):
    data = list()

    try:
        with open(input_file, 'rt', encoding='utf-8-sig') as f:
            dialect = csv.Sniffer().sniff(f.readline())
            f.seek(0)

            reader = csv.reader(f, delimiter=dialect.delimiter, quotechar=dialect.quotechar)

            for row in reader:
                data.append(row)
    except:
        with open(input_file, 'rt') as f:
            dialect = csv.Sniffer().sniff(f.readline())
            f.seek(0)

            reader = csv.reader(f, delimiter=dialect.delimiter, quotechar=dialect.quotechar)

            for row in reader:
                data.append(row)

    return data


class GUI:
    def __init__(self):
        self.cws_set = set()
        self.data1 = list()
        self.data2 = list()
        self.common = list()
        self.only_in1 = list()
        self.only_in2 = list()

        self.window = Tk()

        self.window.title("Σύγκριση αρχείων απεργιών")
        self.window.resizable(False, False)
        self.create_widgets()

    def check_date(self, date_str):
        try:
            d1, m1, y1 = date_str.split("/")
            if len(d1) == 1:
                d1 = '0' + d1
            if len(m1) == 1:
                m1 = '0' + m1
            if len(y1) == 2:
                y1 = '20' + y1
            if len(y1) != 4:
                raise Exception

            date1 = date(int(y1), int(m1), int(d1))

            self.from_date.set(f"{d1}/{m1}/{y1}")
        except:
            showwarning("Προσοχή ...",
                        "Η σωστή μορφοποίηση για την ημερομηνία είναι 'Ημέρα/Μήνας/Έτος'.\n"
                        "Για τη διόρθωση των λαθών:\n"
                        "1) Η τιμή της 'Ημέρας' πρέπει να είναι μεταξύ 1 και 31.\n"
                        "2) Η τιμή της 'Ημέρας' σχετίζεται με την τιμή του 'Μήνα' "
                        "(π.χ. ο Φεβρουάριος σίγουρα δεν μπορεί να έχει 30 ή 31 ημέρες).\n"
                        "3) Η τιμή του 'Μήνα' πρέπει να είναι μεταξύ 1 και 12.\n"
                        "4) Η τιμή του 'Έτους' πρέπει να είναι μεγαλύτερη του 1.")
            return False

        return True

    def save_file(self, data, filename):
        wb = Workbook()
        ws = wb.active

        for row in data:
            ws.append(row)

        set_cols_width(ws)

        output_file = os.path.join(self.output_dir_name.get(), filename)

        not_saved = True

        while not_saved:
            try:
                wb.save(output_file)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message=f"Παρακαλώ κλείστε το αρχείο '{filename}' ώστε να ολοκληρωθεί η αποθήκευση.")
            else:
                not_saved = False

    def get_output_dir_name(self):
        d_name = filedialog.askdirectory(initialdir="./data/",
                                         title="Επιλέξτε τον φάκελο που θα αποθηκευτούν τα αρχεία")

        if d_name == "":
            return

        self.output_dir_name.set(d_name)
        self.ntr_output_dir_name.configure(state='disabled')
        self.btn_get_output_dir.configure(state='disabled')
        self.ntr_from_date.configure(state='normal')
        year = date.today().year
        self.from_date.set(f'01/09/{year}')
        self.update_cws_listbox()
        self.btn_run.configure(state='normal')

    def update_cws_set(self, data):
        for item in data[1:]:
            self.cws_set.add(item[14])

    def update_cws_listbox(self):
        sorted_list = list()
        for item in self.cws_set:
            sorted_list.append(item)

        sorted_list.sort()
        for item in sorted_list:
            self.lb_cws.insert(END, item)

    def get_data1_filename(self):
        f_name = filedialog.askopenfilename(initialdir="./data/",
                                            title="Επιλέξτε το πρώτο αρχείο άδειας",
                                            filetypes=(("csv files", "*.csv"), ("all files", "*.*")))

        if f_name == "":
            return

        self.data1_filename.set(f_name)
        self.ntr_data1_filename.configure(state='disabled')
        self.btn_get_data1.configure(state='disabled')
        self.data1 = parse_csv_data(f_name)
        self.update_cws_set(self.data1)
        self.ntr_data2_filename.configure(state='readonly')
        self.btn_get_data2.configure(state='normal')

    def get_data2_filename(self):
        f_name = filedialog.askopenfilename(initialdir="./data/",
                                            title="Επιλέξτε το πρώτο αρχείο άδειας",
                                            filetypes=(("csv files", "*.csv"), ("all files", "*.*")))

        if f_name == "":
            return

        if f_name == self.data1_filename.get() or f_name == self.data1_filename.get()[:-1]:
            showwarning(title="Προσοχή ...", message="Έχετε ήδη επιλέξει αυτό το αρχείο ως πρώτο αρχείο.")
            return

        self.data2_filename.set(f_name)
        self.ntr_data2_filename.configure(state='disabled')
        self.btn_get_data2.configure(state='disabled')
        self.data2 = parse_csv_data(f_name)
        self.update_cws_set(self.data2)
        self.ntr_output_dir_name.configure(state='readonly')
        self.btn_get_output_dir.configure(state='normal')

    def filter_data(self, data):
        filtered_data = list()

        d1, m1, y1 = self.from_date.get().split("/")
        date1 = date(int(y1), int(m1), int(d1))

        cws_selections = self.lb_cws.curselection()
        cws_selections_text = list()

        for item in cws_selections:
            cws_selections_text.append(self.lb_cws.get(item))

        for item in data[1:]:
            date2_str = item[6]
            cws = item[14]

            d2, m2, y2 = date2_str.split("/")
            date2 = date(int(y2), int(m2), int(d2))

            if date2 > date1 and cws in cws_selections_text:
                filtered_data.append(item)

        return filtered_data

    def create_lists(self):
        for item in self.filtered_data1:
            if item in self.filtered_data2:
                self.common.append(item)
            else:
                self.only_in1.append(item)

        for item in self.filtered_data2:
            if item not in self.filtered_data1:
                self.only_in2.append(item)

    def export_lists(self):
        header = self.data1[0]

        output_dir = self.output_dir_name.get()

        if len(self.common) != 0:
            save_list = list()
            save_list.append(header)
            save_list += self.common
            self.save_file(save_list, os.path.join(output_dir, 'common.xlsx'))

        if len(self.only_in1) != 0:
            save_list = list()
            save_list.append(header)
            save_list += self.only_in1
            self.save_file(save_list, os.path.join(output_dir, 'onlyIn1stFile.xlsx'))

        if len(self.only_in2) != 0:
            save_list = list()
            save_list.append(header)
            save_list += self.only_in2
            self.save_file(save_list, os.path.join(output_dir, 'onlyIn2ndFile.xlsx'))

        if len(self.only_in1) == 0 and len(self.only_in2) == 0:
            showinfo(title="Απουσία αρχείων διαφορών",
                     message="Δεν υπήρχαν διαφορές μεταξύ των αρχείων για το φιλτράρισμα που επιλέξατε.")

    def create_diff_report(self):
        not_found_in_1st = ''
        not_found_in_2nd = ''
        header = ['Περιφέρεια',
                  'Διεύθυνση',
                  'Κωδικός Φορέα Υπηρέτησης',
                  'Ονομασία Φορέα Υπηρέτησης',
                  'Τύπος (Απεργία ή Στάση Εργασίας)',
                  'Ώρες Στάσης Εργασίας',
                  'Ημερομηνία Απεργίας/Στάσης Εργασίας',
                  'A.M.',
                  'A.Φ.M.',
                  'Φύλο',
                  'Επώνυμο',
                  'Όνομα',
                  'Πατρώνυμο',
                  'Κωδικός Κύριας Ειδικότητας',
                  'Σχέση Εργασίας',
                  'ΦΕΚ/ΑΔΑ Διορισμού/Πρόσληψης',
                  'Θέση Προσωπικού Φακέλου',
                  'Κωδικός Οργανικής/Προσωρινής Τοποθέτησης',
                  'Ονομασία Οργανικής/Προσωρινής Τοποθέτησης']

        print(20 * '-', 'Analysis', 20 * '-')
        for item1 in self.only_in1:
            last_name1 = item1[10]
            first_name1 = item1[11]
            father_name1 = item1[12]
            specialty1 = item1[13]
            type1 = item1[4]
            from_date1 = item1[6]

            found = False
            for item2 in self.only_in2:
                last_name2 = item2[10]
                first_name2 = item2[11]
                father_name2 = item2[12]
                specialty2 = item2[13]
                type2 = item2[4]
                from_date2 = item2[6]

                if last_name1 == last_name2 and \
                        first_name1 == first_name2 and \
                        father_name1 == father_name2 and \
                        specialty1 == specialty2 and \
                        type1 == type2 and \
                        from_date1 == from_date2:
                    found = True
                    diffs = ''
                    for i in range(len(header)):
                        if item1[i] != item2[i]:
                            diffs += f"{header[i]}: '{item1[i]}' <--> '{item2[i]}'\n"

                    if diffs != '':
                        print(f"{last_name1} {first_name1} {father_name1} {specialty1} {type1} {from_date1}:")
                        print(diffs)

            if not found:
                not_found_in_2nd += f"{last_name1} {first_name1} {father_name1} {specialty1} {type1} {from_date1}\n"

        if not_found_in_2nd != '':
            print(20 * '-', 'Not found in 2nd file', 20 * '-')
            print(not_found_in_2nd)

        for item2 in self.only_in2:
            last_name2 = item2[10]
            first_name2 = item2[11]
            father_name2 = item2[12]
            specialty2 = item2[13]
            type2 = item2[4]
            from_date2 = item2[6]

            found = False
            for item1 in self.only_in1:
                last_name1 = item1[10]
                first_name1 = item1[11]
                father_name1 = item1[12]
                specialty1 = item1[13]
                type1 = item1[4]
                from_date1 = item1[6]

                if last_name1 == last_name2 and \
                        first_name1 == first_name2 and \
                        father_name1 == father_name2 and \
                        specialty1 == specialty2 and \
                        type1 == type2 and \
                        from_date1 == from_date2:
                    found = True

            if not found:
                not_found_in_1st += f"{last_name2} {first_name2} {father_name2} {specialty2} {type2} {from_date2}\n"

        if not_found_in_1st != '':
            print(20 * '-', 'Not found in 1st file', 20 * '-')
            print(not_found_in_1st)

        if not_found_in_1st == '' and not_found_in_2nd == '':
            print('Filtered entries were the same.')

    def run(self):
        if not self.check_date(self.from_date.get()):
            return

        if len(self.lb_cws.curselection()) == 0:
            showwarning("Προσοχή ...", "Πρέπει να επιλέξετε τουλάχιστον μια 'Τρέχουσα Σχέση Εργασίας'.")

            return

        self.ntr_from_date.configure(state='disabled')
        self.lb_cws.configure(state='disabled')
        self.btn_run.configure(state='disabled')

        self.filtered_data1 = self.filter_data(self.data1)
        self.filtered_data2 = self.filter_data(self.data2)
        conform_data(self.filtered_data1)
        conform_data(self.filtered_data2)
        self.create_lists()
        self.create_diff_report()
        self.export_lists()

        showinfo(title="Ολοκλήρωση εκτέλεσης", message="Η σύγκριση των αρχείων αδειών ολοκληρώθηκε.")

    def create_widgets(self):
        self.f_data = Frame(self.window)

        self.l_data1 = Label(self.f_data, text="Πρώτο αρχείο:")
        self.l_data1.grid(column=0, row=0, padx=10, pady=5, sticky=E)

        self.data1_filename = StringVar()
        self.ntr_data1_filename = Entry(self.f_data, width=128, state='readonly', textvariable=self.data1_filename)
        self.ntr_data1_filename.grid(column=1, row=0, padx=10, pady=5, sticky=W)

        self.btn_get_data1 = Button(self.f_data, text="Επιλέξτε αρχείο...", command=self.get_data1_filename)
        self.btn_get_data1.grid(column=2, row=0, padx=10, pady=5)

        self.l_data2 = Label(self.f_data, text="Δεύτερο αρχείο:")
        self.l_data2.grid(column=0, row=1, padx=10, pady=5, sticky=E)

        self.data2_filename = StringVar()
        self.ntr_data2_filename = Entry(self.f_data, width=128, state='disabled', textvariable=self.data2_filename)
        self.ntr_data2_filename.grid(column=1, row=1, padx=10, pady=5, sticky=W)

        self.btn_get_data2 = Button(self.f_data, text="Επιλέξτε αρχείο...", command=self.get_data2_filename,
                                    state='disabled')
        self.btn_get_data2.grid(column=2, row=1, padx=10, pady=5)

        self.l_output_dir_name = Label(self.f_data, text="Φάκελος για αποθήκευση:")
        self.l_output_dir_name.grid(column=0, row=2, padx=10, pady=5, sticky=E)

        self.output_dir_name = StringVar()
        self.ntr_output_dir_name = Entry(self.f_data, width=128, state='disabled', textvariable=self.output_dir_name)
        self.ntr_output_dir_name.grid(column=1, row=2, padx=10, pady=5, sticky=W)

        self.btn_get_output_dir = Button(self.f_data, text="Επιλέξτε φάκελο...", command=self.get_output_dir_name,
                                         state='disabled')
        self.btn_get_output_dir.grid(column=2, row=2, padx=10, pady=5)

        self.lf_filter = LabelFrame(self.f_data, text="Φιλτράρισμα")
        self.lf_filter.grid(column=0, row=3, columnspan=3, padx=10, pady=10, sticky=EW)

        self.l_from_date = Label(self.lf_filter, text="Από:")
        self.l_from_date.grid(column=0, row=0, padx=10, pady=5, sticky=E)

        self.from_date = StringVar()
        self.ntr_from_date = Entry(self.lf_filter, width=125, state='disabled', textvariable=self.from_date)
        self.ntr_from_date.grid(column=1, row=0, pady=5)

        self.l_current_work_status = Label(self.lf_filter, text="Τρέχουσα Σχέση Εργασίας:")
        self.l_current_work_status.grid(column=0, row=1, padx=10, pady=5, sticky=NE)

        sb_listbox_cws = Scrollbar(self.lf_filter, orient=VERTICAL)
        sb_listbox_cws.grid(row=1, column=2, sticky=N + S + W)

        self.lb_cws = Listbox(self.lf_filter, selectmode='multiple', yscrollcommand=sb_listbox_cws.set)
        self.lb_cws.grid(column=1, row=1, pady=5, sticky=E + W)

        sb_listbox_cws.configure(command=self.lb_cws.yview)

        self.btn_run = Button(self.f_data, text="Εκτέλεση", command=self.run, state='disabled')
        self.btn_run.grid(column=1, row=10, padx=10, pady=5)

        self.f_data.pack()


gui = GUI()
gui.window.mainloop()
