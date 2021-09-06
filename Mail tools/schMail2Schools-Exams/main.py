from tkinter import *
from tkinter import filedialog
from tkinter.scrolledtext import ScrolledText
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from openpyxl import load_workbook
from os import walk
from threading import Thread
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.header import Header
from email.utils import formataddr, formatdate
import os
import smtplib
import ssl


class GUI:
    def __init__(self):
        self.window = Tk()
        self.window.bind_all("<Key>", self._onKeyRelease, "+")
        self.window.title("Αποστολή μαζικής αλληλογραφίας σε Σχολικές Μονάδες")
        self.window.resizable(False, False)
        self.create_widgets()

    def _onKeyRelease(self, event):
        ctrl = (event.state & 0x4) != 0
        if event.keycode == 88 and ctrl and event.keysym.lower() != "x":
            event.widget.event_generate("<<Cut>>")

        if event.keycode == 86 and ctrl and event.keysym.lower() != "v":
            event.widget.event_generate("<<Paste>>")

        if event.keycode == 67 and ctrl and event.keysym.lower() != "c":
            event.widget.event_generate("<<Copy>>")

    def debugCkbChange(self):
        if self.debugChecked.get():
            self.ntrDebugRecipient.configure(state='normal')
        else:
            self.ntrDebugRecipient.configure(state='disabled')

    def send_mail(self, receiver_email, subject, body, attachments):
        smtp_username = self.user.get()
        smtp_pass = self.passwd.get()
        sender_email = self.sch_mail.get()

        # Create a multipart message and set headers
        message = MIMEMultipart()
        message["From"] = formataddr(
            (str(Header('Διεύθυνση Δευτεροβάθμιας Εκπαίδευσης Ηρακλείου', 'utf-8')), sender_email))
        message["To"] = receiver_email
        message["Subject"] = subject
        message["Bcc"] = sender_email
        message["Date"] = formatdate(localtime=True)

        # Add body to email
        message.attach(MIMEText(body, "plain"))

        for filename in attachments:
            # Open attachemnt file in binary mode
            with open(filename, "rb") as attachment:
                # Add file as application/octet-stream
                # Email client can usually download this automatically as attachment
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())

            # Encode file in ASCII characters to send by email
            encoders.encode_base64(part)

            # Add header as key/value pair to attachment part
            fname = os.path.basename(filename)
            part.add_header('Content-Disposition', 'attachment', filename=fname)

            # Add attachment to message and convert message to string
            message.attach(part)

        text = message.as_string()

        # Log in to server using secure context and send email
        context = ssl.create_default_context()
        with smtplib.SMTP("mail.sch.gr", 587) as server:
            server.starttls(context=context)
            server.login(smtp_username, smtp_pass)
            server.sendmail(sender_email, [receiver_email, sender_email], text)

    def parseXlsxData(self, test):
        workbook = load_workbook(filename=self.addressBookFilename.get())
        sheet = workbook.active

        self.schools = list()

        for row in sheet.iter_rows():
            entry = list()
            for cell in row:
                if cell.value is None:
                    text = ""
                else:
                    text = str(cell.value).strip()

                entry.append(text)

            self.schools.append(entry)

        inputDirectory = self.filesDirName.get()
        debugMode = self.debugChecked.get()

        self.stPreview.delete('1.0', END)

        previewText = ""
        if test:
            if debugMode:
                previewText += "-" * 20 + " Δοκιμαστική λειτουργία " + "-" * 20 + "\n"
                previewText += "Αποστολή στον παραλήπτη: " + self.debugRecipient.get() + "\n"
                previewText += 64 * "-" + "\n\n"
            previewText += "-" * 20 + " Θέμα " + "-" * 20 + "\n"
            previewText += self.subject.get() + "\n\n"
            previewText += "-" * 20 + " Σώμα " + "-" * 20 + "\n"
            previewText += self.stBody.get('1.0', END) + "\n\n"
            previewText += "-" * 20 + " Κοινό συνημμένο " + "-" * 20 + "\n"
            if self.attachmentFilename.get() == "":
                previewText += "Χωρίς κοινό συνημμένο." + "\n\n"
            else:
                previewText += self.attachmentFilename.get() + "\n\n"
            previewText += "-" * 20 + " Βιλίο διευθύνσεων " + "-" * 20 + "\n"
            previewText += self.addressBookFilename.get() + "\n\n"
            previewText += "-" * 20 + " Φάκελος αρχείων " + "-" * 20 + "\n"
            previewText += self.filesDirName.get() + "\n\n"
            previewText += "-" * 20 + " Αποστολή ... " + "-" * 20 + "\n\n"

        excludedSchools = ""
        errors = ""
        exceptions = ""

        if not test:
            showinfo(title="Έναρξη αποστολής",
                     message="Μην τερματίσετε την εφαρμογή μέχρι να εμφανιστεί το μήνημα ολοκλήρωσης της αποστολής.")

        filesCounter = 0
        sendCounter = 0
        for sch in self.schools[1:]:
            schoolName = sch[0]
            schoolCode = sch[1]
            excluded = sch[3]

            if not debugMode:
                recipient_email = sch[2]
            else:
                recipient_email = self.debugRecipient.get()

            if not excluded:
                attachmentsList = list()
                appendCommonAttachment = True

                for path, dirs, files in walk(inputDirectory):
                    for file in files:
                        if schoolCode in file:
                            filesCounter += 1

                            if self.attachmentFilename.get() != "" and appendCommonAttachment:
                                attachmentsList.append(self.attachmentFilename.get())
                                appendCommonAttachment = False

                            attachmentsList.append(path + "\\" + file)

                if len(attachmentsList) == 0:
                    errors += "To σχολείο '{} ({})' δεν έχει αρχεία για αποστολή.\n".format(schoolName, schoolCode)
                else:
                    if test:
                        previewText += "-" * 20 + " {} ({}) ".format(schoolName, schoolCode) + "-" * 20

                        if len(attachmentsList) == 1:
                            previewText += "\nΑποστολή του παρκάτω αρχείου στον παραλήπτη '{}':\n".format(
                                recipient_email)
                            previewText += "1: '{}'\n".format(attachmentsList[0])
                        else:
                            previewText += "\nΑποστολή των παρακάτω αρχείων στον παραλήπτη '{}':\n".format(
                                recipient_email)
                            i = 0
                            skipAttachment = True
                            for a in attachmentsList:
                                if self.attachmentFilename.get() != "" and skipAttachment:
                                    previewText += "Κοινό συνημμένο: '{}'\n".format(a)
                                    skipAttachment = False
                                else:
                                    i += 1
                                    previewText += "{}: '{}'\n".format(i, a)
                    else:
                        try:
                            self.send_mail(receiver_email=recipient_email, subject=self.subject.get(),
                                           body=self.stBody.get('1.0', END), attachments=attachmentsList)
                        except:
                            exceptions += "Η αποστολή στον παραλήπτη '{}' απέτυχε.\n".format(recipient_email)
                        else:
                            sendCounter += 1
                            self.stPreview.insert(INSERT, "{:2d}) Η αποστολή στον παραλήπτη '{}: {}' πέτυχε.\n".format(
                                sendCounter, schoolName, recipient_email))
            else:
                excludedSchools += "Για το σχολείο '{}' υπήρχε εξαίρεση.\n".format(schoolName)

        if test:
            previewText += "\n" + 64 * "-" + "\n"
            previewText += "Πλήθος αρχείων: {} \n".format(filesCounter)
            previewText += 64 * "-" + "\n\n"

            if excludedSchools != "":
                previewText += "\n" + "-" * 20 + " Εξαιρέσεις " + "-" * 20 + "\n"
                previewText += excludedSchools

            if errors != "":
                previewText += "\n" + "-" * 20 + " Λάθη " + "-" * 20 + "\n"
                previewText += errors
        else:
            if exceptions != "":
                previewText += "\n" + "-" * 20 + " Αποτυχημένες αποστολές " + "-" * 20 + "\n"
                previewText += exceptions

        self.stPreview.insert(INSERT, previewText)

        if not test:
            showinfo(title="Ολοκλήρωση αποστολής", message="Η αποστολή ολοκληρώθηκε.")

    def getAttachmentFilename(self):
        fName = filedialog.askopenfilename(initialdir="./data/", title="Επιλέξτε το έγγραφο",
                                           filetypes=(("pdf files", "*.pdf"), ("all files", "*.*")))

        if fName == "":
            return

        self.attachmentFilename.set(fName)

    def nextGoToTabSettings(self):
        if self.subject.get() == "" and self.stBody.get('1.0', END) == "\n":
            showwarning(title="Κενά πεδία...", message="Παρακαλώ συμπληρώστε τα πεδία 'Θέμα' και 'Σώμα' του μηνύματος.")
            return
        elif self.subject.get() == "":
            showwarning(title="Κενό πεδίο...", message="Παρακαλώ συμπληρώστε το πεδίο 'Θέμα' του μηνύματος.")
            return
        elif self.stBody.get('1.0', END) == "\n":
            showwarning(title="Κενό πεδίο...", message="Παρακαλώ συμπληρώστε το πεδίο 'Σώμα' του μηνύματος.")
            return

        self.tabControl.tab(0, state='disabled')
        self.tabControl.tab(1, state='normal')
        self.tabControl.select(1)

    def nextGoToTabPreview(self):
        if self.addressBookFilename.get() == "" or self.filesDirName.get() == "" or self.sch_mail.get() == "" or self.user.get() == "" or self.passwd.get() == "":
            showwarning(title="Κενά πεδία...", message="Παρακαλώ συμπληρώστε όλα τα πεδία για να προχωρήσετε.")
            return

        if self.debugChecked.get() and self.debugRecipient.get() == "":
            showwarning(title="Κενό πεδίο...",
                        message="Έχετε ενεργοποιήσει τη δοκιμαστική λειτουργία. Για να προχωρήσετε πρέπει υποχρεωτικά να συμπληρώσετε το πεδίο 'Παραλήπτης για δοκιμαστική αποστολή'.")
            return

        self.tabControl.tab(1, state='disabled')
        self.tabControl.tab(2, state='normal')
        self.tabControl.select(2)

    def getAddressBookFilename(self):
        fName = filedialog.askopenfilename(initialdir="./mail/", title="Επιλέξτε το Βιβλίο Διευθύνσεων",
                                           filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if fName == "":
            return

        self.addressBookFilename.set(fName)

    def getFilesDirName(self):
        dName = filedialog.askdirectory(initialdir="./data/", title="Επιλέξτε τον φάκελο με τα αρχεία προς αποστολή")

        if dName == "":
            return

        self.filesDirName.set(dName)

    def prevGoToTabSubjectBody(self):
        self.tabControl.tab(1, state='disabled')
        self.tabControl.tab(0, state='normal')
        self.tabControl.select(0)

    def prevGoToTabSettings(self):
        self.stPreview.delete('1.0', END)
        self.notPreviewed = True
        self.btnSend.configure(state='disabled')

        self.tabControl.tab(2, state='disabled')
        self.tabControl.tab(1, state='normal')
        self.tabControl.select(1)

    def preview(self):
        self.notPreviewed = False
        self.btnSend.configure(state='normal')
        self.parseXlsxData(test=True)

    def send(self):
        self.btnPrevGoToSettings.configure(state='disabled')
        self.btnPreview.configure(state='disabled')
        self.btnSend.configure(state='disabled')

        self.runThread = Thread(target=self.parseXlsxData, args=[False])
        self.runThread.start()

    def create_widgets(self):
        # Tabs
        self.tabControl = Notebook(self.window)
        self.tabSubjectBody = Frame(self.tabControl)
        self.tabControl.add(self.tabSubjectBody, text="Δημιουργία μηνύματος")
        self.tabSettings = Frame(self.tabControl)
        self.tabControl.add(self.tabSettings, text="Επιλογές/Ρυθμίσεις", state='disabled')
        self.tabPreview = Frame(self.tabControl)
        self.tabControl.add(self.tabPreview, text="Προεπισκόπηση/Αποστολή", state='disabled')
        self.tabControl.pack(expand=1, fill="both")

        # Tab: SubjectBody
        self.fSubjectBody = Frame(self.tabSubjectBody)

        self.lSubject = Label(self.fSubjectBody, text="Θέμα:")
        self.lSubject.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.subject = StringVar()
        self.ntrSubject = Entry(self.fSubjectBody, width=110, textvariable=self.subject)
        self.ntrSubject.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.lBody = Label(self.fSubjectBody, text="Σώμα:")
        self.lBody.grid(column=0, row=1, padx=10, pady=10, sticky=NE)

        self.stBody = ScrolledText(self.fSubjectBody, height=10, wrap=WORD)
        self.stBody.grid(column=1, row=1, padx=10, pady=10)

        self.lAttachment = Label(self.fSubjectBody, text="Κοινό συνημμένο αρχείο\n(π.χ. διαβιβαστικό):")
        self.lAttachment.grid(column=0, row=2, padx=10, pady=10, sticky=E)

        self.attachmentFilename = StringVar()
        self.ntrAttachmentFilename = Entry(self.fSubjectBody, width=110, state='readonly',
                                           textvariable=self.attachmentFilename)
        self.ntrAttachmentFilename.grid(column=1, row=2, padx=10, pady=10, sticky=W)

        self.btnOpenAttachment = Button(self.fSubjectBody, text="Επιλέξτε αρχείο...",
                                        command=self.getAttachmentFilename)
        self.btnOpenAttachment.grid(column=2, row=2, padx=10, pady=10)

        self.btnNextGoToSettings = Button(self.fSubjectBody, text="Επόμενο", command=self.nextGoToTabSettings)
        self.btnNextGoToSettings.grid(column=2, row=10, padx=10, pady=10)

        self.fSubjectBody.pack()

        # Tab: Settings
        self.fSettings = Frame(self.tabSettings)

        self.lAddressBook = Label(self.fSettings, text="Βιβλίο Διευθύνσεων:")
        self.lAddressBook.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.addressBookFilename = StringVar()
        self.ntrAddressBookFilename = Entry(self.fSettings, width=110, state='readonly',
                                            textvariable=self.addressBookFilename)
        self.ntrAddressBookFilename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenAddressBook = Button(self.fSettings, text="Επιλέξτε αρχείο...", command=self.getAddressBookFilename)
        self.btnOpenAddressBook.grid(column=2, row=0, padx=10, pady=10)

        self.lFiles = Label(self.fSettings, text="Φάκελος με αρχεία προς αποστολή:")
        self.lFiles.grid(column=0, row=1, padx=10, pady=10, sticky=E)

        self.filesDirName = StringVar()
        self.ntrFilesDirName = Entry(self.fSettings, width=110, state='readonly', textvariable=self.filesDirName)
        self.ntrFilesDirName.grid(column=1, row=1, padx=10, pady=10, sticky=W)

        self.btnOpenFilesDirName = Button(self.fSettings, text="Επιλέξτε φάκελο...", command=self.getFilesDirName)
        self.btnOpenFilesDirName.grid(column=2, row=1, padx=10, pady=10)

        self.lSchMail = Label(self.fSettings, text="Sch mail:")
        self.lSchMail.grid(column=0, row=2, padx=10, pady=10, sticky=E)

        self.sch_mail = StringVar()
        self.ntrSchMail = Entry(self.fSettings, width=110, textvariable=self.sch_mail)
        self.ntrSchMail.grid(column=1, row=2, padx=10, pady=10, sticky=W)

        self.lUser = Label(self.fSettings, text="User:")
        self.lUser.grid(column=0, row=3, padx=10, pady=10, sticky=E)

        self.user = StringVar()
        self.ntrUser = Entry(self.fSettings, width=110, textvariable=self.user)
        self.ntrUser.grid(column=1, row=3, padx=10, pady=10, sticky=W)

        self.lPasswd = Label(self.fSettings, text="Password:")
        self.lPasswd.grid(column=0, row=4, padx=10, pady=10, sticky=E)

        self.passwd = StringVar()
        self.ntrPasswd = Entry(self.fSettings, width=110, textvariable=self.passwd)
        self.ntrPasswd.grid(column=1, row=4, padx=10, pady=10, sticky=W)

        self.lfDebug = LabelFrame(self.fSettings, text="Δοκιμαστική λειτουργία")
        self.lfDebug.grid(column=0, row=5, columnspan=3, padx=10, pady=10)

        self.lDebugRecipient = Label(self.lfDebug, text="Παραλήπτης για δοκιμαστική\nαποστολή:")
        self.lDebugRecipient.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.debugRecipient = StringVar()
        self.ntrDebugRecipient = Entry(self.lfDebug, width=110, textvariable=self.debugRecipient, state='disabled')
        self.ntrDebugRecipient.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.debugChecked = BooleanVar()
        self.ckbDebugChecked = Checkbutton(self.lfDebug, text="Ενεργοποίηση δοκιμαστικής\nλειτουργίας",
                                           variable=self.debugChecked, command=self.debugCkbChange)
        self.ckbDebugChecked.grid(column=2, row=0, padx=10, pady=10)

        self.btnPrevGoToSubjectBody = Button(self.fSettings, text="Προηγούμενο", command=self.prevGoToTabSubjectBody)
        self.btnPrevGoToSubjectBody.grid(column=0, row=10, padx=10, pady=10)

        self.btnNextGoToPreview = Button(self.fSettings, text="Επόμενο", command=self.nextGoToTabPreview)
        self.btnNextGoToPreview.grid(column=2, row=10, padx=10, pady=10)

        self.fSettings.pack()

        # Tab: Preview
        self.fPreview = Frame(self.tabPreview)

        self.stPreview = ScrolledText(self.fPreview, width=120, height=16)
        self.stPreview.grid(column=0, row=0, columnspan=3, padx=10, pady=10)

        self.btnPrevGoToSettings = Button(self.fPreview, text="Προηγούμενο", command=self.prevGoToTabSettings)
        self.btnPrevGoToSettings.grid(column=0, row=10, padx=10, pady=10)

        self.btnPreview = Button(self.fPreview, text="Προεπισκόπηση", command=self.preview)
        self.btnPreview.grid(column=1, row=10, padx=10, pady=10)

        self.notPreviewed = True

        self.btnSend = Button(self.fPreview, text="Αποστολή", command=self.send, state='disabled')
        self.btnSend.grid(column=2, row=10, padx=10, pady=10)

        self.fPreview.pack()


gui = GUI()
gui.window.mainloop()
