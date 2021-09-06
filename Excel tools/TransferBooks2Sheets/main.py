from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from openpyxl import *
from openpyxl.utils import get_column_letter
from os import walk


class GUI():
    def __init__(self):
        self.window = Tk()

        self.window.title("Συνένωση αρχείων excel σε ένα")
        self.window.resizable(False, False)
        self.create_widgets()


    def getFilesDirName(self):
        dName = filedialog.askdirectory(initialdir="./data/", title="Επιλέξτε τον φάκελο με τα αρχεία προς συνένωση")

        if dName == "":
            return

        self.filesDirName.set(dName)
        self.btnRun.configure(state='normal')


    def run(self):
        inputDirectory = self.filesDirName.get()

        out_wb = Workbook()

        files_count = 0
        for path, dirs, files in walk(inputDirectory):
            for file in files:
                if file[-5:] == ".xlsx":
                    files_count += 1

        self.pbProgress['maximum'] = files_count
        self.pbProgress['value'] = 0
        self.pbProgress.update()

        count = 0
        for path, dirs, files in walk(inputDirectory):
            for file in files:
                if file[-5:] != ".xlsx":
                    continue

                self.lProgress.configure(text=f'Εκτέλεση σε εξέλιξη... ({count + 1}/{files_count})')
                self.pbProgress['value'] = count + 1
                self.pbProgress.update()

                out_ws = out_wb.create_sheet(
                    file.replace("Κενό - ", "Κ-").replace("Πλεόνασμα - ", "Π-").replace(" - ", "-").
                        replace(".xlsx", "")[0:30])

                workbook = load_workbook(filename=path + "\\" + file)
                sheet = workbook.active

                mr = sheet.max_row
                mc = sheet.max_column

                # copying the cell values from source
                # excel file to destination excel file
                for i in range(1, mr + 1):
                    for j in range(1, mc + 1):
                        # reading cell value from source excel file
                        c = sheet.cell(row=i, column=j)

                        # writing the read value to destination excel file
                        out_ws.cell(row=i, column=j).value = c.value

                column_widths = []
                for row in out_ws.iter_rows():
                    for i, cell in enumerate(row):
                        try:
                            column_widths[i] = max(column_widths[i], len(str(cell.value)))
                        except IndexError:
                            column_widths.append(len(str(cell.value)))

                for i, column_width in enumerate(column_widths):
                    out_ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

                count += 1

        sn = out_wb.sheetnames
        out_wb.remove(out_wb[sn[0]])
        outputFile = "output.xlsx"

        notSaved = True

        while notSaved:
            try:
                out_wb.save(outputFile)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message="Παρακαλώ κλείστε το αρχείο '{}' ώστε να ολοκληρωθεί η αποθήκευση.".format(
                                outputFile))
            else:
                notSaved = False

        showinfo(title="Αρχείο εξόδου",
                 message=f'Η συνένωση έχει αποθηκευτεί στο αρχείο "{outputFile}" στον φάκελο εκτέλεσης του προγράμματος.')


    def create_widgets(self):
        self.fData = Frame(self.window)

        self.lData = Label(self.fData, text="Φάκελος με αρχεία\nπρος συνένωση:")
        self.lData.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.filesDirName = StringVar()
        self.ntrFilesDirName = Entry(self.fData, width=100, state='readonly', textvariable=self.filesDirName)
        self.ntrFilesDirName.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenData = Button(self.fData, text="Επιλέξτε φάκελο...", command=self.getFilesDirName)
        self.btnOpenData.grid(column=2, row=0, padx=10, pady=10)

        self.lProgress = Label(self.fData, text="Αναμονή για εκτέλεση ...")
        self.lProgress.grid(column=0, row=1, columnspan=3, padx=10, pady=10)

        self.pbProgress = Progressbar(self.fData, orient='horizontal', length=400, mode='determinate')
        self.pbProgress.grid(column=0, row=2, columnspan=3, padx=10, pady=10)

        self.btnRun = Button(self.fData, text="Εκτέλεση συνένωσης", command=self.run, state='disabled')
        self.btnRun.grid(column=1, row=10, padx=10, pady=10)

        self.fData.pack()


gui = GUI()
gui.window.mainloop()
