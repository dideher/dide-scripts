from tkinter import *
from tkinter import filedialog
from tkinter.ttk import *
from tkinter.messagebox import showwarning, showinfo
from openpyxl import *
from sortedcontainers import SortedSet
import os


class GUI():
    def __init__(self):
        self.window = Tk()

        self.window.title("Μετασχηματισμός αρχείου με Κενά/Πλεονάσματα")
        self.window.resizable(False, False)
        self.create_widgets()

    def is_float(self, s):
        return s.replace('.', '', 1).isdecimal()

    def parseXlsxData(self):
        workbook = load_workbook(filename=self.dataFilename.get())
        sheet = workbook.active

        self.xlsxData = list()

        for row in sheet.iter_rows():
            entry = list()
            for i, cell in enumerate(row):
                if i > 4:
                    break

                if cell.value is None:
                    text = ""
                else:
                    text = str(cell.value).strip()

                if self.is_float(text):
                    text = text.replace(".", ",")

                entry.append(text)

            self.xlsxData.append(entry)

    def createSpecialtiesTypes(self):
        self.generalEducationSpcTypes = SortedSet()
        self.specialEducationSpcTypes = SortedSet()
        self.miscSpcTypes = SortedSet()

        for item in self.xlsxData[1:]:
            if item[3] == 'Γενικής Παιδείας - Πανελλαδικώς Εξεταζόμενα Μαθήματα':
                self.generalEducationSpcTypes.add(f'{item[1]} - Γενικής Παιδείας - Πανελλαδικώς Εξεταζόμενα Μαθήματα')
                self.generalEducationSpcTypes.add(
                    f'{item[1]} - Γενικής Παιδείας - μη Πανελλαδικώς Εξεταζόμενα Μαθήματα')
                self.generalEducationSpcTypes.add(f'{item[1]} - Γενικής Παιδείας (Σύνολο)')
            elif item[3] == 'Γενικής Παιδείας - μη Πανελλαδικώς Εξεταζόμενα Μαθήματα':
                self.generalEducationSpcTypes.add(
                    f'{item[1]} - Γενικής Παιδείας - μη Πανελλαδικώς Εξεταζόμενα Μαθήματα')
            elif 'Ειδικής Αγωγής' in item[3]:
                self.specialEducationSpcTypes.add(f'{item[1]} - {item[3]}')
            else:
                self.miscSpcTypes.add(f'{item[1]} - {item[3]}')

    def getSchools(self):
        self.generalEducationSchools = SortedSet()
        self.specialEducationSchools = SortedSet()
        self.miscSchools = SortedSet()

        for item in self.xlsxData[1:]:
            if item[3] == 'Γενικής Παιδείας - Πανελλαδικώς Εξεταζόμενα Μαθήματα' or \
                    item[3] == 'Γενικής Παιδείας - μη Πανελλαδικώς Εξεταζόμενα Μαθήματα':
                self.generalEducationSchools.add(item[0])
            elif 'Ειδικής Αγωγής' in item[3]:
                self.specialEducationSchools.add(item[0])
            else:
                self.miscSchools.add(item[0])

    def createTables(self):
        self.createGEStable()
        self.createSEStable()
        self.createMStable()

    def createMStable(self):
        self.msTable = list()
        header = list()
        header.append('Σχολείο')
        header += self.miscSpcTypes[:]

        self.msTable.append(header)
        for sch in self.miscSchools:
            entry = list()
            entry.append(sch)
            sch_values = [0] * len(self.miscSpcTypes)

            for item in self.xlsxData[1:]:
                if item[0] != sch:
                    continue

                if item[3] == 'Γενικής Παιδείας - Πανελλαδικώς Εξεταζόμενα Μαθήματα' or \
                        item[3] == 'Γενικής Παιδείας - μη Πανελλαδικώς Εξεταζόμενα Μαθήματα' or \
                        'Ειδικής Αγωγής' in item[3]:
                    continue

                spcType = f'{item[1]} - {item[3]}'
                indexScpType = self.miscSpcTypes.index(spcType)

                if item[2] == 'Κενό':
                    sch_values[indexScpType] -= int(item[4])
                else:
                    sch_values[indexScpType] += int(item[4])

            entry += sch_values

            self.msTable.append(entry)

    def createSEStable(self):
        self.sesTable = list()
        header = list()
        header.append('Σχολείο')
        header += self.specialEducationSpcTypes[:]

        self.sesTable.append(header)
        for sch in self.specialEducationSchools:
            entry = list()
            entry.append(sch)
            sch_values = [0] * len(self.specialEducationSpcTypes)

            for item in self.xlsxData[1:]:
                if item[0] != sch:
                    continue

                if 'Ειδικής Αγωγής' in item[3]:
                    spcType = f'{item[1]} - {item[3]}'
                    indexScpType = self.specialEducationSpcTypes.index(spcType)

                    if item[2] == 'Κενό':
                        sch_values[indexScpType] -= int(item[4])
                    else:
                        sch_values[indexScpType] += int(item[4])
                else:
                    continue

            entry += sch_values

            self.sesTable.append(entry)

    def createGEStable(self):
        self.gesTable = list()
        header = list()
        header.append('Σχολείο')
        header += self.generalEducationSpcTypes[:]

        self.gesTable.append(header)
        for sch in self.generalEducationSchools:
            entry = list()
            entry.append(sch)
            sch_values = [0] * len(self.generalEducationSpcTypes)

            for item in self.xlsxData[1:]:
                if item[0] != sch:
                    continue

                if item[3] == 'Γενικής Παιδείας - Πανελλαδικώς Εξεταζόμενα Μαθήματα' or \
                        item[3] == 'Γενικής Παιδείας - μη Πανελλαδικώς Εξεταζόμενα Μαθήματα':
                    spcType = f'{item[1]} - {item[3]}'
                    indexScpType = self.generalEducationSpcTypes.index(spcType)

                    if item[2] == 'Κενό':
                        sch_values[indexScpType] -= int(item[4])
                    else:
                        sch_values[indexScpType] += int(item[4])
                else:
                    continue

                spcType = f'{item[1]} - Γενικής Παιδείας (Σύνολο)'
                if spcType in self.generalEducationSpcTypes:
                    indexScpType = self.generalEducationSpcTypes.index(spcType)
                    if item[2] == 'Κενό':
                        sch_values[indexScpType] -= int(item[4])
                    else:
                        sch_values[indexScpType] += int(item[4])

            entry += sch_values

            self.gesTable.append(entry)

    def saveFiles(self):
        self.saveFile('Γενικής Παιδείας.xlsx', self.gesTable)
        self.saveFile('Ειδικής Αγωγής.xlsx', self.sesTable)
        self.saveFile('Υπόλοιπα.xlsx', self.msTable)

    def saveFile(self, outFile, data):
        wb = Workbook()
        ws = wb.active

        for row in data:
            ws.append(row)

        notSaved = True

        outputFile = os.path.join(self.outputDirName.get(), outFile)
        while notSaved:
            try:
                wb.save(outputFile)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message=f"Παρακαλώ κλείστε το αρχείο '{outputFile}' ώστε να ολοκληρωθεί η αποθήκευση.")
            else:
                notSaved = False

    def getOutputDirName(self):
        dName = filedialog.askdirectory(initialdir="./data/", title="Επιλέξτε τον φάκελο που θα αποθηκευτούν τα αρχεία")

        if dName == "":
            return

        self.outputDirName.set(dName)
        self.btnRun.configure(state='normal')

    def getDataFilename(self):
        fName = filedialog.askopenfilename(initialdir="./data/",
                                           title="Επιλέξτε το αρχείο excel",
                                           filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if fName == "":
            return

        self.dataFilename.set(fName)
        self.parseXlsxData()
        self.ntrOutputDirName.configure(state='readonly')
        self.btnOpenOutputDir.configure(state='normal')

    def run(self):
        self.btnRun.configure(state='disabled')
        self.parseXlsxData()
        self.getSchools()
        self.createSpecialtiesTypes()
        self.createTables()
        self.saveFiles()
        showinfo(title="Ολοκλήρωση εκτέλεσης", message="Ο μετασχηματισμός ολοκληρώθηκε.")

    def create_widgets(self):
        self.fData = Frame(self.window)

        self.lData = Label(self.fData, text="Αρχείο:")
        self.lData.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.dataFilename = StringVar()
        self.ntrDataFilename = Entry(self.fData, width=128, state='readonly', textvariable=self.dataFilename)
        self.ntrDataFilename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenData = Button(self.fData, text="Επιλέξτε αρχείο...", command=self.getDataFilename)
        self.btnOpenData.grid(column=2, row=0, padx=10, pady=10)

        self.lOutputDirName = Label(self.fData, text="Φάκελος για αποθήκευση των αρχείων:")
        self.lOutputDirName.grid(column=0, row=2, padx=10, pady=10, sticky=E)

        self.outputDirName = StringVar()
        self.ntrOutputDirName = Entry(self.fData, width=128, state='disabled', textvariable=self.outputDirName)
        self.ntrOutputDirName.grid(column=1, row=2, padx=10, pady=10, sticky=W)

        self.btnOpenOutputDir = Button(self.fData, text="Επιλέξτε φάκελο...", command=self.getOutputDirName,
                                       state='disabled')
        self.btnOpenOutputDir.grid(column=2, row=2, padx=10, pady=10)

        self.btnRun = Button(self.fData, text="Εκτέλεση μετασχηματισμού", command=self.run, state='disabled')
        self.btnRun.grid(column=1, row=10, padx=10, pady=10)

        self.fData.pack()


gui = GUI()
gui.window.mainloop()
