from tkinter import *
from tkinter import filedialog
from tkinter.ttk import *
from tkinter.messagebox import showwarning, showinfo
from openpyxl import *
from openpyxl.utils import get_column_letter
import os


class GUI():
    def __init__(self):
        self.window = Tk()

        self.window.title("Διαχωρισμός φύλλου excel σε πολλά (Αντί για filter/select/copy/new file/paste/save)")
        self.window.resizable(False, False)
        self.create_widgets()


    def is_float(self, s):
        return s.replace('.', '', 1).isdecimal()


    def parseXlsxData(self):
        workbook = load_workbook(filename=self.dataFilename.get())
        sheet = workbook.active

        self.xlsxData = list()

        for row in sheet.iter_rows():
            entry = list()
            for cell in row:
                if cell.value is None:
                    text = ""
                else:
                    text = str(cell.value).strip()

                if self.is_float(text):
                    text = text.replace(".", ",")

                entry.append(text)

            self.xlsxData.append(entry)


    def splitData(self):
        data = dict()
        skipHeader = True

        for r in self.xlsxData:
            if skipHeader:
                skipHeader = False
                continue

            temp = r.copy()

            key = r[self.cbFilterCols.current()]

            if key == "":
                key = "_EMPTY CELLS_"

            if key not in data:
                data[key] = list()
            data[key].append(temp)

        logb = Workbook()
        logs = logb.active
        logh = ["Σχολείο", "Πλήθος μαθητών"]
        logs.append(logh)

        for key in data:
            logd = [key, len(data[key])]
            logs.append(logd)

        notSaved = True

        outFile = "_log.xlsx"
        outputFile = os.path.join(self.outputDirName.get(), outFile)
        while notSaved:
            try:
                logb.save(outputFile)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message="Παρακαλώ κλείστε το αρχείο '{}' ώστε να ολοκληρωθεί η αποθήκευση.".format(
                                outputFile))
            else:
                notSaved = False

        for key in data:
            wb = Workbook()
            ws = wb.active

            ws.append(self.xlsxData[0])
            for row in data[key]:
                ws.append(row)

            column_widths = []
            for row in ws.iter_rows():
                for i, cell in enumerate(row):
                    try:
                        column_widths[i] = max(column_widths[i], len(str(cell.value)))
                    except IndexError:
                        column_widths.append(len(str(cell.value)))

            for i, column_width in enumerate(column_widths):
                ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

            outFile = (key.replace("<", "_").replace(">", "_").replace(":", "_").replace("\"", "_").replace("/", "_")
                       .replace("\\", "_").replace("|", "_").replace("?", "_").replace("*", "_"))
            outFile += ".xlsx"

            notSaved = True

            outputFile = os.path.join(self.outputDirName.get(), outFile)
            while notSaved:
                try:
                    wb.save(outputFile)
                except:
                    showwarning(title="Αρχείο σε χρήση...",
                                message="Παρακαλώ κλείστε το αρχείο '{}' ώστε να ολοκληρωθεί η αποθήκευση.".format(
                                    outputFile))
                else:
                    notSaved = False


    def getOutputDirName(self):
        dName = filedialog.askdirectory(initialdir="./data/", title="Επιλέξτε τον φάκελο που θα αποθηκευτούν τα αρχεία")

        if dName == "":
            return

        self.outputDirName.set(dName)
        self.btnRun.configure(state='normal')


    def getDataFilename(self):
        fName = filedialog.askopenfilename(initialdir="./data/",
                                           title="Επιλέξτε το αρχείο excel",
                                           filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if fName == "":
            return

        self.dataFilename.set(fName)
        self.parseXlsxData()
        self.cbFilterCols.configure(state='readonly')
        self.cbFilterCols['values'] = self.xlsxData[0]


    def cbFilterColsSelect(self, eventObject):
        if self.cbFilterCols.current() != -1:
            self.ntrOutputDirName.configure(state='readonly')
            self.btnOpenOutputDir.configure(state='normal')


    def run(self):
        self.splitData()
        showinfo(title="Ολοκλήρωση εκτέλεσης", message="Ο διαχωρισμός ολοκληρώθηκε.")


    def create_widgets(self):
        self.fData = Frame(self.window)

        self.lData = Label(self.fData, text="Αρχείο:")
        self.lData.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.dataFilename = StringVar()
        self.ntrDataFilename = Entry(self.fData, width=128, state='readonly', textvariable=self.dataFilename)
        self.ntrDataFilename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenData = Button(self.fData, text="Επιλέξτε αρχείο...", command=self.getDataFilename)
        self.btnOpenData.grid(column=2, row=0, padx=10, pady=10)

        self.lFilterCol = Label(self.fData, text="Στήλη για διαχωρισμό:")
        self.lFilterCol.grid(column=0, row=1, padx=10, pady=10, sticky=E)

        self.filterCols = StringVar()
        self.cbFilterCols = Combobox(self.fData, width=40, textvariable=self.filterCols, state='disabled')
        self.cbFilterCols.bind("<<ComboboxSelected>>", self.cbFilterColsSelect)
        self.cbFilterCols.grid(column=1, row=1, padx=10, pady=10, sticky='NSEW')

        self.lOutputDirName = Label(self.fData, text="Φάκελος για αποθήκευση των αρχείων:")
        self.lOutputDirName.grid(column=0, row=2, padx=10, pady=10, sticky=E)

        self.outputDirName = StringVar()
        self.ntrOutputDirName = Entry(self.fData, width=128, state='disabled', textvariable=self.outputDirName)
        self.ntrOutputDirName.grid(column=1, row=2, padx=10, pady=10, sticky=W)

        self.btnOpenOutputDir = Button(self.fData, text="Επιλέξτε φάκελο...", command=self.getOutputDirName, state='disabled')
        self.btnOpenOutputDir.grid(column=2, row=2, padx=10, pady=10)

        self.btnRun = Button(self.fData, text="Εκτέλεση διαχωρισμού", command=self.run, state='disabled')
        self.btnRun.grid(column=1, row=10, padx=10, pady=10)

        self.fData.pack()


gui = GUI()
gui.window.mainloop()
