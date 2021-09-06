from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from openpyxl import *
from openpyxl.utils import get_column_letter
from os import walk


class GUI():
    def __init__(self):
        self.window = Tk()

        self.window.title("Συνένωση αρχείων excel σε ένα")
        self.window.resizable(False, False)
        self.create_widgets()

    def getFilesDirName(self):
        dName = filedialog.askdirectory(initialdir="./data/", title="Επιλέξτε τον φάκελο με τα αρχεία προς συνένωση")

        if dName == "":
            return

        self.filesDirName.set(dName)
        self.btnRun.configure(state='normal')

    def is_float(self, s):
        return s.replace('.', '', 1).isdecimal()

    def clean_spaces(self, text):
        while "  " in text:
            text = text.replace("  ", " ")

        return text

    def run(self):
        self.btnRun.configure(state='disabled')
        inputDirectory = self.filesDirName.get()

        header = list()
        data = list()

        for path, dirs, files in walk(inputDirectory):
            for file in files:
                if file[-5:] == ".xlsx":
                    workbook = load_workbook(filename=path + "\\" + file)
                    sheet = workbook.active

                    for i, row in enumerate(sheet.iter_rows()):
                        entry = list()
                        for cell in row:
                            if cell.value is None:
                                text = ""
                            else:
                                text = self.clean_spaces(str(cell.value))

                            if self.is_float(text):
                                text = text.replace(".", ",")
                            entry.append(text)

                        if i == 0:
                            header.clear()
                            header.append(entry)
                        else:
                            data.append(entry)

        wb = Workbook()
        ws = wb.active

        ws.append(header[0])

        for entry in data:
            ws.append(entry)

        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

        outputFile = "output.xlsx"

        notSaved = True

        while notSaved:
            try:
                wb.save(outputFile)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message="Παρακαλώ κλείστε το αρχείο '{}' ώστε να ολοκληρωθεί η αποθήκευση.".format(
                                outputFile))
            else:
                notSaved = False

        showinfo(title="Αρχείο εξόδου",
                 message=f'Η συνένωση έχει αποθηκευτεί στο αρχείο "{outputFile}" στον φάκελο εκτέλεσης του προγράμματος.')
        self.btnRun.configure(state='normal')

    def create_widgets(self):
        self.fData = Frame(self.window)

        self.lData = Label(self.fData, text="Φάκελος με αρχεία\nπρος συνένωση:")
        self.lData.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.filesDirName = StringVar()
        self.ntrFilesDirName = Entry(self.fData, width=100, state='readonly', textvariable=self.filesDirName)
        self.ntrFilesDirName.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenData = Button(self.fData, text="Επιλέξτε φάκελο...", command=self.getFilesDirName)
        self.btnOpenData.grid(column=2, row=0, padx=10, pady=10)

        self.btnRun = Button(self.fData, text="Εκτέλεση συνένωσης", command=self.run, state='disabled')
        self.btnRun.grid(column=1, row=10, padx=10, pady=10)

        self.fData.pack()


gui = GUI()
gui.window.mainloop()
