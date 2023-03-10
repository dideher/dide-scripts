from tkinter import *
from tkinter import filedialog
from tkinter.ttk import *
from tkinter.messagebox import showwarning, showinfo
from openpyxl import *
from openpyxl.utils import get_column_letter
import os


class GUI:
    def __init__(self):
        self.window = Tk()

        self.window.title("Διαχωρισμός αρχείου xlsx σε πακέτα")
        self.window.resizable(False, False)
        self.create_widgets()

    def is_float(self, s):
        return s.replace('.', '', 1).isdecimal()

    def parse_xlsx_data(self):
        wb = load_workbook(filename=self.data_filename.get())
        sheet = wb.active

        self.data = list()

        for row in sheet.iter_rows():
            entry = list()
            for cell in row:
                if cell.value is None:
                    text = ""
                else:
                    text = str(cell.value).strip()

                # if self.is_float(text):
                #     text = text.replace(".", ",")

                entry.append(text)

            self.data.append(entry)

    def check_value(self, value):
        value = value.strip()
        entries_count = int(self.entries_count.get())

        if value.isdigit():
            value_num = int(value) - 1
            if 0 <= value_num < entries_count:
                return True

        return False

    def split_data(self):
        packet_size = int(self.packet_size.get())
        counter = 1
        packet_num = 1

        self.split_list = list()
        self.split_list.append(list())

        for row in self.data[1:]:
            self.split_list[packet_num - 1].append(row)
            if counter < packet_size:
                counter += 1
            else:
                counter = 1
                packet_num += 1
                self.split_list.append(list())

    def export_parts(self):
        header = self.data[0]

        for pn, part in enumerate(self.split_list):
            wb = Workbook()
            ws = wb.active

            ws.append(header)
            for row in part:
                ws.append(row)

            self.set_cols_width(ws)
            self.safe_save(wb, f"part {pn + 1:03}.xlsx")

    def set_cols_width(self, ws):
        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

    def safe_save(self, wb, out_file):
        not_saved = True

        output_file = os.path.join(self.output_dir.get(), out_file)
        while not_saved:
            try:
                wb.save(output_file)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message=f"Παρακαλώ κλείστε το αρχείο '{output_file}' ώστε να ολοκληρωθεί η αποθήκευση.")
            else:
                not_saved = False

    def get_output_dir(self):
        d_name = filedialog.askdirectory(initialdir="./data/",
                                         title="Επιλέξτε τον φάκελο που θα αποθηκευτούν τα αρχεία")

        if d_name == "":
            return

        self.output_dir.set(d_name)
        self.btn_run.configure(state='normal')

    def get_data(self):
        f_name = filedialog.askopenfilename(initialdir="./data/",
                                            title="Επιλέξτε το αρχείο xlsx",
                                            filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if f_name == "":
            return

        self.data_filename.set(f_name)
        self.parse_xlsx_data()
        self.ntr_entries_count.configure(state='readonly')
        self.entries_count.set(str(len(self.data) - 1))
        self.ntr_packet_size.configure(state='normal')
        self.packet_size.set(str(round(len(self.data) / 2)))
        self.ntr_output_dir.configure(state='readonly')
        self.btn_get_output_dir.configure(state='normal')

    def run(self):
        if not self.check_value(self.packet_size.get()):
            showwarning(title='Λάθος τιμή πεδίου',
                        message='Η τιμή του μεγέθους των πακέτων πρέπει να είναι αριθμός μεγαλύτερος του 0 και '
                                f'μικρότερος ή ίσος από {self.entries_count.get()}.')
            return

        self.split_data()
        self.export_parts()

        showinfo(title="Ολοκλήρωση εκτέλεσης", message="Ο διαχωρισμός ολοκληρώθηκε.")

    def create_widgets(self):
        self.f_data = Frame(self.window)

        self.l_data = Label(self.f_data, text="Αρχείο:")
        self.l_data.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.data_filename = StringVar()
        self.ntr_data_filename = Entry(self.f_data, width=128, state='readonly', textvariable=self.data_filename)
        self.ntr_data_filename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btn_get_data = Button(self.f_data, text="Επιλέξτε αρχείο...", command=self.get_data)
        self.btn_get_data.grid(column=2, row=0, padx=10, pady=10)

        self.l_entries_count = Label(self.f_data, text="Πλήθος εγγραφών αρχείου:")
        self.l_entries_count.grid(column=0, row=1, padx=10, pady=10, sticky=E)

        self.entries_count = StringVar()
        self.ntr_entries_count = Entry(self.f_data, width=128, state='disabled', textvariable=self.entries_count)
        self.ntr_entries_count.grid(column=1, row=1, padx=10, pady=10, sticky=W)

        self.l_packet_size = Label(self.f_data, text="Πλήθος εγγραφών ανά πακέτο:")
        self.l_packet_size.grid(column=0, row=2, padx=10, pady=10, sticky=E)

        self.packet_size = StringVar()
        self.ntr_packet_size = Entry(self.f_data, width=128, state='disabled', textvariable=self.packet_size)
        self.ntr_packet_size.grid(column=1, row=2, padx=10, pady=10, sticky=W)

        self.l_output_dir = Label(self.f_data, text="Φάκελος για αποθήκευση των αρχείων:")
        self.l_output_dir.grid(column=0, row=3, padx=10, pady=10, sticky=E)

        self.output_dir = StringVar()
        self.ntr_output_dir = Entry(self.f_data, width=128, state='disabled', textvariable=self.output_dir)
        self.ntr_output_dir.grid(column=1, row=3, padx=10, pady=10, sticky=W)

        self.btn_get_output_dir = Button(self.f_data, text="Επιλέξτε φάκελο...", command=self.get_output_dir,
                                         state='disabled')
        self.btn_get_output_dir.grid(column=2, row=3, padx=10, pady=10)

        self.btn_run = Button(self.f_data, text="Εκτέλεση διαχωρισμού", command=self.run, state='disabled')
        self.btn_run.grid(column=1, row=10, padx=10, pady=10)

        self.f_data.pack()


gui = GUI()
gui.window.mainloop()
