from tkinter import *
from tkinter import filedialog
from tkinter.ttk import *
from tkinter.messagebox import showwarning, showinfo
from openpyxl import *
from openpyxl.utils import get_column_letter
import os, re


class GUI():
    def __init__(self):
        self.data1 = list()
        self.data2 = list()
        self.data3 = list()
        self.data4 = list()
        self.file1Selections = [-1, -1, -1, -1]
        self.file2Selections = [-1, -1, -1, -1]

        self.window = Tk()

        self.window.title("Διαχωρισμός αρχείου excel με πολλαπλές στήλες")
        self.window.resizable(False, False)
        self.create_widgets()

    def conformText(self, text):
        text1 = (text.upper().replace("Ά", "Α").replace("Έ", "Ε").replace("Ή", "Η").replace("Ί", "Ι")
                 .replace("Ϊ́", "Ϊ").replace("Ύ", "Υ").replace("Ϋ́", "Ϋ").replace("Ό", "Ο").replace("Ώ", "Ω"))

        text2 = re.sub(r'([ ]+)', r' ', text1)

        return text2

    def is_float(self, s):
        return s.replace('.', '', 1).isdecimal()

    def parseXlsxData(self, file, data):
        workbook = load_workbook(filename=file)
        sheet = workbook.active

        for row in sheet.iter_rows():
            entry = list()
            for cell in row:
                if cell.value is None:
                    text = ""
                else:
                    text = str(cell.value).strip()

                if self.is_float(text):
                    text = text.replace(".", ",")

                entry.append(text)

            data.append(entry)

    def setColsWidth(self, ws):
        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

    def saveFileMatches(self):
        wb = Workbook()
        ws = wb.active

        for row in self.data3:
            ws.append(row)

        self.setColsWidth(ws)

        outputFile = os.path.join(self.outputDirName.get(), "output_matches.xlsx")

        notSaved = True

        while notSaved:
            try:
                wb.save(outputFile)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message=f"Παρακαλώ κλείστε το αρχείο '{outputFile}' ώστε να ολοκληρωθεί η αποθήκευση.")
            else:
                notSaved = False

    def saveFileNoMatches(self):
        wb = Workbook()
        ws = wb.active

        for row in self.data4:
            ws.append(row)

        self.setColsWidth(ws)

        outputFile = os.path.join(self.outputDirName.get(), "output_no_matches.xlsx")

        notSaved = True

        while notSaved:
            try:
                wb.save(outputFile)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message=f"Παρακαλώ κλείστε το αρχείο '{outputFile}' ώστε να ολοκληρωθεί η αποθήκευση.")
            else:
                notSaved = False

    def getOutputDirName(self):
        dName = filedialog.askdirectory(initialdir="./data/", title="Επιλέξτε τον φάκελο που θα αποθηκευτεί το αρχείο")

        if dName == "":
            return

        self.outputDirName.set(dName)
        self.ntrOutputDirName.configure(state='disabled')
        self.btnOpenOutputDir.configure(state='disabled')
        self.cbFile1_1.configure(state='readonly')
        self.cbFile2_1.configure(state='readonly')
        self.btnAddRelation2.configure(state='normal')
        self.btnRun.configure(state='normal')

    def getData1Filename(self):
        fName = filedialog.askopenfilename(initialdir="./data/",
                                           title="Επιλέξτε το πρώτο αρχείο excel",
                                           filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if fName == "":
            return

        self.data1Filename.set(fName)
        self.ntrData1Filename.configure(state='disabled')
        self.btnOpenData1.configure(state='disabled')
        self.parseXlsxData(fName, self.data1)
        self.ntrData2Filename.configure(state='readonly')
        self.btnOpenData2.configure(state='normal')

        self.cbFile1_1['values'] = self.data1[0]
        self.cbFile1_2['values'] = self.data1[0]
        self.cbFile1_3['values'] = self.data1[0]
        self.cbFile1_4['values'] = self.data1[0]

    def getData2Filename(self):
        fName = filedialog.askopenfilename(initialdir="./data/",
                                           title="Επιλέξτε το δεύτερο αρχείο excel",
                                           filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if fName == "":
            return

        if fName == self.data1Filename.get():
            showwarning(title="Προσοχή ...", message="Έχετε ήδη επιλέξει αυτό το αρχείο ως πρώτο αρχείο.")
            return

        self.data2Filename.set(fName)
        self.ntrData2Filename.configure(state='disabled')
        self.btnOpenData2.configure(state='disabled')
        self.parseXlsxData(fName, self.data2)
        self.ntrOutputDirName.configure(state='readonly')
        self.btnOpenOutputDir.configure(state='normal')

        self.cbFile2_1['values'] = self.data2[0]
        self.cbFile2_2['values'] = self.data2[0]
        self.cbFile2_3['values'] = self.data2[0]
        self.cbFile2_4['values'] = self.data2[0]

    def addRelation2(self):
        if (self.cbFile1_1.current() == -1 or self.cbFile2_1.current() == -1):
            showwarning(title="Προσοχή ...", message="Πρέπει να κάνετε την αντιστοίχιση πριν προσθέσετε νέα.")
            return

        self.cbFile1_2.configure(state='readonly')
        self.cbFile2_2.configure(state='readonly')
        self.btnAddRelation2.configure(state='disabled')
        self.btnAddRelation3.configure(state='normal')
        self.btnRemoveRelation2.configure(state='normal')

    def removeRelation2(self):
        self.file1Selections[2 - 1] = -1
        self.file2Selections[2 - 1] = -1
        self.file1_2.set('')
        self.cbFile1_2.configure(state='disabled')
        self.file2_2.set('')
        self.cbFile2_2.configure(state='disabled')
        self.btnRemoveRelation2.configure(state='disabled')
        self.btnAddRelation3.configure(state='disabled')
        self.btnAddRelation2.configure(state='normal')

    def addRelation3(self):
        if (self.cbFile1_2.current() == -1 or self.cbFile2_2.current() == -1):
            showwarning(title="Προσοχή ...", message="Πρέπει να κάνετε την αντιστοίχιση πριν προσθέσετε νέα.")
            return

        self.cbFile1_3.configure(state='readonly')
        self.cbFile2_3.configure(state='readonly')
        self.btnAddRelation3.configure(state='disabled')
        self.btnRemoveRelation2.configure(state='disabled')
        self.btnAddRelation4.configure(state='normal')
        self.btnRemoveRelation3.configure(state='normal')

    def removeRelation3(self):
        self.file1Selections[3 - 1] = -1
        self.file2Selections[3 - 1] = -1
        self.file1_3.set('')
        self.cbFile1_3.configure(state='disabled')
        self.file2_3.set('')
        self.cbFile2_3.configure(state='disabled')
        self.btnRemoveRelation3.configure(state='disabled')
        self.btnAddRelation4.configure(state='disabled')
        self.btnAddRelation3.configure(state='normal')
        self.btnRemoveRelation2.configure(state='normal')

    def addRelation4(self):
        if (self.cbFile1_3.current() == -1 or self.cbFile2_3.current() == -1):
            showwarning(title="Προσοχή ...", message="Πρέπει να κάνετε την αντιστοίχιση πριν προσθέσετε νέα.")
            return

        self.cbFile1_4.configure(state='readonly')
        self.cbFile2_4.configure(state='readonly')
        self.btnAddRelation4.configure(state='disabled')
        self.btnRemoveRelation3.configure(state='disabled')
        self.btnRemoveRelation4.configure(state='normal')

    def removeRelation4(self):
        self.file1Selections[4 - 1] = -1
        self.file2Selections[4 - 1] = -1
        self.file1_4.set('')
        self.cbFile1_4.configure(state='disabled')
        self.file2_4.set('')
        self.cbFile2_4.configure(state='disabled')
        self.btnRemoveRelation4.configure(state='disabled')
        self.btnAddRelation4.configure(state='normal')
        self.btnRemoveRelation3.configure(state='normal')

    def cbFile1_1Select(self, eventObject):
        if self.cbFile1_1.current() != self.file1Selections[1 - 1] and self.cbFile1_1.current() in self.file1Selections:
            showwarning(title="Προσοχή ...", message="Έχετε ήδη μια αντιστοίχιση με αυτό το πεδίο.")
            if self.file1Selections[1 - 1] != -1:
                self.cbFile1_1.current(self.file1Selections[1 - 1])
            else:
                self.file1_1.set('')
            return

        self.file1Selections[1 - 1] = self.cbFile1_1.current()

    def cbFile2_1Select(self, eventObject):
        if self.cbFile2_1.current() != self.file2Selections[1 - 1] and self.cbFile2_1.current() in self.file2Selections:
            showwarning(title="Προσοχή ...", message="Έχετε ήδη μια αντιστοίχιση με αυτό το πεδίο.")
            if self.file2Selections[1 - 1] != -1:
                self.cbFile2_1.current(self.file2Selections[1 - 1])
            else:
                self.file2_1.set('')
            return

        self.file2Selections[1 - 1] = self.cbFile2_1.current()

    def cbFile1_2Select(self, eventObject):
        if self.cbFile1_2.current() != self.file1Selections[2 - 1] and self.cbFile1_2.current() in self.file1Selections:
            showwarning(title="Προσοχή ...", message="Έχετε ήδη μια αντιστοίχιση με αυτό το πεδίο.")
            if self.file1Selections[2 - 1] != -1:
                self.cbFile1_2.current(self.file1Selections[2 - 1])
            else:
                self.file1_2.set('')
            return

        self.file1Selections[2 - 1] = self.cbFile1_2.current()

    def cbFile2_2Select(self, eventObject):
        if self.cbFile2_2.current() != self.file2Selections[2 - 1] and self.cbFile2_2.current() in self.file2Selections:
            showwarning(title="Προσοχή ...", message="Έχετε ήδη μια αντιστοίχιση με αυτό το πεδίο.")
            if self.file2Selections[2 - 1] != -1:
                self.cbFile2_2.current(self.file2Selections[2 - 1])
            else:
                self.file2_2.set('')
            return

        self.file2Selections[2 - 1] = self.cbFile2_2.current()

    def cbFile1_3Select(self, eventObject):
        if self.cbFile1_3.current() != self.file1Selections[3 - 1] and self.cbFile1_3.current() in self.file1Selections:
            showwarning(title="Προσοχή ...", message="Έχετε ήδη μια αντιστοίχιση με αυτό το πεδίο.")
            if self.file1Selections[3 - 1] != -1:
                self.cbFile1_3.current(self.file1Selections[3 - 1])
            else:
                self.file1_3.set('')
            return

        self.file1Selections[3 - 1] = self.cbFile1_3.current()

    def cbFile2_3Select(self, eventObject):
        if self.cbFile2_3.current() != self.file2Selections[3 - 1] and self.cbFile2_3.current() in self.file2Selections:
            showwarning(title="Προσοχή ...", message="Έχετε ήδη μια αντιστοίχιση με αυτό το πεδίο.")
            if self.file2Selections[3 - 1] != -1:
                self.cbFile2_3.current(self.file2Selections[3 - 1])
            else:
                self.file2_3.set('')
            return

        self.file2Selections[3 - 1] = self.cbFile2_3.current()

    def cbFile1_4Select(self, eventObject):
        if self.cbFile1_4.current() != self.file1Selections[4 - 1] and self.cbFile1_4.current() in self.file1Selections:
            showwarning(title="Προσοχή ...", message="Έχετε ήδη μια αντιστοίχιση με αυτό το πεδίο.")
            if self.file1Selections[4 - 1] != -1:
                self.cbFile1_4.current(self.file1Selections[4 - 1])
            else:
                self.file1_4.set('')
            return

        self.file1Selections[4 - 1] = self.cbFile1_4.current()

    def cbFile2_4Select(self, eventObject):
        if self.cbFile2_4.current() != self.file2Selections[4 - 1] and self.cbFile2_4.current() in self.file2Selections:
            showwarning(title="Προσοχή ...", message="Έχετε ήδη μια αντιστοίχιση με αυτό το πεδίο.")
            if self.file2Selections[4 - 1] != -1:
                self.cbFile2_4.current(self.file2Selections[4 - 1])
            else:
                self.file2_4.set('')
            return

        self.file2Selections[4 - 1] = self.cbFile2_4.current()

    def run(self):
        if (self.file1Selections[1 - 1] == -1 or self.file2Selections[1 - 1] == -1):
            showwarning(title="Προσοχή ...", message="Πρέπει να κάνετε τουλάχιστον μια αντιστοίχιση.")
            return
        else:
            for i in range(1, 4):
                if (self.file2Selections[i] == -1 and self.file1Selections[i] != -1) or (
                        self.file2Selections[i] != -1 and self.file1Selections[i] == -1):
                    showwarning(title="Προσοχή ...",
                                message=f"Η {i + 1}η αντιστοίχιση δεν είναι ολοκληρωμένη. Συμπληρώστε το πεδίο ή αφαιρέσετε την.")
                    return

        self.cbFile1_1.configure(state='disabled')
        self.cbFile1_2.configure(state='disabled')
        self.cbFile1_3.configure(state='disabled')
        self.cbFile1_4.configure(state='disabled')
        self.cbFile2_1.configure(state='disabled')
        self.cbFile2_2.configure(state='disabled')
        self.cbFile2_3.configure(state='disabled')
        self.cbFile2_4.configure(state='disabled')
        self.btnAddRelation2.configure(state='disabled')
        self.btnAddRelation3.configure(state='disabled')
        self.btnAddRelation4.configure(state='disabled')
        self.btnRemoveRelation2.configure(state='disabled')
        self.btnRemoveRelation3.configure(state='disabled')
        self.btnRemoveRelation4.configure(state='disabled')

        self.btnRun.configure(state='disabled')

        self.data3.append(self.data1[0])
        self.data4.append(self.data1[0])

        selections_count = 0
        for item in self.file1Selections:
            if item != -1:
                selections_count += 1
            else:
                break

        input_list = self.data1[1:]

        for i in range(selections_count):
            output_list = list()
            for row1 in input_list:
                match = False

                for row2 in self.data2[1:]:
                    if row2[self.file2Selections[i]] == '':
                        break

                    if self.conformText(row1[self.file1Selections[i]]) == self.conformText(
                            row2[self.file2Selections[i]]):
                        match = True
                        break

                if match:
                    output_list.append(row1)

            input_list = output_list
            self.data3 = output_list

        self.data3 = [self.data1[0]] + self.data3
        self.saveFileMatches()

        for item in self.data1[1:]:
            if item not in self.data3[1:]:
                self.data4.append(item)

        self.saveFileNoMatches()

        showinfo(title="Ολοκλήρωση εκτέλεσης", message="Ο διαχωρισμός ολοκληρώθηκε.")

        self.window.destroy()

    def create_widgets(self):
        self.fData = Frame(self.window)

        self.lData1 = Label(self.fData, text="Αρχείο για διαχωρισμό:")
        self.lData1.grid(column=0, row=0, padx=10, pady=5, sticky=E)

        self.data1Filename = StringVar()
        self.ntrData1Filename = Entry(self.fData, width=128, state='readonly', textvariable=self.data1Filename)
        self.ntrData1Filename.grid(column=1, row=0, padx=10, pady=5, sticky=W)

        self.btnOpenData1 = Button(self.fData, text="Επιλέξτε αρχείο...", command=self.getData1Filename)
        self.btnOpenData1.grid(column=2, row=0, padx=10, pady=5)

        self.lData2 = Label(self.fData, text="Αρχείο με τιμές για το φιλτράρισμα:")
        self.lData2.grid(column=0, row=1, padx=10, pady=5, sticky=E)

        self.data2Filename = StringVar()
        self.ntrData2Filename = Entry(self.fData, width=128, state='disabled', textvariable=self.data2Filename)
        self.ntrData2Filename.grid(column=1, row=1, padx=10, pady=5, sticky=W)

        self.btnOpenData2 = Button(self.fData, text="Επιλέξτε αρχείο...", command=self.getData2Filename,
                                   state='disabled')
        self.btnOpenData2.grid(column=2, row=1, padx=10, pady=5)

        self.lOutputDirName = Label(self.fData, text="Φάκελος για αποθήκευση:")
        self.lOutputDirName.grid(column=0, row=2, padx=10, pady=5, sticky=E)

        self.outputDirName = StringVar()
        self.ntrOutputDirName = Entry(self.fData, width=128, state='disabled', textvariable=self.outputDirName)
        self.ntrOutputDirName.grid(column=1, row=2, padx=10, pady=5, sticky=W)

        self.btnOpenOutputDir = Button(self.fData, text="Επιλέξτε φάκελο...", command=self.getOutputDirName,
                                       state='disabled')
        self.btnOpenOutputDir.grid(column=2, row=2, padx=10, pady=5)

        self.lfRelateFrame = LabelFrame(self.fData, text="Αντιστοιχίσεις")
        self.lfRelateFrame.grid(column=0, row=3, columnspan=3, padx=10, pady=5)

        self.lFile1 = Label(self.lfRelateFrame, text="Πεδία από το πρώτο αρχείο")
        self.lFile1.grid(column=0, row=0, padx=10, pady=1)

        self.lFile2 = Label(self.lfRelateFrame, text="Πεδία από το δεύτερο αρχείο")
        self.lFile2.grid(column=2, row=0, padx=10, pady=1)

        self.lAdd = Label(self.lfRelateFrame, text="Προσθήκη νέας\nαντιστοίχισης", justify=CENTER)
        self.lAdd.grid(column=3, row=0, padx=10, pady=1)

        self.lRemove = Label(self.lfRelateFrame, text="Αφαίρεση τρέχουσας\nαντιστοίχισης", justify=CENTER)
        self.lRemove.grid(column=4, row=0, padx=10, pady=1)
        # ----------------------------------------------------------------------------------------------------------
        self.file1_1 = StringVar()
        self.cbFile1_1 = Combobox(self.lfRelateFrame, width=60, textvariable=self.file1_1, state='disabled')
        self.cbFile1_1.bind("<<ComboboxSelected>>", self.cbFile1_1Select)
        self.cbFile1_1.grid(column=0, row=1, padx=10, pady=5)

        self.lEq1 = Label(self.lfRelateFrame, text="==")
        self.lEq1.grid(column=1, row=1, padx=10, pady=5)

        self.file2_1 = StringVar()
        self.cbFile2_1 = Combobox(self.lfRelateFrame, width=60, textvariable=self.file2_1, state='disabled')
        self.cbFile2_1.bind("<<ComboboxSelected>>", self.cbFile2_1Select)
        self.cbFile2_1.grid(column=2, row=1, padx=10, pady=5)

        self.btnAddRelation2 = Button(self.lfRelateFrame, text="+", command=self.addRelation2, state='disabled')
        self.btnAddRelation2.grid(column=3, row=1, padx=10, pady=5)
        # ----------------------------------------------------------------------------------------------------------

        # ----------------------------------------------------------------------------------------------------------
        self.file1_2 = StringVar()
        self.cbFile1_2 = Combobox(self.lfRelateFrame, width=60, textvariable=self.file1_2, state='disabled')
        self.cbFile1_2.bind("<<ComboboxSelected>>", self.cbFile1_2Select)
        self.cbFile1_2.grid(column=0, row=2, padx=10, pady=5)

        self.lEq2 = Label(self.lfRelateFrame, text="==")
        self.lEq2.grid(column=1, row=2, padx=10, pady=5)

        self.file2_2 = StringVar()
        self.cbFile2_2 = Combobox(self.lfRelateFrame, width=60, textvariable=self.file2_2, state='disabled')
        self.cbFile2_2.bind("<<ComboboxSelected>>", self.cbFile2_2Select)
        self.cbFile2_2.grid(column=2, row=2, padx=10, pady=5)

        self.btnAddRelation3 = Button(self.lfRelateFrame, text="+", command=self.addRelation3, state='disabled')
        self.btnAddRelation3.grid(column=3, row=2, padx=10, pady=5)

        self.btnRemoveRelation2 = Button(self.lfRelateFrame, text="-", command=self.removeRelation2, state='disabled')
        self.btnRemoveRelation2.grid(column=4, row=2, padx=10, pady=5)
        # ----------------------------------------------------------------------------------------------------------

        # ----------------------------------------------------------------------------------------------------------
        self.file1_3 = StringVar()
        self.cbFile1_3 = Combobox(self.lfRelateFrame, width=60, textvariable=self.file1_3, state='disabled')
        self.cbFile1_3.bind("<<ComboboxSelected>>", self.cbFile1_3Select)
        self.cbFile1_3.grid(column=0, row=3, padx=10, pady=5)

        self.lEq3 = Label(self.lfRelateFrame, text="==")
        self.lEq3.grid(column=1, row=3, padx=10, pady=5)

        self.file2_3 = StringVar()
        self.cbFile2_3 = Combobox(self.lfRelateFrame, width=60, textvariable=self.file2_3, state='disabled')
        self.cbFile2_3.bind("<<ComboboxSelected>>", self.cbFile2_3Select)
        self.cbFile2_3.grid(column=2, row=3, padx=10, pady=5)

        self.btnAddRelation4 = Button(self.lfRelateFrame, text="+", command=self.addRelation4, state='disabled')
        self.btnAddRelation4.grid(column=3, row=3, padx=10, pady=5)

        self.btnRemoveRelation3 = Button(self.lfRelateFrame, text="-", command=self.removeRelation3, state='disabled')
        self.btnRemoveRelation3.grid(column=4, row=3, padx=10, pady=5)
        # ----------------------------------------------------------------------------------------------------------

        # ----------------------------------------------------------------------------------------------------------
        self.file1_4 = StringVar()
        self.cbFile1_4 = Combobox(self.lfRelateFrame, width=60, textvariable=self.file1_4, state='disabled')
        self.cbFile1_4.bind("<<ComboboxSelected>>", self.cbFile1_4Select)
        self.cbFile1_4.grid(column=0, row=4, padx=10, pady=5)

        self.lEq4 = Label(self.lfRelateFrame, text="==")
        self.lEq4.grid(column=1, row=4, padx=10, pady=5)

        self.file2_4 = StringVar()
        self.cbFile2_4 = Combobox(self.lfRelateFrame, width=60, textvariable=self.file2_4, state='disabled')
        self.cbFile2_4.bind("<<ComboboxSelected>>", self.cbFile2_4Select)
        self.cbFile2_4.grid(column=2, row=4, padx=10, pady=5)

        self.btnRemoveRelation4 = Button(self.lfRelateFrame, text="-", command=self.removeRelation4, state='disabled')
        self.btnRemoveRelation4.grid(column=4, row=4, padx=10, pady=5)
        # ----------------------------------------------------------------------------------------------------------

        self.btnRun = Button(self.fData, text="Εκτέλεση", command=self.run, state='disabled')
        self.btnRun.grid(column=1, row=10, padx=10, pady=5)

        self.fData.pack()


gui = GUI()
gui.window.mainloop()
