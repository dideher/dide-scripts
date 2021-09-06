from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from openpyxl import *
from openpyxl.utils import get_column_letter
from os import walk
import os
import xlrd


class GUI():
    def __init__(self):
        self.window = Tk()

        self.window.title("Συνένωση αρχείων excel σε ένα")
        self.window.resizable(False, False)
        self.create_widgets()


    def getFilesDirName(self):
        dName = filedialog.askdirectory(initialdir="./data/", title="Επιλέξτε τον φάκελο με τα αρχεία προς συνένωση")

        if dName == "":
            return

        self.filesDirName.set(dName)
        self.btnRun.configure(state='normal')


    def is_float(self, s):
        return s.replace('.', '', 1).isdecimal()


    def clean_spaces(self, text):
        while "  " in text:
            text = text.replace("  ", " ")

        return text


    def convertXls(self):
        inputDirectory = self.filesDirName.get()

        for path, dirs, files in walk(inputDirectory):
            for file in files:
                if file[-4:] == ".xls":
                    book_xls = xlrd.open_workbook(os.path.join(path, file))
                    book_xlsx = Workbook()

                    sheet_names = book_xls.sheet_names()
                    for sheet_index, sheet_name in enumerate(sheet_names):
                        sheet_xls = book_xls.sheet_by_name(sheet_name)
                        if sheet_index == 0:
                            sheet_xlsx = book_xlsx.active
                            sheet_xlsx.title = sheet_name
                        else:
                            sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

                        for row in range(0, sheet_xls.nrows):
                            for col in range(0, sheet_xls.ncols):
                                sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)

                    book_xlsx.save(os.path.join(path, file + 'x'))


    def checkHeader(self, header, row):
        for item1, item2 in zip(header, row):
            if item1 != item2.value:
                return False

        return True


    def checkC3forSchoolname(self, schoolName):
        return schoolName != '(επιλέξτε το σχολείο σας)'


    def run(self):
        self.btnRun.configure(state='disabled')

        self.convertXls()

        inputDirectory = self.filesDirName.get()

        header = [
                    "Α/Α",
                    "Επώνυμο",
                    "Όνομα",
                    "Πατρώνυμο",
                    "Κλάδος",
                    "Βαθμός",
                    "Αριθμός Μητρώου Υπουργείου Παιδείας (για μονιμους)",
                    "ΑΦΜ",
                    "Νεα Ελληνική Γλώσσα και Λογοτεχνία",
                    "Αρχαία Ελληνικά (Ο.Π.)",
                    "Ιστορία (Ο.Π.)",
                    "Κοινωνιολογία (Ο.Π.)",
                    "Μαθηματικά (Ο.Π.)",
                    "Φυσική (Ο.Π.)",
                    "Χημεία (Ο.Π.)",
                    "Βιολογία (Ο.Π.)",
                    "Μαθηματικά (Ο.Π.)",
                    "Πληροφορική (Ο.Π.)",
                    "Οικονομία",
                    "Τηλέφωνο",
                    "Παρατηρήσεις",
                    "Λύκειο όπου διδάσκει"
        ]

        schoolsPending = {
                    "1ο ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ",
                    "2ο ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ",
                    "3ο ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ",
                    "4ο ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ",
                    "5ο ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ",
                    "6ο ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ",
                    "7ο ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ",
                    "8ο ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ",
                    "10ο ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ",
                    "11ο ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ",
                    "13ο ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ",
                    "ΓΕΝΙΚΟ ΛΥΚΕΙΟ Ν.ΑΛΙΚΑΡΝΑΣΣΟΥ",
                    "ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΑΓΙΑΣ ΒΑΡΒΑΡΑΣ",
                    "ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΑΓΙΟΥ ΜΥΡΩΝΑ",
                    "ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΑΓΙΩΝ ΔΕΚΑ",
                    "ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΑΡΚΑΛΟΧΩΡΙΟΥ",
                    "ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΑΡΧΑΝΩΝ",
                    "ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΑΣΗΜΙΟΥ",
                    "ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΒΙΑΝΝΟΥ",
                    "ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΓΑΖΙΟΥ",
                    "ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΓΟΥΒΩΝ",
                    "ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΕΠΙΣΚΟΠΗΣ",
                    "ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΚΑΣΤΕΛΙΟΥ",
                    "ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΚΡΟΥΣΩΝΑ",
                    "ΓΕΝΙΚΟ ΛΥΚΕΙΟ Λ. ΧΕΡΣΟΝΗΣΟΥ",
                    "ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΜΑΛΙΩΝ",
                    "ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΜΕΛΕΣΩΝ",
                    "ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΜΟΙΡΩΝ",
                    "ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΜΟΧΟΥ",
                    "ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΠΟΜΠΙΑΣ",
                    "ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΤΥΜΠΑΚΙΟΥ",
                    "ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΧΑΡΑΚΑ",
                    "ΕΣΠΕΡΙΝΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ",
                    "ΚΑΛΛΙΤΕΧΝΙΚΟ ΣΧΟΛΕΙΟ (ΛΥΚΕΙΑΚΕΣ ΤΑΞΕΙΣ)",
                    "ΜΟΥΣΙΚΟ ΣΧΟΛΕΙΟ - ΛΥΚΕΙΑΚΕΣ ΤΑΞΕΙΣ",
                    "ΠΡΟΤΥΠΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ",
                    "ΤΟ ΠΑΓΚΡΗΤΙΟΝ - ΓΕΝ. ΛΥΚΕΙΟ",
                    "ΕΣΠΕΡΙΝΟ ΓΥΜΝΑΣΙΟ ΤΥΜΠΑΚΙΟΥ - ΛΥΚΕΙΑΚΕΣ ΤΑΞΕΙΣ"
        }

        data = list()
        errors = ""
        errors_counter = 0

        for path, dirs, files in walk(inputDirectory):
            for file in files:
                if file[-5:] == ".xlsx":
                    workbook = load_workbook(filename=path + "\\" + file)
                    sheet = workbook.active

                    if (sheet['C3'].value is None):
                        errors_counter += 1
                        errors += f'{errors_counter}: "{file}" Δεν έχει όνομα σχολείου στο κελί C3.\n'
                        continue

                    schoolName = sheet['C3'].value.strip()

                    if not self.checkC3forSchoolname(schoolName):
                        errors_counter += 1
                        errors += f'{errors_counter}: "{file}" Δεν έχει όνομα σχολείου στο κελί C3.\n'
                        continue

                    if schoolName not in schoolsPending:
                        errors_counter += 1
                        errors += f'{errors_counter}: "{file}" Υπάρχει και άλλο αρχείο για το σχολείο "{schoolName}".\n'
                        continue
                    else:
                        schoolsPending.remove(schoolName)

                    if not self.checkHeader(header[:-3], sheet[7]):
                        errors_counter += 1
                        errors += f'{errors_counter}: "{file}" Δεν έχει τις σωστές στήλες.\n'
                        continue

                    endOfData = False

                    for i, row in enumerate(sheet.iter_rows()):
                        rowStartsWithNumber = False
                        entry = list()

                        for j, cell in enumerate(row):
                            if cell.value is None:
                                text = ""
                            else:
                                text = self.clean_spaces(str(cell.value))

                            if j == 21 and text == "":
                                text = schoolName

                            if j == 0 and text.isdecimal() and int(text) >= 1:
                                rowStartsWithNumber = True

                            if j > 21:
                                break

                            if rowStartsWithNumber and j == 1 and text == "":
                                endOfData = True
                                break

                            entry.append(text)

                        if rowStartsWithNumber:
                            if endOfData:
                                break
                            else:
                                data.append(entry)

        wb = Workbook()
        ws = wb.active

        ws.append(header)

        for entry in data:
            ws.append(entry)

        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

        outputFile = "output.xlsx"

        notSaved = True

        while notSaved:
            try:
                wb.save(outputFile)
            except:
                showwarning(title="Αρχείο σε χρήση...", message="Παρακαλώ κλείστε το αρχείο '{}' ώστε να ολοκληρωθεί η αποθήκευση.".format(outputFile))
            else:
                notSaved = False

        showinfo(title="Αρχείο εξόδου",
                    message=f'Η συνένωση έχει αποθηκευτεί στο αρχείο "{outputFile}" στον φάκελο εκτέλεσης του προγράμματος.')

        print(f'{"-" * 20} Λάθη {"-" * 20}')
        print(errors)
        print(f'{"-" * 20} Εκκρεμότητες {"-" * 20}')
        print(schoolsPending)


    def create_widgets(self):
        self.fData = Frame(self.window)

        self.lData = Label(self.fData, text="Φάκελος με αρχεία\nπρος συνένωση:")
        self.lData.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.filesDirName = StringVar()
        self.ntrFilesDirName = Entry(self.fData, width=100, state='readonly', textvariable=self.filesDirName)
        self.ntrFilesDirName.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenData = Button(self.fData, text="Επιλέξτε φάκελο...", command=self.getFilesDirName)
        self.btnOpenData.grid(column=2, row=0, padx=10, pady=10)

        self.btnRun = Button(self.fData, text="Εκτέλεση συνένωσης", command=self.run, state='disabled')
        self.btnRun.grid(column=1, row=10, padx=10, pady=10)

        self.fData.pack()


gui = GUI()
gui.window.mainloop()
