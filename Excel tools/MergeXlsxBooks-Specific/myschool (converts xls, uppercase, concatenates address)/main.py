from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from openpyxl import *
from openpyxl.utils import get_column_letter
from os import walk
import os
import xlrd
import re


class GUI():
    def __init__(self):
        self.window = Tk()

        self.window.title("Συνένωση αρχείων excel σε ένα")
        self.window.resizable(False, False)
        self.create_widgets()

    def getFilesDirName(self):
        dName = filedialog.askdirectory(initialdir="./data/", title="Επιλέξτε τον φάκελο με τα αρχεία προς συνένωση")

        if dName == "":
            return

        self.filesDirName.set(dName)
        self.btnRun.configure(state='normal')


    def toUpperCase(self, text):
        text1 = (str(text).upper().replace(".", ". ").replace(" .", ". ").replace("Ά", "Α").replace("Έ", "Ε")
                 .replace("Ή", "Η").replace("Ί", "Ι").replace("Ϊ́", "Ϊ").replace("Ύ", "Υ").replace("Ϋ́", "Ϋ")
                 .replace("Ό", "Ο").replace("Ώ", "Ω").strip())

        text2 = re.sub(r'([ ]+)', r' ', text1)
        text3 = re.sub(r'([0-9]+)Ο', r'\1ο', text2)

        return text3

    def convertXls(self):
        inputDirectory = self.filesDirName.get()

        for path, dirs, files in walk(inputDirectory):
            for file in files:
                if file[-4:] == ".xls":
                    book_xls = xlrd.open_workbook(os.path.join(path, file))
                    book_xlsx = Workbook()

                    sheet_names = book_xls.sheet_names()
                    for sheet_index, sheet_name in enumerate(sheet_names):
                        sheet_xls = book_xls.sheet_by_name(sheet_name)
                        if sheet_index == 0:
                            sheet_xlsx = book_xlsx.active
                            sheet_xlsx.title = sheet_name
                        else:
                            sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

                        for row in range(0, sheet_xls.nrows):
                            for col in range(0, sheet_xls.ncols):
                                sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)

                    book_xlsx.save(os.path.join(path, file + 'x'))

    def run(self):
        self.btnRun.configure(state='disabled')
        showinfo(title="Εκτέλεση συνένωσης",
                 message=f'Μην τερματίσετε την εφαρμογή μέχρι να εμφανιστεί το μήνυμα ολοκλήρωσης.')

        self.convertXls()

        inputDirectory = self.filesDirName.get()

        header = ["Α/Α", "ΕΠΩΝΥΜΟ ΜΑΘΗΤΗ", "ΟΝΟΜΑ ΜΑΘΗΤΗ", "ΟΝΟΜΑ ΠΑΤΕΡΑ", "ΔΙΕΥΘΥΝΣΗ, ΟΔΟΣ - ΑΡΙΘΜΟΣ",
                  "ΔΙΕΥΘΥΝΣΗ, Τ.Κ.", "ΔΙΕΥΘΥΝΣΗ, ΠΕΡΙΟΧΗ", "ΔΙΕΥΘΥΝΣΗ (ΣΥΓΚΕΝΤΡΩΤΙΚΗ)", "ΣΧΟΛΕΙΟ"]

        data = list()

        for path, dirs, files in walk(inputDirectory):
            for file in files:
                if file[-5:] == ".xlsx":
                    school = ""
                    workbook = load_workbook(filename=path + "\\" + file)

                    for sheet in workbook.worksheets:
                        for i, row in enumerate(sheet.iter_rows()):
                            rowContainsData = False
                            entry = list()

                            for j, cell in enumerate(row):
                                if cell.value is None:
                                    text = ""
                                else:
                                    text = self.toUpperCase(str(cell.value))

                                if "ΣΧΟΛΕΙΟ" in text:
                                    school = text

                                if j == 0 and text.isdecimal() and int(text) >= 1:
                                    rowContainsData = True

                                if rowContainsData:
                                    if j == 4:
                                        continue
                                    elif j > 7:
                                        entry.append(f"{entry[4]} {entry[5]} {entry[6]}")
                                        if school == "":
                                            school = f"[Όνομα αρχείου] {file}"

                                        entry.append(school)
                                        break
                                    else:
                                        entry.append(text)

                            if rowContainsData:
                                data.append(entry)

        wb = Workbook()
        ws = wb.active

        ws.append(header)

        for entry in data:
            ws.append(entry)

        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

        outputFile = "output.xlsx"

        notSaved = True

        while notSaved:
            try:
                wb.save(outputFile)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message="Παρακαλώ κλείστε το αρχείο '{}' ώστε να ολοκληρωθεί η αποθήκευση.".format(
                                outputFile))
            else:
                notSaved = False

        showinfo(title="Αρχείο εξόδου",
                 message=f'Η συνένωση έχει αποθηκευτεί στο αρχείο "{outputFile}" στον φάκελο εκτέλεσης του προγράμματος.')

        self.btnRun.configure(state='normal')


    def create_widgets(self):
        self.fData = Frame(self.window)

        self.lData = Label(self.fData, text="Φάκελος με αρχεία\nπρος συνένωση:")
        self.lData.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.filesDirName = StringVar()
        self.ntrFilesDirName = Entry(self.fData, width=100, state='readonly', textvariable=self.filesDirName)
        self.ntrFilesDirName.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenData = Button(self.fData, text="Επιλέξτε φάκελο...", command=self.getFilesDirName)
        self.btnOpenData.grid(column=2, row=0, padx=10, pady=10)

        self.btnRun = Button(self.fData, text="Εκτέλεση συνένωσης", command=self.run, state='disabled')
        self.btnRun.grid(column=1, row=10, padx=10, pady=10)

        self.fData.pack()


gui = GUI()
gui.window.mainloop()
