from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showinfo
from tkinter.ttk import *
from openpyxl import *
from openpyxl.utils import get_column_letter
from os import walk
import os
import xlrd


class GUI:
    def __init__(self):
        self.window = Tk()

        self.window.title("Μετατροπή αρχείων xls σε xlsx")
        self.window.resizable(False, False)
        self.create_widgets()

    def getFilesDirName(self):
        dName = filedialog.askdirectory(initialdir="./data/", title="Επιλέξτε τον φάκελο με τα αρχεία προς μετατροπή")

        if dName == "":
            return

        self.filesDirName.set(dName)
        self.btnRun.configure(state='normal')

    def setColsWidth(self, ws):
        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

    def convertXls(self):
        inputDirectory = self.filesDirName.get()

        for path, dirs, files in walk(inputDirectory):
            for file in files:
                if file[-4:] == ".xls":
                    book_xls = xlrd.open_workbook(os.path.join(path, file))
                    book_xlsx = Workbook()

                    sheet_names = book_xls.sheet_names()
                    for sheet_index, sheet_name in enumerate(sheet_names):
                        sheet_xls = book_xls.sheet_by_name(sheet_name)
                        if sheet_index == 0:
                            sheet_xlsx = book_xlsx.active
                            sheet_xlsx.title = sheet_name
                        else:
                            sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

                        for row in range(0, sheet_xls.nrows):
                            for col in range(0, sheet_xls.ncols):
                                sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)

                        self.setColsWidth(sheet_xlsx)

                    book_xlsx.save(os.path.join(path, file + 'x'))

    def run(self):
        self.btnRun.configure(state='disabled')
        self.convertXls()
        showinfo(title="Ολοκλήρωση εκτέλεσης", message="Η μετροπή των αρχείων ολοκληρώθηκε.")

    def create_widgets(self):
        self.fData = Frame(self.window)

        self.lData = Label(self.fData, text="Φάκελος με αρχεία\nπρος μετατροπή:")
        self.lData.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.filesDirName = StringVar()
        self.ntrFilesDirName = Entry(self.fData, width=100, state='readonly', textvariable=self.filesDirName)
        self.ntrFilesDirName.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenData = Button(self.fData, text="Επιλέξτε φάκελο...", command=self.getFilesDirName)
        self.btnOpenData.grid(column=2, row=0, padx=10, pady=10)

        self.btnRun = Button(self.fData, text="Εκτέλεση", command=self.run, state='disabled')
        self.btnRun.grid(column=1, row=10, padx=10, pady=10)

        self.fData.pack()


gui = GUI()
gui.window.mainloop()
