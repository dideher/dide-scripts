from tkinter import *
from tkinter import filedialog
from tkinter.ttk import *
from tkinter.messagebox import showwarning, showinfo
from openpyxl import *
from openpyxl.utils import get_column_letter
import os
import xlrd


class GUI():
    def __init__(self):
        self.data1 = list()
        self.data2 = list()
        self.common = list()
        self.onlyIn1 = list()
        self.onlyIn2 = list()

        self.window = Tk()

        self.window.title("Διαχωρισμός αρχείων excel σε κοινές και διαφορετικές εγγραφές")
        self.window.resizable(False, False)
        self.create_widgets()

    def is_float(self, s):
        return s.replace('.', '', 1).isdecimal()

    def convertXls(self, file):
        book_xls = xlrd.open_workbook(file)
        book_xlsx = Workbook()

        sheet_names = book_xls.sheet_names()
        for sheet_index, sheet_name in enumerate(sheet_names):
            sheet_xls = book_xls.sheet_by_name(sheet_name)
            if sheet_index == 0:
                sheet_xlsx = book_xlsx.active
                sheet_xlsx.title = sheet_name
            else:
                sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

            for row in range(0, sheet_xls.nrows):
                for col in range(0, sheet_xls.ncols):
                    sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)

        book_xlsx.save(file + 'x')

    def parseXlsxData(self, file, data):
        workbook = load_workbook(filename=file)
        sheet = workbook.active

        for row in sheet.iter_rows():
            entry = list()
            for cell in row:
                if cell.value is None:
                    text = ""
                else:
                    text = str(cell.value).strip()

                if self.is_float(text):
                    text = text.replace(".", ",")

                entry.append(text)

            data.append(entry)

    def setColsWidth(self, ws):
        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

    def saveFile(self, l, filename):
        wb = Workbook()
        ws = wb.active

        for row in l:
            ws.append(row)

        self.setColsWidth(ws)

        outputFile = os.path.join(self.outputDirName.get(), filename)

        notSaved = True

        while notSaved:
            try:
                wb.save(outputFile)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message=f"Παρακαλώ κλείστε το αρχείο '{filename}' ώστε να ολοκληρωθεί η αποθήκευση.")
            else:
                notSaved = False

    def getOutputDirName(self):
        dName = filedialog.askdirectory(initialdir="./data/", title="Επιλέξτε τον φάκελο που θα αποθηκευτούν τα αρχεία")

        if dName == "":
            return

        self.outputDirName.set(dName)
        self.ntrOutputDirName.configure(state='disabled')
        self.btnOpenOutputDir.configure(state='disabled')
        self.btnRun.configure(state='normal')

    def getData1Filename(self):
        fName = filedialog.askopenfilename(initialdir="./data/",
                                           title="Επιλέξτε το πρώτο αρχείο excel",
                                           filetypes=(
                                               ("xlsx files", "*.xlsx"), ("xls files", "*.xls"), ("all files", "*.*")))

        if fName == "":
            return

        if fName.endswith('.xls'):
            self.convertXls(fName)
            fName = fName + 'x'

        self.data1Filename.set(fName)
        self.ntrData1Filename.configure(state='disabled')
        self.btnOpenData1.configure(state='disabled')
        self.parseXlsxData(fName, self.data1)
        self.ntrData2Filename.configure(state='readonly')
        self.btnOpenData2.configure(state='normal')

    def getData2Filename(self):
        fName = filedialog.askopenfilename(initialdir="./data/",
                                           title="Επιλέξτε το δεύτερο αρχείο excel",
                                           filetypes=(
                                               ("xlsx files", "*.xlsx"), ("xls files", "*.xls"), ("all files", "*.*")))

        if fName == "":
            return

        if fName == self.data1Filename.get() or fName == self.data1Filename.get()[:-1]:
            showwarning(title="Προσοχή ...", message="Έχετε ήδη επιλέξει αυτό το αρχείο ως πρώτο αρχείο.")
            return

        if fName.endswith('.xls'):
            self.convertXls(fName)
            fName = fName + 'x'

        self.data2Filename.set(fName)
        self.ntrData2Filename.configure(state='disabled')
        self.btnOpenData2.configure(state='disabled')
        self.parseXlsxData(fName, self.data2)
        self.ntrOutputDirName.configure(state='readonly')
        self.btnOpenOutputDir.configure(state='normal')

    def run(self):
        self.btnRun.configure(state='disabled')

        for item in self.data1[1:]:
            if item in self.data2[1:]:
                self.common.append(item)
            else:
                self.onlyIn1.append(item)

        for item in self.data2[1:]:
            if item not in self.data1[1:]:
                self.onlyIn2.append(item)

        header = self.data1[0]

        outputDir = self.outputDirName.get()

        if len(self.common) != 0:
            saveList = list()
            saveList.append(header)
            saveList += self.common
            self.saveFile(saveList, os.path.join(outputDir, 'common.xlsx'))

        if len(self.onlyIn1) != 0:
            saveList = list()
            saveList.append(header)
            saveList += self.onlyIn1
            self.saveFile(saveList, os.path.join(outputDir, 'onlyIn1stFile.xlsx'))

        if len(self.onlyIn2) != 0:
            saveList = list()
            saveList.append(header)
            saveList += self.onlyIn2
            self.saveFile(saveList, os.path.join(outputDir, 'onlyIn2ndFile.xlsx'))

        showinfo(title="Ολοκλήρωση εκτέλεσης", message="Ο διαχωρισμός ολοκληρώθηκε.")

    def create_widgets(self):
        self.fData = Frame(self.window)

        self.lData1 = Label(self.fData, text="Πρώτο αρχείο:")
        self.lData1.grid(column=0, row=0, padx=10, pady=5, sticky=E)

        self.data1Filename = StringVar()
        self.ntrData1Filename = Entry(self.fData, width=128, state='readonly', textvariable=self.data1Filename)
        self.ntrData1Filename.grid(column=1, row=0, padx=10, pady=5, sticky=W)

        self.btnOpenData1 = Button(self.fData, text="Επιλέξτε αρχείο...", command=self.getData1Filename)
        self.btnOpenData1.grid(column=2, row=0, padx=10, pady=5)

        self.lData2 = Label(self.fData, text="Δεύτερο αρχείο:")
        self.lData2.grid(column=0, row=1, padx=10, pady=5, sticky=E)

        self.data2Filename = StringVar()
        self.ntrData2Filename = Entry(self.fData, width=128, state='disabled', textvariable=self.data2Filename)
        self.ntrData2Filename.grid(column=1, row=1, padx=10, pady=5, sticky=W)

        self.btnOpenData2 = Button(self.fData, text="Επιλέξτε αρχείο...", command=self.getData2Filename,
                                   state='disabled')
        self.btnOpenData2.grid(column=2, row=1, padx=10, pady=5)

        self.lOutputDirName = Label(self.fData, text="Φάκελος για αποθήκευση:")
        self.lOutputDirName.grid(column=0, row=2, padx=10, pady=5, sticky=E)

        self.outputDirName = StringVar()
        self.ntrOutputDirName = Entry(self.fData, width=128, state='disabled', textvariable=self.outputDirName)
        self.ntrOutputDirName.grid(column=1, row=2, padx=10, pady=5, sticky=W)

        self.btnOpenOutputDir = Button(self.fData, text="Επιλέξτε φάκελο...", command=self.getOutputDirName,
                                       state='disabled')
        self.btnOpenOutputDir.grid(column=2, row=2, padx=10, pady=5)

        self.btnRun = Button(self.fData, text="Εκτέλεση", command=self.run, state='disabled')
        self.btnRun.grid(column=1, row=10, padx=10, pady=5)

        self.fData.pack()


gui = GUI()
gui.window.mainloop()
