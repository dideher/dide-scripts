from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


class GUI:
    def __init__(self):
        self.window = Tk()

        self.window.title("Καθαρισμός αρχείου xlsx από στήλες που δεν χρειάζονται")
        self.window.resizable(False, False)
        self.create_widgets()

    def transpose_data(self):
        data = list()

        for item in self.headers:
            data.append(item[0])

        self.headers = data

    def get_headers(self):
        f_name = filedialog.askopenfilename(initialdir="./data", title="Επιλέξτε το αρχείο xlsx",
                                            filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if f_name == "":
            return

        self.headers = self.parse_xlsx_data(f_name)

        entries_listed_in_rows = True
        for item in self.headers:
            if not (type(item) is list and len(item) == 1):
                entries_listed_in_rows = False
                break

        if entries_listed_in_rows:
            self.transpose_data()
        else:
            self.headers = self.headers[0]

        self.btn_get_headers.configure(state='disabled')
        self.headers_filename.set(f_name)

        self.btn_get_xlsx.configure(state='normal')

    def get_xlsx(self):
        f_name = filedialog.askopenfilename(initialdir="./data", title="Επιλέξτε το αρχείο xlsx",
                                            filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if f_name == "":
            return

        self.xlsx_data = self.parse_xlsx_data(f_name)

        self.btn_get_xlsx.configure(state='disabled')
        self.xlsx_filename.set(f_name)

        self.btn_run.configure(state='normal')

    def parse_xlsx_data(self, file):
        workbook = load_workbook(filename=file)
        sheet = workbook.active

        data = list()

        for row in sheet.iter_rows():
            entry = list()
            for cell in row:
                if cell.value is None:
                    text = ""
                else:
                    text = str(cell.value).strip()

                entry.append(text)

            data.append(entry)

        return data

    def save_file(self, data, output_file):
        wb = Workbook()
        ws = wb.active

        for entry in data:
            ws.append(entry)

        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

        not_saved = True

        while not_saved:
            try:
                wb.save(output_file)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message=f"Παρακαλώ κλείστε το αρχείο '{output_file}' ώστε να ολοκληρωθεί η αποθήκευση.")
            else:
                not_saved = False

    def clean_data(self):
        clean_data = list()
        headers_index_list = list()

        if self.xlsx_data[0] == self.headers:
            showinfo(title="Καθαρό αρχείο...",
                     message=f"Το αρχείο είναι ήδη καθαρό.")

            return None

        cols = ''
        for item in self.headers:
            if item not in self.xlsx_data[0]:
                cols += f"- {item}\n"
            else:
                headers_index_list.append(self.xlsx_data[0].index(item))

        if cols != '':
            showwarning(title="Λάθος τύπος αρχείου...",
                        message=f"Το αρχείο πρέπει να έχει επιπλέον τις εξής στήλες:\n{cols}")

            return None

        for item in self.xlsx_data:
            entry = list()

            for i in headers_index_list:
                entry.append(item[i])

            clean_data.append(entry)

        return clean_data

    def run(self):
        self.btn_run.configure(state='disabled')

        clean_data = self.clean_data()
        output_file = "output.xlsx"

        if clean_data:
            self.save_file(clean_data, output_file)

            showinfo(title='Ολοκλήρωση Εκτέλεσης',
                     message=f'Η δημιουργία του αρχείου xlsx ολοκληρώθηκε.')

        self.window.destroy()

    def create_widgets(self):
        self.f_data = Frame(self.window)

        self.l_headers = Label(self.f_data, text="Στήλες που πρέπει να υπάρχουν (xlsx):")
        self.l_headers.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.headers_filename = StringVar()
        self.headers_filename.set('')
        self.ntr_headers_filename = Entry(self.f_data, width=128, state='readonly', textvariable=self.headers_filename)
        self.ntr_headers_filename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btn_get_headers = Button(self.f_data, text="Επιλέξτε αρχείο...", command=self.get_headers)
        self.btn_get_headers.grid(column=2, row=0, padx=10, pady=10)

        self.l_xlsx = Label(self.f_data, text="Αρχείο για καθαρισμό (xlsx):")
        self.l_xlsx.grid(column=0, row=1, padx=10, pady=10, sticky=E)

        self.xlsx_filename = StringVar()
        self.xlsx_filename.set('')
        self.ntr_xlsx_filename = Entry(self.f_data, width=128, state='readonly', textvariable=self.xlsx_filename)
        self.ntr_xlsx_filename.grid(column=1, row=1, padx=10, pady=10, sticky=W)

        self.btn_get_xlsx = Button(self.f_data, text="Επιλέξτε αρχείο...", command=self.get_xlsx,
                                   state='disabled')
        self.btn_get_xlsx.grid(column=2, row=1, padx=10, pady=10)

        self.btn_run = Button(self.f_data, text="Εκτέλεση", command=self.run, state='disabled')
        self.btn_run.grid(column=1, row=10, padx=10, pady=10)

        self.f_data.pack()


gui = GUI()
gui.window.mainloop()
