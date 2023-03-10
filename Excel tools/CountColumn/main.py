from tkinter import *
from tkinter import filedialog
from tkinter.ttk import *
from tkinter.messagebox import showwarning, showinfo
from openpyxl import *
from openpyxl.utils import get_column_letter
import os


class GUI:
    def __init__(self):
        self.window = Tk()

        self.window.title("Καταμέτρηση πλήθους τιμών στήλης αρχείου xlsx")
        self.window.resizable(False, False)
        self.create_widgets()

    def parse_xlsx_data(self):
        wb = load_workbook(filename=self.data_filename.get())
        sheet = wb.active

        self.xlsx_data = list()

        for row in sheet.iter_rows():
            entry = list()
            for cell in row:
                if cell.value is None:
                    text = ""
                else:
                    text = str(cell.value).strip()

                entry.append(text)

            self.xlsx_data.append(entry)

    def split_data(self):
        data = dict()

        for row in self.xlsx_data[1:]:
            key = row[self.cb_filter_cols.current()]

            if key == "":
                key = "_EMPTY CELLS_"

            if key not in data:
                data[key] = list()
            data[key].append(row)

        logb = Workbook()
        logs = logb.active
        logh = [self.filter_cols.get(), "Πλήθος"]
        logs.append(logh)

        for key in data:
            logd = [key, len(data[key])]
            logs.append(logd)

        self.set_cols_width(logs)

        out_file = "_log.xlsx"
        output_file = os.path.join(self.output_dir.get(), out_file)
        self.safe_save(logb, output_file)

    def set_cols_width(self, ws):
        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

    def safe_save(self, wb, output_file):
        not_saved = True

        while not_saved:
            try:
                wb.save(output_file)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message=f"Παρακαλώ κλείστε το αρχείο '{output_file}' ώστε να ολοκληρωθεί η αποθήκευση.")
            else:
                not_saved = False

    def get_output_dir(self):
        d_name = filedialog.askdirectory(initialdir="./data/",
                                         title="Επιλέξτε τον φάκελο που θα αποθηκευτεί το αρχείο")

        if d_name == "":
            return

        self.output_dir.set(d_name)
        self.btn_run.configure(state='normal')

    def get_data(self):
        f_name = filedialog.askopenfilename(initialdir="./data/",
                                            title="Επιλέξτε το αρχείο xlsx",
                                            filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if f_name == "":
            return

        self.data_filename.set(f_name)
        self.parse_xlsx_data()
        self.cb_filter_cols.configure(state='readonly')
        self.cb_filter_cols['values'] = self.xlsx_data[0]

    def cb_filter_cols_select(self, event_object):
        if self.cb_filter_cols.current() != -1:
            self.ntr_output_dir.configure(state='readonly')
            self.btn_get_output_dir.configure(state='normal')

    def run(self):
        self.split_data()
        showinfo(title="Ολοκλήρωση εκτέλεσης", message="Η καταμέτρηση ολοκληρώθηκε.")

    def create_widgets(self):
        self.f_data = Frame(self.window)

        self.l_data = Label(self.f_data, text="Αρχείο:")
        self.l_data.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.data_filename = StringVar()
        self.ntr_data_filename = Entry(self.f_data, width=128, state='readonly', textvariable=self.data_filename)
        self.ntr_data_filename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btn_get_data = Button(self.f_data, text="Επιλέξτε αρχείο...", command=self.get_data)
        self.btn_get_data.grid(column=2, row=0, padx=10, pady=10)

        self.l_filter_col = Label(self.f_data, text="Στήλη για καταμέτρηση:")
        self.l_filter_col.grid(column=0, row=1, padx=10, pady=10, sticky=E)

        self.filter_cols = StringVar()
        self.cb_filter_cols = Combobox(self.f_data, width=40, textvariable=self.filter_cols, state='disabled')
        self.cb_filter_cols.bind("<<ComboboxSelected>>", self.cb_filter_cols_select)
        self.cb_filter_cols.grid(column=1, row=1, padx=10, pady=10, sticky='NSEW')

        self.l_output_dir = Label(self.f_data, text="Φάκελος για αποθήκευση του αρχείου:")
        self.l_output_dir.grid(column=0, row=2, padx=10, pady=10, sticky=E)

        self.output_dir = StringVar()
        self.ntr_output_dir = Entry(self.f_data, width=128, state='disabled', textvariable=self.output_dir)
        self.ntr_output_dir.grid(column=1, row=2, padx=10, pady=10, sticky=W)

        self.btn_get_output_dir = Button(self.f_data, text="Επιλέξτε φάκελο...", command=self.get_output_dir,
                                         state='disabled')
        self.btn_get_output_dir.grid(column=2, row=2, padx=10, pady=10)

        self.btn_run = Button(self.f_data, text="Εκτέλεση", command=self.run, state='disabled')
        self.btn_run.grid(column=1, row=10, padx=10, pady=10)

        self.f_data.pack()


gui = GUI()
gui.window.mainloop()
