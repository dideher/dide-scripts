from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import csv


class GUI():
    def __init__(self):
        self.window = Tk()

        self.window.title("Μετατροπή ενός αρχείου csv σε xlsx")
        self.window.resizable(False, False)
        self.create_widgets()

    def getInputFilename(self):
        fName = filedialog.askopenfilename(initialdir=".", title="Επιλέξτε το αρχείο csv",
                                           filetypes=(("csv files", "*.csv"), ("all files", "*.*")))

        if fName == "":
            return

        self.btnOpenInputFile.configure(state='disabled')
        self.inputFilename.set(fName)

        self.data = self.parseCsvData(fName)
        self.btnRun.configure(state='normal')

    def parseCsvData(self, inputFile):
        data = list()

        try:
            with open(inputFile, 'rt', encoding='utf-8-sig') as f:
                dialect = csv.Sniffer().sniff(f.readline())
                f.seek(0)

                reader = csv.reader(f, delimiter=dialect.delimiter, quotechar=dialect.quotechar)

                for row in reader:
                    data.append(row)
        except:
            with open(inputFile, 'rt') as f:
                dialect = csv.Sniffer().sniff(f.readline())
                f.seek(0)

                reader = csv.reader(f, delimiter=dialect.delimiter, quotechar=dialect.quotechar)

                for row in reader:
                    data.append(row)

        return data

    def saveFile(self, data, outputFile):
        wb = Workbook()
        ws = wb.active

        for entry in data:
            ws.append(entry)

        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

        notSaved = True

        while notSaved:
            try:
                wb.save(outputFile)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message="Παρακαλώ κλείστε το αρχείο '{}' ώστε να ολοκληρωθεί η αποθήκευση.".format(
                                outputFile))
            else:
                notSaved = False

    def run(self):
        self.btnRun.configure(state='disabled')

        inputFile = self.inputFilename.get()
        outputFile = inputFile.replace(".csv", ".xlsx")

        self.saveFile(self.data, outputFile)

        showinfo(title='Ολοκλήρωση Εκτέλεσης',
                 message=f'Η μετατροπή του αρχείου csv σε xlsx ολοκληρώθηκε.')

    def create_widgets(self):
        self.fData = Frame(self.window)

        self.lInputFile = Label(self.fData, text="Αρχείο csv:")
        self.lInputFile.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.inputFilename = StringVar()
        self.inputFilename.set('')
        self.ntrInputFilename = Entry(self.fData, width=128, state='readonly', textvariable=self.inputFilename)
        self.ntrInputFilename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenInputFile = Button(self.fData, text="Επιλέξτε αρχείο...", command=self.getInputFilename)
        self.btnOpenInputFile.grid(column=2, row=0, padx=10, pady=10)

        self.btnRun = Button(self.fData, text="Εκτέλεση", command=self.run, state='disabled')
        self.btnRun.grid(column=1, row=10, padx=10, pady=10)

        self.fData.pack()


gui = GUI()
gui.window.mainloop()
