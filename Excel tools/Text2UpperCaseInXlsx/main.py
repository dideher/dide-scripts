from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from openpyxl import *
from openpyxl.utils import get_column_letter
from parseXlsxData import *


class GUI():
    def __init__(self):
        self.window = Tk()

        self.window.title("Μετατροπή του κειμένου ενός αρχείου xlsx σε κεφαλαία")
        self.window.resizable(False, False)
        self.create_widgets()

    def getInputFilename(self):
        fName = filedialog.askopenfilename(initialdir=".", title="Επιλέξτε το αρχείο xlsx",
                                           filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if fName == "":
            return

        self.btnOpenInputFile.configure(state='disabled')
        self.inputFilename.set(fName)
        self.data = parseXlsxData(fName)
        self.btnRun.configure(state='normal')

    def saveFile(self, data, outputFile):
        wb = Workbook()
        ws = wb.active

        for entry in data:
            ws.append(entry)

        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

        notSaved = True

        while notSaved:
            try:
                wb.save(outputFile)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message="Παρακαλώ κλείστε το αρχείο '{}' ώστε να ολοκληρωθεί η αποθήκευση.".format(
                                outputFile))
            else:
                notSaved = False

    def run(self):
        self.btnRun.configure(state='disabled')

        inputFile = self.inputFilename.get()
        outputFile = inputFile.replace(".xlsx", "_uppercase.xlsx")
        self.saveFile(self.data, outputFile)

        showinfo(title='Ολοκλήρωση Εκτέλεσης',
                 message=f'Η μετατροπή του κειμένου του αρχείου xlsx σε κεφαλαία ολοκληρώθηκε.')

    def create_widgets(self):
        self.fData = Frame(self.window)

        self.lInputFile = Label(self.fData, text="Αρχείο xlsx:")
        self.lInputFile.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.inputFilename = StringVar()
        self.inputFilename.set('')
        self.ntrInputFilename = Entry(self.fData, width=128, state='readonly', textvariable=self.inputFilename)
        self.ntrInputFilename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenInputFile = Button(self.fData, text="Επιλέξτε αρχείο...", command=self.getInputFilename)
        self.btnOpenInputFile.grid(column=2, row=0, padx=10, pady=10)

        self.btnRun = Button(self.fData, text="Εκτέλεση", command=self.run, state='disabled')
        self.btnRun.grid(column=1, row=10, padx=10, pady=10)

        self.fData.pack()


gui = GUI()
gui.window.mainloop()