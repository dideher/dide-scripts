from openpyxl import load_workbook
import re


def parseXlsxData(fileName):
    workbook = load_workbook(filename=fileName)
    sheet = workbook.active

    data = list()

    for row in sheet.iter_rows():
        entry = list()
        for cell in row:
            if cell.value is None:
                entry.append("")
            else:
                text1 = (str(cell.value).upper().replace(".", ". ").replace(" .", ". ").replace("Ά", "Α").replace("Έ", "Ε")
						.replace("Ή", "Η").replace("Ί", "Ι").replace("Ϊ́", "Ϊ").replace("Ύ", "Υ").replace("Ϋ́", "Ϋ")
						.replace("Ό", "Ο").replace("Ώ", "Ω").strip())

                text2 = re.sub(r'([ ]+)', r' ', text1)
                entry.append(re.sub(r'([0-9]+)Ο', r'\1ο', text2))

        data.append(entry)

    return data