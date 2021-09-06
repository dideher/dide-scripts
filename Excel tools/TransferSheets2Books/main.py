from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from openpyxl import *
from openpyxl.utils import get_column_letter
import os


class GUI():
    def __init__(self):
        self.window = Tk()

        self.window.title("Διαχωρισμός αρχείου excel σε πολλά")
        self.window.resizable(False, False)
        self.create_widgets()

    def getOutputDirName(self):
        dName = filedialog.askdirectory(initialdir="./data/", title="Επιλέξτε τον φάκελο που θα αποθηκευτούν τα αρχεία")

        if dName == "":
            return

        self.outputDirName.set(dName)
        self.btnRun.configure(state='normal')

    def getDataFilename(self):
        fName = filedialog.askopenfilename(initialdir="./data/",
                                           title="Επιλέξτε το αρχείο excel",
                                           filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if fName == "":
            return

        self.dataFilename.set(fName)
        self.ntrOutputDirName.configure(state='normal')
        self.btnOpenOutputDir.configure(state='normal')

    def run(self):
        self.btnRun.configure(state='disabled')
        file = self.dataFilename.get()

        workbook = load_workbook(filename=file)
        sheets_count = len(workbook.worksheets)
        self.pbProgress['maximum'] = sheets_count
        self.pbProgress['value'] = 0
        self.pbProgress.update()

        count = 0
        for sheet in workbook.worksheets:
            self.lProgress.configure(text=f'Εκτέλεση σε εξέλιξη... ({count + 1}/{sheets_count})')
            self.pbProgress['value'] = count + 1
            self.pbProgress.update()

            out_wb = Workbook()
            out_ws = out_wb.active
            out_ws.title = sheet.title

            mr = sheet.max_row
            mc = sheet.max_column

            # copying the cell values from source
            # excel file to destination excel file
            for i in range(1, mr + 1):
                for j in range(1, mc + 1):
                    # reading cell value from source excel file
                    c = sheet.cell(row=i, column=j)

                    # writing the read value to destination excel file
                    out_ws.cell(row=i, column=j).value = c.value

            column_widths = []
            for row in out_ws.iter_rows():
                for i, cell in enumerate(row):
                    try:
                        column_widths[i] = max(column_widths[i], len(str(cell.value)))
                    except IndexError:
                        column_widths.append(len(str(cell.value)))

            for i, column_width in enumerate(column_widths):
                out_ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

            outputFile = f"{out_ws.title}.xlsx"

            self.safe_save(out_wb, outputFile)
            count += 1

        showinfo(title="Ολοκλήρωση εκτέλεσης",
                 message=f'Ο διαχωρισμός του αρχείου excel ολοκληρώθηκε.')



    def safe_save(self, wb, outFile):
        notSaved = True

        outputFile = os.path.join(self.outputDirName.get(), outFile)
        while notSaved:
            try:
                wb.save(outputFile)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message="Παρακαλώ κλείστε το αρχείο '{}' ώστε να ολοκληρωθεί η αποθήκευση.".format(
                                outputFile))
            else:
                notSaved = False


    def create_widgets(self):
        self.fData = Frame(self.window)

        self.lData = Label(self.fData, text="Αρχείο:")
        self.lData.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.dataFilename = StringVar()
        self.ntrDataFilename = Entry(self.fData, width=128, state='readonly', textvariable=self.dataFilename)
        self.ntrDataFilename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenData = Button(self.fData, text="Επιλέξτε αρχείο...", command=self.getDataFilename)
        self.btnOpenData.grid(column=2, row=0, padx=10, pady=10)

        self.lOutputDirName = Label(self.fData, text="Φάκελος για αποθήκευση των αρχείων:")
        self.lOutputDirName.grid(column=0, row=1, padx=10, pady=10, sticky=E)

        self.outputDirName = StringVar()
        self.ntrOutputDirName = Entry(self.fData, width=128, state='disabled', textvariable=self.outputDirName)
        self.ntrOutputDirName.grid(column=1, row=1, padx=10, pady=10, sticky=W)

        self.btnOpenOutputDir = Button(self.fData, text="Επιλέξτε φάκελο...", command=self.getOutputDirName,
                                       state='disabled')
        self.btnOpenOutputDir.grid(column=2, row=1, padx=10, pady=10)

        self.lProgress = Label(self.fData, text="Αναμονή για εκτέλεση ...")
        self.lProgress.grid(column=0, row=2, columnspan=3, padx=10, pady=10)

        self.pbProgress = Progressbar(self.fData, orient='horizontal', length=400, mode='determinate')
        self.pbProgress.grid(column=0, row=3, columnspan=3, padx=10, pady=10)

        self.btnRun = Button(self.fData, text="Εκτέλεση", command=self.run, state='disabled')
        self.btnRun.grid(column=0, row=10, columnspan=3, padx=10, pady=10)

        self.fData.pack()


gui = GUI()
gui.window.mainloop()
