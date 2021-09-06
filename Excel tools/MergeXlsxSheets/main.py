from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from openpyxl import *
from openpyxl.utils import get_column_letter


class GUI():
    def __init__(self):
        self.window = Tk()

        self.window.title("Συνένωση φύλλων excel σε ένα")
        self.window.resizable(False, False)
        self.create_widgets()


    def getDataFilename(self):
        fName = filedialog.askopenfilename(initialdir="./data/",
                                           title="Επιλέξτε το αρχείο excel",
                                           filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if fName == "":
            return

        self.dataFilename.set(fName)
        self.btnRun.configure(state='normal')


    def is_float(self, s):
        return s.replace('.', '', 1).isdecimal()


    def clean_spaces(self, text):
        while "  " in text:
            text = text.replace("  ", " ")

        return text


    def run(self):
        self.btnRun.configure(state='disabled')
        file = self.dataFilename.get()
        workbook = load_workbook(filename=file)

        header = list()
        data = list()

        for sheet in workbook.worksheets:
            header.clear()
            skipHeader = True

            for row in sheet.iter_rows():
                if self.headerChecked.get() and skipHeader:
                    for cell in row:
                        if cell.value is None:
                            text = ""
                        else:
                            text = self.clean_spaces(str(cell.value))
                        header.append(text)

                    if self.sheetNameInColChecked.get():
                        header.append("ΑΠΟ ΦΥΛΛΟ")

                    skipHeader = False
                    continue

                entry = list()
                for cell in row:
                    if cell.value is None:
                        text = ""
                    else:
                        text = self.clean_spaces(str(cell.value))

                    if self.is_float(text):
                        text = text.replace(".", ",")
                    entry.append(text)

                if self.sheetNameInColChecked.get():
                    entry.append(sheet.title)

                data.append(entry)

        wb = Workbook()
        ws = wb.active

        if self.headerChecked.get():
            ws.append(header)

        for entry in data:
            ws.append(entry)

        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

        outputFile = file.replace(".xlsx", "_output.xlsx")

        notSaved = True

        while notSaved:
            try:
                wb.save(outputFile)
            except:
                showwarning(title="Αρχείο σε χρήση...", message="Παρακαλώ κλείστε το αρχείο '{}' ώστε να ολοκληρωθεί η αποθήκευση.".format(outputFile))
            else:
                notSaved = False

        showinfo(title="Αρχείο εξόδου",
                    message="Η συνένωση έχει αποθηκευτεί στο αρχείο: " + outputFile)
        self.btnRun.configure(state='normal')


    def create_widgets(self):
        self.fData = Frame(self.window)

        self.lData = Label(self.fData, text="Αρχείο:")
        self.lData.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.dataFilename = StringVar()
        self.ntrDataFilename = Entry(self.fData, width=128, state='readonly', textvariable=self.dataFilename)
        self.ntrDataFilename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenData = Button(self.fData, text="Επιλέξτε αρχείο...", command=self.getDataFilename)
        self.btnOpenData.grid(column=2, row=0, padx=10, pady=10)

        self.headerChecked = BooleanVar()
        self.ckbHeaderChecked = Checkbutton(self.fData, text="Τα φύλλα περιέχουν επικεφαλίδες", variable=self.headerChecked)
        self.ckbHeaderChecked.grid(column=1, row=1, padx=10, pady=10)

        self.sheetNameInColChecked = BooleanVar()
        self.ckbSheetNameInColChecked = Checkbutton(self.fData, text="Δημιουργία κελιού με το όνομα του φύλλου προέλευσης", variable=self.sheetNameInColChecked)
        self.ckbSheetNameInColChecked.grid(column=1, row=2, padx=10, pady=10)

        self.btnRun = Button(self.fData, text="Εκτέλεση συνένωσης", command=self.run, state='disabled')
        self.btnRun.grid(column=1, row=10, padx=10, pady=10)

        self.fData.pack()


gui = GUI()
gui.window.mainloop()
