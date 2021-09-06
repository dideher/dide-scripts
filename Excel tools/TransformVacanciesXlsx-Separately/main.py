from tkinter import *
from tkinter import filedialog
from tkinter.ttk import *
from tkinter.messagebox import showwarning, showinfo
from openpyxl import *
from openpyxl.utils import get_column_letter
import collections
import os


class GUI():
    def __init__(self):
        self.window = Tk()

        self.window.title("Μετασχηματισμός αρχείου με Κενά/Πλεονάσματα")
        self.window.resizable(False, False)
        self.create_widgets()


    def is_float(self, s):
        return s.replace('.', '', 1).isdecimal()


    def parseXlsxData(self):
        workbook = load_workbook(filename=self.dataFilename.get())
        sheet = workbook.active

        self.xlsxData = list()

        for row in sheet.iter_rows():
            entry = list()
            for i, cell in enumerate(row):
                if i > 4:
                    break

                if cell.value is None:
                    text = ""
                else:
                    text = str(cell.value).strip()

                if self.is_float(text):
                    text = text.replace(".", ",")

                entry.append(text)

            self.xlsxData.append(entry)


    def splitData(self, data, column):
        dataDict = dict()

        for r in data:
            temp = r.copy()

            key = r[column]

            if key == "":
                key = "_EMPTY CELLS_"

            if key not in dataDict:
                dataDict[key] = list()
            dataDict[key].append(temp)

        return dataDict


    def getVacanciesSurplus(self):
        self.vacanciesSurplus = self.splitData(self.xlsxData[1:], 2)


    def getVacanciesSurplusTypes(self):
        self.vacanciesSurplusTypes = dict()

        for key in self.vacanciesSurplus:
            temp = self.splitData(self.vacanciesSurplus[key], 3)
            self.vacanciesSurplusTypes[key] = temp

    def getSpecialties(self, data):
        dataDict = dict()

        for entry in data:
            key = entry[1]
            if key not in dataDict:
                dataDict[key] = int(entry[4])
            else:
                dataDict[key] += int(entry[4])

        return dataDict


    def getSchools(self, data, specialties):
        dataDict = dict()

        for entry in data:
            key = entry[0]
            if key not in dataDict:
                dataDict[key] = {k: 0 for k in specialties}

        return dataDict


    def transform(self, data):
        outData = list()
        specialties = collections.OrderedDict(sorted(self.getSpecialties(data).items()))
        schools = collections.OrderedDict(sorted(self.getSchools(data, specialties).items()))

        header = list()
        header.append('Σχολείο')
        for spc in specialties:
            header.append(spc)
        outData.append(header)

        for entry in data:
            schools[entry[0]][entry[1]] = int(entry[4])

        for sch in schools:
            temp = list()
            temp.append(sch)
            for spc in schools[sch]:
                if schools[sch][spc] == 0:
                    temp.append("")
                else:
                    temp.append(schools[sch][spc])

            outData.append(temp)

        sums = list()
        sums.append('Σύνολα')
        for spc in specialties:
            sums.append(specialties[spc])
        outData.append(sums)

        return outData


    def saveFiles(self):
        for entry in self.vacanciesSurplusTypes:
            data = self.vacanciesSurplusTypes[entry]
            for key in data:
                wb = Workbook()
                ws = wb.active

                outData = self.transform(data[key])
                for row in outData:
                    ws.append(row)

                column_widths = []
                for row in ws.iter_rows():
                    for i, cell in enumerate(row):
                        try:
                            column_widths[i] = max(column_widths[i], len(str(cell.value)))
                        except IndexError:
                            column_widths.append(len(str(cell.value)))

                for i, column_width in enumerate(column_widths):
                    ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

                outFile = f'{entry} - '
                outFile += (key.replace("<", "_").replace(">", "_").replace(":", "_").replace("\"", "_").replace("/", "_")
                           .replace("\\", "_").replace("|", "_").replace("?", "_").replace("*", "_"))
                outFile += ".xlsx"

                notSaved = True

                outputFile = os.path.join(self.outputDirName.get(), outFile)
                while notSaved:
                    try:
                        wb.save(outputFile)
                    except:
                        showwarning(title="Αρχείο σε χρήση...",
                                    message="Παρακαλώ κλείστε το αρχείο '{}' ώστε να ολοκληρωθεί η αποθήκευση.".format(
                                        outputFile))
                    else:
                        notSaved = False


    def getOutputDirName(self):
        dName = filedialog.askdirectory(initialdir="./data/", title="Επιλέξτε τον φάκελο που θα αποθηκευτούν τα αρχεία")

        if dName == "":
            return

        self.outputDirName.set(dName)
        self.btnRun.configure(state='normal')


    def getDataFilename(self):
        fName = filedialog.askopenfilename(initialdir="./data/",
                                           title="Επιλέξτε το αρχείο excel",
                                           filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if fName == "":
            return

        self.dataFilename.set(fName)
        self.parseXlsxData()
        self.ntrOutputDirName.configure(state='readonly')
        self.btnOpenOutputDir.configure(state='normal')

    def run(self):
        self.parseXlsxData()
        self.getVacanciesSurplus()
        self.getVacanciesSurplusTypes()
        self.saveFiles()
        showinfo(title="Ολοκλήρωση εκτέλεσης", message="Ο μετασχηματισμός ολοκληρώθηκε.")


    def create_widgets(self):
        self.fData = Frame(self.window)

        self.lData = Label(self.fData, text="Αρχείο:")
        self.lData.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.dataFilename = StringVar()
        self.ntrDataFilename = Entry(self.fData, width=128, state='readonly', textvariable=self.dataFilename)
        self.ntrDataFilename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenData = Button(self.fData, text="Επιλέξτε αρχείο...", command=self.getDataFilename)
        self.btnOpenData.grid(column=2, row=0, padx=10, pady=10)

        self.lOutputDirName = Label(self.fData, text="Φάκελος για αποθήκευση των αρχείων:")
        self.lOutputDirName.grid(column=0, row=2, padx=10, pady=10, sticky=E)

        self.outputDirName = StringVar()
        self.ntrOutputDirName = Entry(self.fData, width=128, state='disabled', textvariable=self.outputDirName)
        self.ntrOutputDirName.grid(column=1, row=2, padx=10, pady=10, sticky=W)

        self.btnOpenOutputDir = Button(self.fData, text="Επιλέξτε φάκελο...", command=self.getOutputDirName, state='disabled')
        self.btnOpenOutputDir.grid(column=2, row=2, padx=10, pady=10)

        self.btnRun = Button(self.fData, text="Εκτέλεση μετασχηματισμού", command=self.run, state='disabled')
        self.btnRun.grid(column=1, row=10, padx=10, pady=10)

        self.fData.pack()


gui = GUI()
gui.window.mainloop()
