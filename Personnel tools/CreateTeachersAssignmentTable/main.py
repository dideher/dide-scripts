from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from openpyxl import *
import docx


class GUI:
    def __init__(self):
        self.window = Tk()

        self.window.title("Δημιουργία πίνακα τοποθετήσεων")
        self.window.resizable(False, False)

        self.f_data = Frame(self.window)

        self.xlsx_data = list()
        self.data = dict()

        self.create_widgets()

    def parse_xlsx_data(self, file):
        wb = load_workbook(filename=file)
        sheet = wb.active

        for row in sheet.iter_rows():
            entry = list()
            for cell in row:
                if cell.value is None:
                    text = ""
                else:
                    text = str(cell.value).strip()

                entry.append(text)

            self.xlsx_data.append(entry)

    def check_for_errors(self):
        errors = ""
        for i, xlsx_entry in enumerate(self.xlsx_data):
            if xlsx_entry[1] == "":
                errors += f'Γραμμή {i + 1:3}: Το πεδίο "ΕΠΙΘΕΤΟ" είναι κενό.\n'
            if xlsx_entry[2] == "":
                errors += f'Γραμμή {i + 1:3}: Το πεδίο "ΟΝΟΜΑ" είναι κενό.\n'
            if xlsx_entry[3] == "":
                errors += f'Γραμμή {i + 1:3}: Το πεδίο "ΠΑΤΡΩΝΥΜΟ" είναι κενό.\n'
            if xlsx_entry[5] == "":
                errors += f'Γραμμή {i + 1:3}: Το πεδίο "ΕΙΔΙΚΟΚΟΤΗΤΑ" είναι κενό.\n'
            if xlsx_entry[9] == "":
                errors += f'Γραμμή {i + 1:3}: Το πεδίο "ΣΧ. ΑΝΑΛΗΨΗΣ" είναι κενό.\n'
            if xlsx_entry[11] == "":
                errors += f'Γραμμή {i + 1:3}: Το πεδίο "ΣΧΟΛΕΙΟ" είναι κενό.\n'
            if xlsx_entry[12] == "":
                errors += f'Γραμμή {i + 1:3}: Το πεδίο "ΩΡΕΣ" είναι κενό.\n'

        for i, xlsx_entry in enumerate(self.xlsx_data[1:]):
            if xlsx_entry[9] == 'ΝΑΙ':
                xlsx_entry[9] = 'NAI'
                print(f'Γραμμή {i + 2:3}: Μετατροπή των ελληνικών χαρακτήρων σε αγγλικούς.')

        if errors != "":
            print(20 * '-', ' ΛΑΘΗ ', 20 * '-')
            print(errors)
            return True

        return False

    def extract_data(self):
        for xlsx_entry in self.xlsx_data[1:]:
            teacher_id = xlsx_entry[1] + xlsx_entry[2] + xlsx_entry[3] + xlsx_entry[5]

            if teacher_id not in self.data:
                self.data[teacher_id] = list()

            self.data[teacher_id].append([f'{xlsx_entry[1]} {xlsx_entry[2]}', xlsx_entry[3], xlsx_entry[5],
                                          xlsx_entry[11], xlsx_entry[12]])

    def create_rows(self):
        doc = docx.Document()

        count_lines = 0
        for entry in self.xlsx_data[1:]:
            if entry[9] != 'NAI':
                count_lines += 1

        for entry in self.data:
            if len(self.data[entry]) == 1:
                count_lines += 1

        table = doc.add_table(rows=count_lines + 1, cols=8)

        table.cell(0, 0).text = 'Α/Α'
        table.cell(0, 1).text = 'ΟΝΟΜΑΤΕΠΩΝΥΜΟ'
        table.cell(0, 2).text = 'ΠΑΤΡΩΝΥΜΟ'
        table.cell(0, 3).text = 'ΕΙΔΙΚΟΤΗΤΑ'
        table.cell(0, 4).text = 'ΣΧΟΛΕΙΟ ΤΟΠΟΘΕΤΗΣΗΣ'
        table.cell(0, 5).text = 'ΩΡΕΣ ΣΧΟΛΕΙΟΥ ΤΟΠΟΘΕΤΗΣΗΣ'
        table.cell(0, 6).text = 'ΣΧΟΛΕΙΟ/Α ΔΙΑΘΕΣΗΣ'
        table.cell(0, 7).text = 'ΩΡΕΣ ΣΧΟΛΕΙΟΥ/ΕΙΩΝ ΔΙΑΘΕΣΗΣ'

        i = 0
        end_row = 0
        for teacher_id in self.data:
            i += 1

            count = len(self.data[teacher_id]) - 1

            start_row = end_row + 1
            if count == 0:
                end_row = start_row
            else:
                end_row = start_row + count - 1

            print(f'AA: {i}, sr: {start_row}, er: {end_row}')
            if start_row != end_row:
                self.merge_cells(table, start_row, end_row, 0, str(i))
                self.merge_cells(table, start_row, end_row, 1, self.data[teacher_id][0][0])
                self.merge_cells(table, start_row, end_row, 2, self.data[teacher_id][0][1])
                self.merge_cells(table, start_row, end_row, 3, self.data[teacher_id][0][2])
                self.merge_cells(table, start_row, end_row, 4, self.data[teacher_id][0][3])
                self.merge_cells(table, start_row, end_row, 5, self.data[teacher_id][0][4])

                for j, entry in enumerate(self.data[teacher_id][1:]):
                    table.cell(start_row + j, 6).text = entry[3]
                    table.cell(start_row + j, 7).text = entry[4]
            else:
                table.cell(start_row, 0).text = str(i)
                table.cell(start_row, 1).text = self.data[teacher_id][0][0]
                table.cell(start_row, 2).text = self.data[teacher_id][0][1]
                table.cell(start_row, 3).text = self.data[teacher_id][0][2]
                table.cell(start_row, 4).text = self.data[teacher_id][0][3]
                table.cell(start_row, 5).text = self.data[teacher_id][0][4]

                if count == 1:
                    table.cell(start_row, 6).text = self.data[teacher_id][1][3]
                    table.cell(start_row, 7).text = self.data[teacher_id][1][4]

        not_saved = True

        while not_saved:
            try:
                doc.save('output.docx')
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message="Παρακαλώ κλείστε το αρχείο 'output.docx' ώστε να ολοκληρωθεί η αποθήκευση.")
            else:
                not_saved = False

    def merge_cells(self, table, start_row, end_row, col, text):
        a = table.cell(start_row, col)
        b = table.cell(end_row, col)
        m = a.merge(b)

        m.text = text

    def get_input_filename(self):
        f_name = filedialog.askopenfilename(initialdir=".", title="Επιλέξτε το αρχείο xlsx",
                                            filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if f_name == "":
            return

        self.btn_open_input_file.configure(state='disabled')
        self.input_filename.set(f_name)

        self.parse_xlsx_data(f_name)

        self.btn_run.configure(state='normal')

    def run(self):
        self.btn_run.configure(state='disabled')

        if self.check_for_errors():
            showwarning(title='Λάθη στο αρχείο εισόδου',
                        message=f'Για να δημιουργηθεί ο πίνακας είναι απαραίτητη η διόρθωση του αρχείου εισόδου.')
            return

        self.extract_data()
        self.create_rows()

        showinfo(title='Ολοκλήρωση Εκτέλεσης',
                 message=f'Η δημιουργία του πίνακα τοποθετήσεων ολοκληρώθηκε.')

        self.window.destroy()

    def create_widgets(self):
        self.l_input_file = Label(self.f_data, text="Αρχείο xlsx:")
        self.l_input_file.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.input_filename = StringVar()
        self.input_filename.set('')
        self.ntr_input_filename = Entry(self.f_data, width=128, state='readonly', textvariable=self.input_filename)
        self.ntr_input_filename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btn_open_input_file = Button(self.f_data, text="Επιλέξτε αρχείο...", command=self.get_input_filename)
        self.btn_open_input_file.grid(column=2, row=0, padx=10, pady=10)

        self.btn_run = Button(self.f_data, text="Εκτέλεση", command=self.run, state='disabled')
        self.btn_run.grid(column=1, row=10, padx=10, pady=10)

        self.f_data.pack()


gui = GUI()
gui.window.mainloop()
