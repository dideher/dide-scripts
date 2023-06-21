from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import date
import sqlite3


class GUI:
    def __init__(self):
        self.window = Tk()

        self.window.title("Δημιουργία αρχείου xlsx για Εργάνη")
        self.window.resizable(False, False)
        self.create_widgets()

    def create_ecd_list(self):
        self.cursor.execute("SELECT DISTINCT end_date FROM teachers")
        q = self.cursor.fetchall()

        ecd_temp_list = list()
        for item in q:
            day, month, year = item[0].split('/')
            if year == str(date.today().year):
                ecd_temp_list.append(f'{year}/{month}/{day}')

        ecd_temp_list.sort(reverse=True)

        ecd_list = list()
        for item in ecd_temp_list:
            year, month, day = item.split('/')
            ecd_list.append(f'{day}/{month}/{year}')

        return ecd_list

    def get_payroll(self):
        f_name = filedialog.askopenfilename(initialdir="./data", title="Επιλέξτε το αρχείο xlsx",
                                            filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if f_name == "":
            return

        self.payroll = self.parse_xlsx_data(f_name)

        header = [
            "A/A",
            "ΕΙΔΟΣ",
            "ΑΡ. ΜΗΤΡΩΟΥ",
            "ΕΠΩΝΥΜΟ",
            "ΟΝΟΜΑ",
            "ΠΑΤΡΩΝΥΜΟ",
            "AΦΜ",
            "AMKA",
            "AMIKA",
            "ΗΜΕΡΕΣ ΕΡΓΑΣΙΑΣ",
            "ΗΜΕΡΕΣ ΑΣΦΑΛΙΣΗΣ",
            "ΣΧ. ΟΡΓ. ΘΕΣΗΣ",
            "ΣΧ. ΥΠΗΡΕΣΙΑΣ",
            "ΙΒΑΝ",
            "ΤΡΑΠΕΖΑ",
            "ΜΚ",
            "ΒΑΘΜΟΣ",
            "ΚΛΑΔΟΣ",
            "ΚΑΤΗΓΟΡΙΑ",
            "ΕΙΔΙΚΟΤΗΤΑ Ν.4521",
            "ΠΛΕΟΝΑΖΟΝ ΧΡΟΝΟΣ",
            "BM",
            "ΥΠΕΡΒΑΛΟΥΣΑ / ΠΡΟΣ. ΔΙΑΦΟΡΑ",
            "ΕΠ. ΔΙΔ. ΠΡΟΕΤ.",
            "ΚΙΝ. ΑΠΟΔ.",
            "ΟΙΚΟΓ. ΕΠ.",
            "ΠΡΟΣ. ΔΙΑΦ.",
            "ΧΡΟΝOΕΠΙΔOMA",
            "ΕΠ. ΘΕΣΗΣ",
            "ΕΠ. ΕΞΩΔΙΔ.",
            "ΕΠ. ΣΠΟΥΔ.",
            "ΕΠ. ΠΑΡΑΠΛΗΓ.",
            "ΕΠ. ΠΡΟΒΛ.",
            "ΕΠ. ΕΙΔ. ΣΥΝΘ.",
            "ΕΠ. ΕΙΔ. ΘΕΣΗΣ",
            "ΕΠ. ΑΝΘΥΓ.",
            "ΕΠ. ΜΕΙΟΝ.",
            "ΕΠ. ΕΣΔΔ",
            "ΟΔΟΙΠΟΡΙΚΑ",
            "ΕΠ. ΠΛΗΡΟΦΟΡΙΚΗΣ",
            "ΕΙΔ. ΕΡΕΥ. ΕΠ.",
            "ΕΞ. ΠΑΡΑΣΤ",
            "ΠΑΓ.ΑΠΟΖΗΜ.",
            "ΕΞ. ΚΙΝΗΣΗΣ",
            "ΠΑΓ.ΑΠΟΖΗΜ. ΒΙΒΛ.",
            "ΑΠΟΖΗΜΙΩΣΗ ΒΟΥΛΕΤΙΚΩΝ ΓΡΑΦΕΙΩΝ",
            "ΑΠΟΖΗΜΙΩΣΗ ΥΠΕΡ. ΑΠΑΣΧ. ΔΙΟΙΚΗΤ.",
            "ΥΠΕΡΩΡΙΕΣ ΚΑΤΑ ΤΙΣ ΕΞΑΙΡ. ΜΕΡΕΣ & ΝΥΧΤΕΣ ΩΡΕΣ",
            "ΑΠΟΖΗΜΙΩΣΗ ΜΗΝΑ ΩΡΟΜΙΣΘΙΩΝ",
            "ΑΠΟΖΗΜΙΩΣΗ ΤΡΙΜΗΝΟΥ ΩΡΟΜΙΣΘΙΩΝ",
            "ΑΠΟΖΗΜΙΩΣΗ ΑΔΕΙΑΣ ΩΡΟΜΙΣΘΙΩΝ",
            "ΑΜΟΙΒΕΣ ΕΠΙΤΗΡ & Β. ΠΡΟΣΩΠ",
            "ΑΜΟΙΒΕΣ Σ/Κ",
            "ΑΜΟΙΒ.ΒΑΘΜ",
            "ΑΜΟΙΒ. ΕΠΙΤΡ",
            "ΜΗ ΧΡΗΣH ΘΕΡΙΝΗΣ ΑΔΕΙΑΣ",
            "ΣΧΟΛΙΚΟΙ ΑΓΩΝΕΣ ΤΡΕΧΟΝΤΟΣ ΕΤΟΥΣ",
            "ΣΧΟΛΙΚΟΙ ΑΓΩΝΕΣ ΠΑΡΕΛΘΟΝΤΟΣ ΕΤΟΥΣ",
            "ΠΑΝΕΛΛΑΔΙΚΕΣ ΠΑΡΕΛΘΟΝΤΟΣ ΕΤΟΥΣ",
            "ΕΞΟΔΑ ΠΑΡΑΣΤΑΣΗΣ (259)",
            "ΝΥΧΤΕΡΙΝΕΣ ΥΠΕΡΩΡΙΕΣ",
            "ΔΙΚΑΣΤΙΚΕΣ ΑΠΟΦΑΣΕΙΣ",
            "ΕΚΤΑΚΤΕΣ ΑΝΑΔΡΟΜΙΚΕΣ ΕΙΣΦΟΡΕΣ",
            "ΑΜΟΙΒΕΣ ΕΠΙΤΗΡ-ΒΟΗΘ ΠΡΟΣΩΠ.(ΑΛΕ:2130205001)",
            "ΑΜΟΙΒΕΣ ΕΠΙΤΗΡ-ΒΟΗΘ ΠΡΟΣΩΠ.(ΑΛΕ:2130289001)",
            "ΑΜΟΙΒΕΣ ΒΑΘΜΟΛΟΓΗΤΩΝ(ΑΛΕ:2130205001)",
            "ΑΜΟΙΒΕΣ ΒΑΘΜΟΛΟΓΗΤΩΝ(ΑΛΕ:2130289001)",
            "ΑΜΟΙΒΕΣ ΕΠΙΤΡ-ΠΑΝ-ΚΠΓ (ΑΛΕ:2130205001)",
            "ΑΜΟΙΒΕΣ ΕΠΙΤΡ-ΠΑΝ-ΚΠΓ (ΑΛΕ:2130289001)",
            "ΜΗ ΧΡΗΣΗ ΘΕΡΙΝΗΣ ΑΔΕΙΑΣ (ΑΛΕ:2130289001)",
            "ΕΞΟΔΑ ΓΙΑ ΛΟΙΠΕΣ ΥΠΗΡ. ΙΔΙΩΤΩΝ(ΑΛΕ:2420989001)",
            "ΕΠΙΜΙΣΘΙΟ ΕΞΩΤΕΡΙΚΟΥ",
            "ΣΥΝ. ΑΚ. ΑΠΟΔ.",
            "ΤΕΑΔΥ",
            "ΤΕΑΔΥ ΠΡ. ΑΜ. ΕΡΓ",
            "ΤΣΜΕΔΕ ΚΣ ΕΡΓ",
            "ΤΣΜΕΔΕ ΕΙΔ.ΠΡΟΣ. ΕΡΓ",
            "ΤΣΜΕΔΕ ΕΦΑΠΑΞ. ΕΡΓ",
            "ΤΣΜΕΔΕ ΕΠΙΚ. ΕΡΓ",
            "ΤΕΑΧ",
            "ΚΥΤ",
            "ΤΣΑΥ",
            "ΙΚΑ",
            "ΝΑΤ",
            "ΤΕΚΑ ΕΡΓ",
            "ΤΣΚΥ",
            "ΤΕΑΔΥ ΤΕΑΜ",
            "ΤΕΑΔΥ ΤΑΔΚΥ",
            "ΤΥΔΚΥ",
            "ΤΕΑΠΠΕΡΤ",
            "ΤΑΜΠΥΕΘΑ",
            "ΤΑΠΙΤ",
            "ΤΠΔΥ-ΤΠΠΕΒΕΒΕΚ ΕΡΓ",
            "ΤΕΑΕΙΓΕ",
            "ΤΑΥΤΕΚΩ (πρ. Τ.Α.Π. ΟΤΕ Ασθένειας)",
            "ΤΑΥΤΕΚΩ-Τ.Π.Πρ.ΟΣΕ",
            "ΤΑΥΤΕΚΩ ΠΡ. ΟΤΕ",
            "ΤΤΑΥΤΕΚΩ ΧΡΗΜΑ",
            "ΤΑΥΤΕΚΩ ΕΟΠΠΥ",
            "ΙΚΑ ΕΤΑΜ (Πρ. ΤΑΠΟΤΕ)",
            "Εργ. Εισφ. υπερ ΟΠΑΔ",
            "ΗΣΑΠ ΥΓ.",
            "ΗΣΑΠ ΣΥΝ.",
            "ΤΕΑΠΟΚΑ",
            "ΚΛ. ΕΠΙΚ. ΑΣΦ. ΔΙΚΗΓ.",
            "ΤΑΜΕΙΟ ΝΟΜΙΚΩΝ ΕΜΜΙΣΘΗ ΕΝΤΟΛΗ",
            "ΤΑΜΕΙΟ ΥΓΕΙΑΣ ΔΙΚΗΓ ΑΘ",
            "Πρόσθετο 20% στα τέλη χαρτοσήμου",
            "Λοιπά τέλη χαρτοσήμου",
            "ΕΔΟΕΑΠ ΕΡΓ",
            "ΤΣΠΕΑΘ ΕΡΓ",
            "ΛΑΠΗΕΑΘ ΕΡΓ",
            "ΕΤΕΑΠ ΕΡΓ",
            "ΕΦΚΑ ΕΡΓ",
            "ΣΥΝ. ΑΚ. ΑΠΟΔ. + ΕΡΓΔ. ΕΙΣΦ.",
            "ΦΟΡΟΣ",
            "Ο.Π.Α.Δ",
            "ΜΤΠΥ",
            "ΜΤΠΥ ΝΕΟΔ.",
            "ΤΕΑΔΥ",
            "ΤΕΑΔΥ ΠΡ. ΑΜ.",
            "ΤΕΑΔΥ ΝΕΟΔ.",
            "ΕΚΤ. ΤΕΑΔΥ",
            "ΣΥΝΤΑΞΗ",
            "ΤΣΜΕΔΕ ΚΣ. ΑΣΦ",
            "ΤΣΜΕΔΕ ΕΙΔ.ΠΡΟΣ ΑΣΦ",
            "ΤΣΜΕΔΕ ΕΦΑΠΑΞ. ΑΣΦ",
            "ΤΣΜΕΔΕ ΕΠΙΚ. ΑΣΦ",
            "ΚΥΤ",
            "ΤΠΔΥ",
            "ΤΣΑΥ ΣΥΝΤ",
            "ΤΣΑΥ ΣΤΕΓΗ",
            "ΤΣΑΥ ΠΡΟΝ",
            "ΥΓΕΙΟΝΟΜ ΤΣΑΥ",
            "ΤΕΑΧ",
            "ΙΚΑ",
            "ΕΠΙΔΟΤΗΣΗ IKA",
            "ΤΥΔΚΥ",
            "ΤΠΔΥ Ν103/75",
            "ΤΕΑΔΥ ΤΕΑΜ",
            "ΤΕΑΔΥ ΤΑΔΚΥ",
            "ΝΑΤ",
            "ΤΕΚΑ",
            "ΤΠΔΥ ΤΑΔΚΥ",
            "ΤΠΔΥ-ΤΠΔΚΥ",
            "ΤΑΠΕΜ",
            "ΤΕΑΠΠΕΡΤ",
            "ΤΑΜΠΥΕΘΑ",
            "ΤΑΠΙΤ",
            "ΤΠΔΥ-ΤΠΠΕΒΕΒΕΚ",
            "ΤΠΔΥ-ΤΠΠΕΒΕΒΕΚ (ΕΙΔΙΚΗ ΕΙΣΦΟΡΑ 1%)",
            "ΤΕΑΕΙΓΕ",
            "ΤΑΥΤΕΚΩ (πρ. Τ.Α.Π. ΟΤΕ Ασθένειας)",
            "ΤΑΥΤΕΚΩ ΠΡ.ΟΤΕ",
            "ΤΑΥΤΕΚΩ-Τ.Π.Πρ.ΟΣΕ",
            "ΤΑΥΤΕΚΩ ΧΡΗΜΑ",
            "ΤΑΥΤΕΚΩ ΕΟΠΠΥ",
            "ΙΚΑ ΕΤΑΜ(Πρ. ΤΑΠΟΤΕ)",
            "ΜΠΤΥ 1%",
            "ΜΤΠΥ ΠΡΟΣΘ. ΑΜΟΙΒ.",
            "ΤΑΜΕΙΟ ΝΟΜΙΚΩΝ",
            "ΤΑΜΕΙΟ ΝΟΜΙΚΩΝ ΕΜΜΙΣΘΗ ΕΝΤΟΛΗ",
            "Πρόσθετο 20% στα τέλη χαρτοσήμου",
            "Λοιπά τέλη χαρτοσήμου",
            "ΕΔΟΕΑΠ",
            "ΤΣΠΕΑΘ",
            "ΛΑΠΗΕΑΘ",
            "ΕΤΕΑΠ",
            "ΤΑΧ. ΤΑΜ.",
            "ΤΑΧ. ΤΑΜ. ΠΡ.",
            "ΤΕΑΔΥ",
            "ΜΤΠΥ",
            "ΤΠΔ",
            "ΟΔΔΥ",
            "ΟΑΕΔ",
            "ΔΙΑΔ. ΑΣΦ. ΕΞΑΓ. ΣΤΡΑΤΟΥ",
            "ΕΞΑΓ. ΤΠΔΥ",
            "ΑΧΡ. ΚΑΤ.",
            "ΑΠΕΡΓΙΑ",
            "ΔΙΑΤΡΟΦΗ",
            "ΑΝΤΙΛΟΓΙΣΜΟΣ ΕΦΚΑ",
            "ΕΜΔΥΔΑΣ Ν.ΜΑΓΝΗΣΙΑΣ",
            "ΕΜΔΥΔΑΣ ΔΩΔ/ΣΟΥ",
            "Π.Ο.ΜΗ.Τ.Ε.Δ.Υ.",
            "ΣΥ.ΜΗ.Τ.Ε.Δ.Υ. ΛΑΡΙΣΑΣ-ΜΑΓΝΗΣΙΑΣ",
            "ΣΥ.ΜΗ.Τ.Ε.Δ.Υ. ΚΑΡΔΙΤΣΑΣ-ΤΡΙΚΑΛΩΝ",
            "ΣΥΛΛΟΓΟΣ",
            "ΕΕΧ",
            "ΚΑΤΑΣΧ.",
            "ΚΑΤΑΣΧΕΣΗ ΑΠΟ ΤΡΙΤΟΥΣ",
            "ΠΟ.ΕΜΔΥΔΑΣ",
            "ΣΥΛ.ΕΙΔ.ΤΕΧ.ΘΕΣΣ.",
            "ΣΥΛ.ΔΙΟΙΚ.ΠΡ.ΘΕΣΣ.",
            "ΕΞΑΓ. ΤΕΑΔΥ",
            "ΕΞΑΓ. ΜΤΠΥ",
            "ΑΡΓ. ΜΤΠΥ",
            "ΟΛΜΕ",
            "ΑΔΕΔΥ",
            "ΤΠΔΥ 1%",
            "ΠΥΝΠΔ 1%",
            "ΤΠΔΥ-ΤΠΔΚΥ 1%",
            "Υπέρ ΟΑΕΔ 1%",
            "Υπέρ Ανεργίας 2%",
            "Εισφορά Αλληλεγγύης",
            "ΠΕΙΘΑΡΧΙΚΟ ΠΡΟΣΤΙΜΟ",
            "ΕΥΠΥ ΥΠΑΙΘ ΜΑΚΕΔΟΝΙΑΣ-ΘΡΑΚΗΣ",
            "Σύλ. Μον. Υπαλ.Γεν. Γραμ. Λαϊκής Επιμόρφωσης",
            "ΓΕΩΤ.Ε.Ε.",
            "Σύλ. Υπαλ. Περιφερ. Υπηρ. ΥΠ",
            "ΟΔΕ",
            "ΤΕΛΩΝΕΙΟ ΘΕΣΣ.",
            "ΤΕΛΩΝΕΙΟ ΠΑΤΡΩΝ",
            "ΕΣΗΕΑ",
            "ΣΥΝ. ΚΡΑΤ.",
            "ΕΙΣΦ. ΔΗΜ.",
            "ΠΛΗΡΩΤΕΟ",
            "ΠΛΗΡ. Α' 15ΗΜΕΡΟ",
            "ΠΛΗΡ. Β' 15ΗΜΕΡΟ",
            "ΕΙΔΟΣ",
            "ΑΠΌ",
            "ΜΕΧΡΙ"
        ]

        if self.payroll[3] != header:
            cols = ''
            for item in header:
                cols += f"- {item}\n"
            showwarning(title="Λάθος τύπος αρχείου...",
                        message=f"Το αρχείο πρέπει να έχει τις εξής στήλες:\n{cols}")

            return

        warnings = ''
        errors = ''

        for item in self.payroll[4:]:
            afm = item[6]
            days_of_work = float(item[9].replace(",", "."))
            days_of_soc_security = item[10]
            basic_salary = float(item[21].replace(",", "."))
            family_bonus = float(item[25].replace(",", "."))

            if days_of_work == 30 and days_of_soc_security == '25':
                basic_salary_fix = basic_salary
                family_bonus_fix = family_bonus
            elif days_of_work < 30 and days_of_soc_security == '25':
                basic_salary_fix = round(basic_salary * 30 / days_of_work)
                family_bonus_fix = round(family_bonus * 30 / days_of_work)

                warnings += f"{afm}: [{basic_salary}, {family_bonus}] --> [{basic_salary_fix}, {family_bonus_fix}]\n"
            else:
                errors += f"{afm}: ΗΜΕΡΕΣ ΕΡΓΑΣΙΑΣ ({days_of_work}), ΗΜΕΡΕΣ ΑΣΦΑΛΙΣΗΣ ({days_of_soc_security})\n"

                continue

            self.cursor.execute(f"INSERT INTO payroll VALUES ('{afm}', '{basic_salary_fix}', '{family_bonus_fix}')")
            self.conn.commit()

        if warnings != '':
            print(20 * '-', ' ΄Έγινε αυτόματη διόρθωση για τα ΑΦΜ ', 20 * '-')
            print(warnings)

        if errors != '':
            print(20 * '-', ' Χρειάζονται διόρθωση ', 20 * '-')
            print(errors)

        self.btn_get_payroll.configure(state='disabled')
        self.payroll_filename.set(f_name)

        self.btn_get_teachers.configure(state='normal')

    def get_teachers(self):
        f_name = filedialog.askopenfilename(initialdir="./data", title="Επιλέξτε το αρχείο xlsx",
                                            filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if f_name == "":
            return

        self.teachers = self.parse_xlsx_data(f_name)

        header = [
            "Αρ.Μητρώου",
            "Α.Φ.Μ",
            "Επώνυμο",
            "Όνομα",
            "Πατρώνυμο",
            "Μητρώνυμο",
            "Φύλο",
            "Είδος Υπαλλήλου",
            "Κλάδος Υπαλλ. Όνομα",
            "Κλάδος Υπαλλ. Κατηγ.",
            "Κλάδος Υπαλλ. Περιγρ.",
            "Ημ. Γέννησης",
            "ΜΚ Ν.3205",
            "Ημ. ΜΚ Ν.3205",
            "Βαθμός Ν.3205",
            "Πλεον. Χρόνος",
            "Ασφαλιστικό",
            "ΑΜΚΑ",
            "ΑΜΙΚΑ",
            "ΑΜ ΤΣΜΕΔΕ",
            "Ημ. Εγγρ. ΤΣΜΕΔΕ",
            "Ημ. Ειδ.Προσ. ΤΣΜΕΔΕ",
            "Ασφάλ. Μέλη",
            "Ημ. Λήξης Ανατροφ.",
            "ΑΜΕΑ",
            "Οικογενειακή Κατάσταση",
            "Βάρη",
            "ΦΕΚ Διορισμού",
            "ΔΟΥ",
            "ΙΒΑΝ",
            "ΑΔΤ",
            "Έπώνυμ. Συζύγου",
            "Παρατηρήσεις",
            "Επώνυμ. Εκτ.",
            "Όνομα Εκτ.",
            "Πατρωνυμ. Εκτ.",
            "Μεταπτυχιακό",
            "Ημερ. Αναγν. Μεταπτ.",
            "Διδακτορικό",
            "Ημερ. Αναγν. Διδακτορ.",
            "Δόση Υπερβαλ.",
            "Από - Δόση Υπερβαλ.",
            "Μέχρι - Δόση Υπερβαλ.",
            "Προϋπηρεσία",
            "Ημ/νία Χορήγησης Νέου ΜΚ",
            "Όνομα Κλάδου Ν4521",
            "Ημ. Διορισμού",
            "Προϋπ. πριν 2011",
            "ΜΚ",
            "Από - ΜΚ",
            "Μέχρι - ΜΚ",
            "Βαθμός",
            "Από - Βαθμός",
            "Μέχρι - Βαθμός",
            "Σχέση Εργ.",
            "Από - Σχέση Εργ.",
            "Μέχρι - Σχέση Εργ.",
            "Καταστ. Εργ.",
            "Από - Καταστ. Εργ.",
            "Μέχρι - Καταστ. Εργ.",
            "Ωράριο",
            "Από - Ωράριο",
            "Μέχρι - Ωράριο",
            "Φορέας Μισθοδοσίας",
            "Κατηγορία Φορ. Μισθ.",
            "Από - Φορέας Μισθοδοσίας",
            "Μέχρι - Φορέας Μισθοδοσίας",
            "Φορέας Οργανικής",
            "Κωδ. Φορέα Οργανικής",
            "Φορέας Υπηρεσίας",
            "Κωδ. Φορέα Υπηρεσίας",
            "Φορέας Απόσπασης",
            "Κωδ. Φορέα Απόσπασης",
            "Φορέας Συμπλήρωσης ωραρίου",
            "Κωδ. Φορέα Συμπλήρωσης ωραρίου"
        ]

        if self.teachers[0][:75] != header:
            cols = ''
            for item in header:
                cols += f"- {item}\n"
            showwarning(title="Λάθος τύπος αρχείου...",
                        message=f"Το αρχείο πρέπει να έχει τις εξής στήλες:\n{cols}")

            return

        reject = ''

        for item in self.teachers[1:]:
            afm = item[1]
            last_name = item[2]
            first_name = item[3]
            father_name = item[4]
            mother_name = item[5]
            gender = item[6]
            birthday = item[11]
            family_status = item[25]
            amka = item[17]
            gov_id = item[30]
            specialty = item[45]
            start_date = item[55]
            end_date = item[56]
            work_type = item[54]
            payment_type = item[64]

            if payment_type == 'Χρηματικά Εντάλματα':
                reject += f"{item}\n"

                continue

            self.cursor.execute(
                f"INSERT INTO teachers VALUES ('{afm}', '{last_name}', '{first_name}', '{father_name}', "
                f"'{mother_name}', '{gender}', '{birthday}', '{family_status}', '{amka}', '{gov_id}', '{specialty}', "
                f"'{start_date}', '{end_date}', '{work_type}')")
            self.conn.commit()

        if reject != '':
            print(20 * '-', ' Χρηματικά Εντάλματα ', 20 * '-')
            print(reject)

        self.btn_get_teachers.configure(state='disabled')
        self.teachers_filename.set(f_name)

        self.cb_end_contract_date['values'] = self.create_ecd_list()
        self.cb_end_contract_date.configure(state='readonly')

        self.btn_run.configure(state='normal')

    def parse_xlsx_data(self, file):
        workbook = load_workbook(filename=file)
        sheet = workbook.active

        data = list()

        for row in sheet.iter_rows():
            entry = list()
            for cell in row:
                if cell.value is None:
                    text = ""
                else:
                    text = str(cell.value).strip()

                entry.append(text)

            data.append(entry)

        return data

    def save_file(self, data, output_file):
        wb = Workbook()
        ws = wb.active

        for entry in data:
            ws.append(entry)

        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

        not_saved = True

        while not_saved:
            try:
                wb.save(output_file)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message=f"Παρακαλώ κλείστε το αρχείο '{output_file}' ώστε να ολοκληρωθεί η αποθήκευση.")
            else:
                not_saved = False

    def calc_children(self, fb):
        if fb == 0:
            return 0
        elif fb == 50:
            return 1
        elif fb == 70:
            return 2
        else:
            children = int((fb - 70) / 50) + 2
            return children

    def create_export_data(self):
        if self.cb_end_contract_date.current() == -1:
            self.cursor.execute(f"SELECT * FROM teachers, payroll "
                                f"WHERE payroll.afm = teachers.afm")
        else:
            ecd = self.end_contract_date.get()
            self.cursor.execute(f"SELECT * FROM teachers, payroll "
                                f"WHERE payroll.afm = teachers.afm AND "
                                f"teachers.end_date = '{ecd}'")

        q = self.cursor.fetchall()

        header = ['ΑΦΜ', 'ΕΠΙΘΕΤΟ', 'ΟΝΟΜΑ', 'ΠΑΤΡΩΝΥΜΟ', 'ΜΗΤΡΩΝΥΜΟ', 'ΦΥΛΟ', 'ΗΜΕΡΟΜΗΝΙΑ ΓΕΝΝΗΣΗΣ',
                  'ΟΙΚΟΓΕΝΕΙΑΚΗ ΚΑΤΑΣΤΑΣΗ', 'ΤΕΚΝΑ', 'ΑΜΚΑ', 'ΑΔΤ', 'ΚΛΑΔΟΣ', 'ΚΩΔΙΚΟΣ ΕΙΔΙΚΟΤΗΤΑΣ ΕΦΚΑ',
                  'ΗΜΕΡΟΜΗΝΙΑ ΠΡΟΣΛΗΨΗΣ', 'ΗΜΕΡΟΜΗΝΙΑ ΑΠΟΛΥΣΗΣ', 'ΜΙΣΘΟΣ', 'ΜΟΡΦΩΤΙΚΟ ΕΠΙΠΕΔΟ', 'ΕΙΔΟΣ ΑΠΑΣΧΟΛΗΣΗΣ',
                  'ΣΥΜΒΑΤΙΚΗ ΗΜΕΡΟΜΗΝΙΑ ΛΗΞΗΣ ΣΥΜΒΑΣΗΣ']

        data = list()
        data.append(header)

        for item in q:
            specialty_code = '24'

            afm = item[0]
            last_name = item[1]
            first_name = item[2]
            father_name = item[3]
            mother_name = item[4]

            gender = item[5]
            if gender == 'Άντρας':
                gender = 'ΑΝΤΡΑΣ'
            else:
                gender = 'ΓΥΝΑΙΚΑ'

            birthday = item[6]

            family_status = item[7]
            if family_status == 'Έγγαμος':
                family_status = 'ΕΓΓΑΜΟΣ'
            elif family_status == 'Άγαμος':
                family_status = 'ΑΓΑΜΟΣ'
            else:
                family_status = 'ΔΙΑΖΕΥΓΜΕΝΟΣ'

            amka = item[8]
            gov_id = item[9]
            specialty = item[10]
            start_date = item[11]
            end_date = item[12]
            basic_salary = item[15]
            family_bonus = item[16]
            children = self.calc_children(family_bonus)
            education = 'ΑΕΙ'

            work_type = item[13]
            if work_type == 'Αναπληρωτές':
                work_type = 'ΠΛΗΡΗΣ'
            else:
                work_type = 'ΜΕΡΙΚΗ'

            typical_end_date = end_date

            salary = basic_salary + family_bonus

            entry = [afm, last_name, first_name, father_name, mother_name, gender, birthday, family_status, children,
                     amka, gov_id, specialty, specialty_code, start_date, end_date, f"{salary:.2f}".replace(".", ","),
                     education, work_type, typical_end_date]

            if entry not in data:
                data.append(entry)

        return data

    def run(self):
        self.btn_run.configure(state='disabled')

        output_file = "ergani.xlsx"

        data = self.create_export_data()

        self.save_file(data, output_file)

        showinfo(title='Ολοκλήρωση Εκτέλεσης',
                 message=f'Η δημιουργία του αρχείου xlsx ολοκληρώθηκε.')

        self.window.destroy()

    def create_db(self):
        self.conn = sqlite3.connect('ergani.db')
        self.cursor = self.conn.cursor()

        self.cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = self.cursor.fetchall()

        for tbl in tables:
            if tbl[0] in ['payroll', 'teachers']:
                self.cursor.execute(f"DROP TABLE {tbl[0]}")
        self.conn.commit()

        self.cursor.execute("""CREATE TABLE payroll (
                                afm text,
                                basic_salary float,
                                family_bonus float
                            )""")

        self.cursor.execute("""CREATE TABLE teachers (
                                afm text,
                                last_name text,
                                first_name text,
                                father_name text,
                                mother_name text,
                                gender text,
                                birthday text,
                                family_status text,
                                amka text,
                                gov_id text,
                                specialty text,
                                start_date text,
                                end_date text,
                                work_type text                                                                            
                            )""")

    def create_widgets(self):
        self.f_data = Frame(self.window)

        self.l_payroll = Label(self.f_data, text="Μισθολογικά (xlsx):")
        self.l_payroll.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.payroll_filename = StringVar()
        self.payroll_filename.set('')
        self.ntr_payroll_filename = Entry(self.f_data, width=128, state='readonly', textvariable=self.payroll_filename)
        self.ntr_payroll_filename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btn_get_payroll = Button(self.f_data, text="Επιλέξτε αρχείο...", command=self.get_payroll, state='normal')
        self.btn_get_payroll.grid(column=2, row=0, padx=10, pady=10)

        self.l_teachers = Label(self.f_data, text="Εκπαιδευτικοί (xlsx):")
        self.l_teachers.grid(column=0, row=1, padx=10, pady=10, sticky=E)

        self.teachers_filename = StringVar()
        self.teachers_filename.set('')
        self.ntr_teachers_filename = Entry(self.f_data, width=128, state='readonly',
                                           textvariable=self.teachers_filename)
        self.ntr_teachers_filename.grid(column=1, row=1, padx=10, pady=10, sticky=W)

        self.btn_get_teachers = Button(self.f_data, text="Επιλέξτε αρχείο...", command=self.get_teachers,
                                       state='disabled')
        self.btn_get_teachers.grid(column=2, row=1, padx=10, pady=10)

        self.l_end_contract_date = Label(self.f_data, text="Ημερομηνία Λήξης Σύμβασης:\n(χωρίς επιλογή για όλες)")
        self.l_end_contract_date.grid(column=0, row=2, padx=10, pady=10, sticky=E)

        self.end_contract_date = StringVar()
        self.cb_end_contract_date = Combobox(self.f_data, width=125, textvariable=self.end_contract_date,
                                             state='disabled')
        self.cb_end_contract_date.grid(column=1, row=2, padx=10, pady=10)

        self.btn_run = Button(self.f_data, text="Εκτέλεση", command=self.run, state='disabled')
        self.btn_run.grid(column=1, row=10, padx=10, pady=10)

        self.create_db()

        self.f_data.pack()


gui = GUI()
gui.window.mainloop()
