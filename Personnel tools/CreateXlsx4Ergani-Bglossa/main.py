from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import csv
import sqlite3


class GUI:
    def __init__(self):
        self.window = Tk()

        self.window.title("Δημιουργία αρχείου xlsx για Εργάνη")
        self.window.resizable(False, False)
        self.create_widgets()

    def create_ecd_list(self):
        self.cursor.execute("SELECT DISTINCT \"Λήξη σύμβασης\" FROM contracts")
        q = self.cursor.fetchall()

        ecd_list = list()
        for item in q:
            ecd_list.append(item[0])

        ecd_list.sort()

        return ecd_list

    def get_contracts(self):
        f_name = filedialog.askopenfilename(initialdir="./data", title="Επιλέξτε το αρχείο csv",
                                            filetypes=(("csv files", "*.csv"), ("all files", "*.*")))

        if f_name == "":
            return

        self.contracts = self.parse_csv_data(f_name)

        header = ['A/A', 'Επίθετο', 'Όνομα', 'Όνομα πατέρα', 'ΑΦΜ', 'Ασφάλιση', 'Ειδικότητα', 'Είδος Απασχόλησης',
                  'Έναρξη σύμβασης', 'Λήξη σύμβασης', 'Ώρες ανά εβδομάδα', 'Απόφαση ΠΥΣΠΕ/ΠΥΣΔΕ',
                  'Αρ. πράξης ΑΠΥΣΠΕ/ΑΠΥΣΔΕ/ΠΥΣΕΕΠ', 'Υπουργική Απόφαση (ΥΑ)', 'ΑΔΑ', 'Πράξη ανάληψης υπηρεσίας',
                  'Απόφαση ΔΕ', 'Απόφαση ορισμού διευθυντή', 'Απόφαση περιφερειακού διευθυντή', 'Έργο', 'Σχολεία']

        # if self.contracts[0] != header:
        #     cols = ''
        #     for item in header:
        #         cols += f"- {item}\n"
        #     showwarning(title="Λάθος τύπος αρχείου...",
        #                 message=f"Το αρχείο πρέπει να έχει τις εξής στήλες:\n{cols}")
        #
        #     return

        for item in self.contracts[1:]:
            if len(item) < 21:
                continue

            values = ''
            for i, v in enumerate(item):
                clean_v = v.replace("'", "")
                values += f"'{clean_v}'"
                if i < len(item) - 1:
                    values += ", "

            self.cursor.execute(f"INSERT INTO contracts VALUES ({values})")
            self.conn.commit()

        self.cb_end_contract_date['values'] = self.create_ecd_list()
        self.cb_end_contract_date.configure(state='readonly')

        self.btn_get_contracts.configure(state='disabled')
        self.contracts_filename.set(f_name)

        self.btn_get_payroll.configure(state='normal')

    def get_payroll(self):
        f_name = filedialog.askopenfilename(initialdir="./data", title="Επιλέξτε το αρχείο csv",
                                            filetypes=(("csv files", "*.csv"), ("all files", "*.*")))

        if f_name == "":
            return

        self.payroll = self.parse_csv_data(f_name)

        header = ['Επίθετο', 'Όνομα', 'Όνομα πατέρα', 'ΑΦΜ', 'Βασικός Μισθός', 'από', 'Οικογενειακό Επίδομα',
                  'Επίδομα προϊσταμένου Νηπιαγωγείου', 'Επίδομα Προϊσταμένου Δημοτικού',
                  'Επιδόματα προβλημ. / παραμ. / μειον. / δυσπρ. περιοχών']
        if self.payroll[0] != header:
            cols = ''
            for item in header:
                cols += f"- {item}\n"
            showwarning(title="Λάθος τύπος αρχείου...",
                        message=f"Το αρχείο πρέπει να έχει τις εξής στήλες:\n{cols}")

            return

        for item in self.payroll[1:]:
            values = ''
            for i, v in enumerate(item):
                clean_v = v.replace("'", "")
                values += f"'{clean_v}'"
                if i < len(item) - 1:
                    values += ", "

            self.cursor.execute(f"INSERT INTO payroll VALUES ({values})")
            self.conn.commit()

        self.btn_get_payroll.configure(state='disabled')
        self.payroll_filename.set(f_name)

        self.btn_get_teachers.configure(state='normal')

    def get_teachers(self):
        f_name = filedialog.askopenfilename(initialdir="./data", title="Επιλέξτε το αρχείο csv",
                                            filetypes=(("csv files", "*.csv"), ("all files", "*.*")))

        if f_name == "":
            return

        self.teachers = self.parse_csv_data(f_name)

        header = ['Α/Α', 'Επίθετο', 'Όνομα', 'Όνομα πατέρα', 'Όνομα μητέρας', 'Φύλο', 'Ημ/νία γέννησης',
                  'Οικογενειακή κατάσταση', 'Αριθμός τέκνων', 'ΑΦΜ', 'ΑΜΚΑ', 'ΔΟΥ', 'ΑΤ ή Διαβατήριο', 'Διεύθυνση',
                  'Τηλέφωνο', 'e-mail', 'Μ.Κ.', 'Έτη προϋπηρεσίας', 'Μήνες προϋπηρεσίας', 'Ημέρες προϋπηρεσίας',
                  'Έχει Master', 'Έχει Διδακτορικό', 'Ειδικότητα', 'IBAN', 'Είναι Ιερέας', 'Ασφάλιση',
                  'Αρ. μητρώου ΙΚΑ', 'Αρ. μητρώου ΤΣΜΕΔΕ', 'ΑΜ εκπαιδευτικού', 'ΚΑΔ', 'Κωδικός ειδικότητας',
                  'Ειδικές περιπτώσεις ασφάλισης (Κωδ.)', 'Έχει απαλλαγή φόρου', 'Αιτιολογία απαλλαγής φόρου']

        if self.teachers[0] != header:
            cols = ''
            for item in header:
                cols += f"- {item}\n"
            showwarning(title="Λάθος τύπος αρχείου...",
                        message=f"Το αρχείο πρέπει να έχει τις εξής στήλες:\n{cols}")

            return

        for item in self.teachers[1:]:
            values = ''
            for i, v in enumerate(item):
                clean_v = v.replace("'", "")
                values += f"'{clean_v}'"
                if i < len(item) - 1:
                    values += ", "

            self.cursor.execute(f"INSERT INTO teachers VALUES ({values})")
            self.conn.commit()

        self.btn_get_teachers.configure(state='disabled')
        self.teachers_filename.set(f_name)

        self.btn_run.configure(state='normal')

    def parse_csv_data(self, input_file):
        data = list()

        try:
            with open(input_file, 'rt', encoding='utf-8-sig') as f:
                dialect = csv.Sniffer().sniff(f.readline())
                f.seek(0)

                reader = csv.reader(f, delimiter=dialect.delimiter, quotechar=dialect.quotechar)

                for row in reader:
                    data.append(row)
        except:
            with open(input_file, 'rt') as f:
                dialect = csv.Sniffer().sniff(f.readline())
                f.seek(0)

                reader = csv.reader(f, delimiter=dialect.delimiter, quotechar=dialect.quotechar)

                for row in reader:
                    data.append(row)

        return data

    def save_file(self, data, output_file):
        wb = Workbook()
        ws = wb.active

        for entry in data:
            ws.append(entry)

        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

        not_saved = True

        while not_saved:
            try:
                wb.save(output_file)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message=f"Παρακαλώ κλείστε το αρχείο '{output_file}' ώστε να ολοκληρωθεί η αποθήκευση.")
            else:
                not_saved = False

    def create_export_data(self):
        if self.cb_end_contract_date.current() == -1:
            self.cursor.execute(f"SELECT * FROM contracts, payroll, teachers "
                                f"WHERE contracts.ΑΦΜ = payroll.ΑΦΜ AND contracts.ΑΦΜ = teachers.ΑΦΜ")
        else:
            ecd = self.end_contract_date.get()
            self.cursor.execute(f"SELECT * FROM contracts, payroll, teachers "
                                f"WHERE contracts.ΑΦΜ = payroll.ΑΦΜ AND contracts.ΑΦΜ = teachers.ΑΦΜ AND "
                                f"contracts.\"Λήξη σύμβασης\" = '{ecd}'")

        q = self.cursor.fetchall()

        header = ['ΑΦΜ', 'ΕΠΙΘΕΤΟ', 'ΟΝΟΜΑ', 'ΠΑΤΡΩΝΥΜΟ', 'ΜΗΤΡΩΝΥΜΟ', 'ΦΥΛΟ', 'ΗΜΕΡΟΜΗΝΙΑ ΓΕΝΝΗΣΗΣ',
                  'ΟΙΚΟΓΕΝΕΙΑΚΗ ΚΑΤΑΣΤΑΣΗ', 'ΤΕΚΝΑ', 'ΑΜΚΑ', 'ΑΔΤ', 'ΚΛΑΔΟΣ', 'ΚΩΔΙΚΟΣ ΕΙΔΙΚΟΤΗΤΑΣ ΕΦΚΑ',
                  'ΗΜΕΡΟΜΗΝΙΑ ΠΡΟΣΛΗΨΗΣ', 'ΗΜΕΡΟΜΗΝΙΑ ΑΠΟΛΥΣΗΣ', 'ΜΙΣΘΟΣ', 'ΜΟΡΦΩΤΙΚΟ ΕΠΙΠΕΔΟ', 'ΕΙΔΟΣ ΑΠΑΣΧΟΛΗΣΗΣ',
                  'ΣΥΜΒΑΤΙΚΗ ΗΜΕΡΟΜΗΝΙΑ ΛΗΞΗΣ ΣΥΜΒΑΣΗΣ']

        data = list()
        data.append(header)

        for item in q:
            afm = item[40]
            last_name = item[32]
            first_name = item[33]
            father_name = item[34]
            mother_name = item[35]
            gender = item[36]
            birthday = item[37]
            family_status = item[38]
            if family_status == 'Έγγαμος/η':
                family_status = 'ΕΓΓΑΜΟΣ'
            elif family_status == 'Αγαμος/η':
                family_status = 'ΑΓΑΜΟΣ'
            else:
                family_status = 'ΔΙΑΖΕΥΓΜΕΝΟΣ'

            children = item[39]
            amka = item[41]
            gov_id = item[43]
            specialty = item[53]
            specialty_code = item[61]
            start_date = item[8]
            end_date = item[9]
            hours = int(item[10])
            basic_salary = float(item[25].replace(".", "").replace(",", "."))
            family_bonus = float(item[27].replace(".", "").replace(",", "."))
            education = 'ΑΕΙ'
            work_type = item[7]
            if work_type == 'Αναπληρωτής Πλήρους Ωραρίου':
                work_type = 'ΠΛΗΡΗΣ'
            else:
                work_type = 'ΜΕΡΙΚΗ'

            typical_end_date = end_date

            # Η διαίρεση πρέπει να γίνεται κανονικά με τις ώρες του κλάδου και όχι με 23
            # Η λύση αυτή δουλεύει για τους εκπαιδευτικούς που έχουμε.
            if hours < 23:
                salary = (basic_salary + family_bonus) * hours / 23
            else:
                salary = basic_salary + family_bonus

            entry = [afm, last_name, first_name, father_name, mother_name, gender, birthday, family_status, children,
                     amka, gov_id, specialty, specialty_code, start_date, end_date, f"{salary:.2f}".replace(".", ","),
                     education, work_type, typical_end_date]

            if entry not in data:
                data.append(entry)

        return data

    def run(self):
        self.btn_run.configure(state='disabled')

        output_file = "ergani.xlsx"

        data = self.create_export_data()

        self.save_file(data, output_file)

        showinfo(title='Ολοκλήρωση Εκτέλεσης',
                 message=f'Η δημιουργία του αρχείου xlsx ολοκληρώθηκε.')

    def create_db(self):
        self.conn = sqlite3.connect('ergani.db')
        self.cursor = self.conn.cursor()

        self.cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = self.cursor.fetchall()

        for tbl in tables:
            if tbl[0] in ['contracts', 'payroll', 'teachers']:
                self.cursor.execute(f"DROP TABLE {tbl[0]}")
        self.conn.commit()

        self.cursor.execute("""CREATE TABLE contracts (
                                "A/A" text,
                                "Επίθετο" text,
                                "Όνομα" text,
                                "Όνομα πατέρα" text,
                                "ΑΦΜ" text,
                                "Ασφάλιση" text,
                                "Ειδικότητα" text,
                                "Είδος Απασχόλησης" text,
                                "Έναρξη σύμβασης" text,
                                "Λήξη σύμβασης" text,
                                "Ώρες ανά εβδομάδα" text,
                                "Απόφαση ΠΥΣΠΕ/ΠΥΣΔΕ" text,
                                "Αρ. πράξης ΑΠΥΣΠΕ/ΑΠΥΣΔΕ/ΠΥΣΕΕΠ" text,
                                "Υπουργική Απόφαση (ΥΑ)" text,
                                "ΑΔΑ" text,
                                "Πράξη ανάληψης υπηρεσίας" text,
                                "Απόφαση ΔΕ" text,
                                "Απόφαση ορισμού διευθυντή" text,
                                "Απόφαση περιφερειακού διευθυντή" text,
                                "Έργο" text,
                                "Σχολεία" text        
                            )""")

        self.cursor.execute("""CREATE TABLE payroll (
                                "Επίθετο" text,
                                "Όνομα" text,
                                "Όνομα πατέρα" text,
                                "ΑΦΜ" text,
                                "Βασικός Μισθός" text,
                                "από" text,
                                "Οικογενειακό Επίδομα" text,
                                "Επίδομα προϊσταμένου Νηπιαγωγείου" text,
                                "Επίδομα Προϊσταμένου Δημοτικού" text,
                                "Επιδόματα προβλημ. / παραμ. / μειον. / δυσπρ. περιοχών" text
                            )""")

        self.cursor.execute("""CREATE TABLE teachers (
                                "Α/Α" text,
                                "Επίθετο" text,
                                "Όνομα" text,
                                "Όνομα πατέρα" text,
                                "Όνομα μητέρας" text,
                                "Φύλο" text,
                                "Ημ/νία γέννησης" text,
                                "Οικογενειακή κατάσταση" text,
                                "Αριθμός τέκνων" text,
                                "ΑΦΜ" text,
                                "ΑΜΚΑ" text,
                                "ΔΟΥ" text,
                                "ΑΤ ή Διαβατήριο" text,
                                "Διεύθυνση" text,
                                "Τηλέφωνο" text,
                                "e-mail" text,
                                "Μ.Κ." text,
                                "Έτη προϋπηρεσίας" text,
                                "Μήνες προϋπηρεσίας" text,
                                "Ημέρες προϋπηρεσίας" text,
                                "Έχει Master" text,
                                "Έχει Διδακτορικό" text,
                                "Ειδικότητα" text,
                                "IBAN" text,
                                "Είναι Ιερέας" text,
                                "Ασφάλιση" text,
                                "Αρ. μητρώου ΙΚΑ" text,
                                "Αρ. μητρώου ΤΣΜΕΔΕ" text,
                                "ΑΜ εκπαιδευτικού" text,
                                "ΚΑΔ" text,
                                "Κωδικός ειδικότητας" text,
                                "Ειδικές περιπτώσεις ασφάλισης (Κωδ.)" text,
                                "Έχει απαλλαγή φόρου" text,
                                "Αιτιολογία απαλλαγής φόρου" text        
                            )""")

    def create_widgets(self):
        self.f_data = Frame(self.window)

        self.l_contracts = Label(self.f_data, text="Συμβάσεις (csv):")
        self.l_contracts.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.contracts_filename = StringVar()
        self.contracts_filename.set('')
        self.ntr_contracts_filename = Entry(self.f_data, width=128, state='readonly',
                                            textvariable=self.contracts_filename)
        self.ntr_contracts_filename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btn_get_contracts = Button(self.f_data, text="Επιλέξτε αρχείο...", command=self.get_contracts)
        self.btn_get_contracts.grid(column=2, row=0, padx=10, pady=10)

        self.l_payroll = Label(self.f_data, text="Μισθολογικά (csv):")
        self.l_payroll.grid(column=0, row=1, padx=10, pady=10, sticky=E)

        self.payroll_filename = StringVar()
        self.payroll_filename.set('')
        self.ntrPayrollFilename = Entry(self.f_data, width=128, state='readonly', textvariable=self.payroll_filename)
        self.ntrPayrollFilename.grid(column=1, row=1, padx=10, pady=10, sticky=W)

        self.btn_get_payroll = Button(self.f_data, text="Επιλέξτε αρχείο...", command=self.get_payroll,
                                      state='disabled')
        self.btn_get_payroll.grid(column=2, row=1, padx=10, pady=10)

        self.l_teachers = Label(self.f_data, text="Εκπαιδευτικοί (csv):")
        self.l_teachers.grid(column=0, row=2, padx=10, pady=10, sticky=E)

        self.teachers_filename = StringVar()
        self.teachers_filename.set('')
        self.ntr_teachers_filename = Entry(self.f_data, width=128, state='readonly',
                                           textvariable=self.teachers_filename)
        self.ntr_teachers_filename.grid(column=1, row=2, padx=10, pady=10, sticky=W)

        self.btn_get_teachers = Button(self.f_data, text="Επιλέξτε αρχείο...", command=self.get_teachers,
                                       state='disabled')
        self.btn_get_teachers.grid(column=2, row=2, padx=10, pady=10)

        self.l_end_contract_date = Label(self.f_data, text="Ημερομηνία Λήξης Σύμβασης:\n(χωρίς επιλογή για όλες)")
        self.l_end_contract_date.grid(column=0, row=3, padx=10, pady=10, sticky=E)

        self.end_contract_date = StringVar()
        self.cb_end_contract_date = Combobox(self.f_data, width=125, textvariable=self.end_contract_date,
                                             state='disabled')
        self.cb_end_contract_date.grid(column=1, row=3, padx=10, pady=10)

        self.btn_run = Button(self.f_data, text="Εκτέλεση", command=self.run, state='disabled')
        self.btn_run.grid(column=1, row=10, padx=10, pady=10)

        self.create_db()

        self.f_data.pack()


gui = GUI()
gui.window.mainloop()
