import xml.etree.ElementTree as ET
from xml.dom import minidom
from tkinter import *
from tkinter import filedialog
from tkinter.ttk import *
from tkinter.messagebox import showinfo
from openpyxl import load_workbook


class GUI:
    def __init__(self):
        self.window = Tk()

        self.window.title("Δημιουργία αρχείου xml για τις απολύσεις στην Εργάνη")
        self.window.resizable(False, False)
        self.create_widgets()

    def parse_xlsx_data(self):
        wb = load_workbook(filename=self.data_filename.get())
        sheet = wb.active

        self.data = list()

        for row in sheet.iter_rows():
            entry = list()
            for cell in row:
                if cell.value is None:
                    text = ""
                else:
                    text = str(cell.value).strip()

                entry.append(text)

            self.data.append(entry)

    def get_data(self):
        f_name = filedialog.askopenfilename(initialdir="./data/",
                                            title="Επιλέξτε το αρχείο xlsx",
                                            filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if f_name == "":
            return

        self.data_filename.set(f_name)
        self.parse_xlsx_data()
        self.btn_run.configure(state='normal')

    def create_xml(self):
        aa_pararthmatos = '0'
        ypiresia_sepe = '17010'
        ypiresia_oaed = '701201'
        kad_kyria = '8531'
        kad_pararthmatos = '8531'
        kallikratis_pararthmatos = '91700101'
        yphkoothta = '048'
        typos_taytothtas = 'ΔAT'
        xaraktirismos = '1'
        sxeshapasxolisis = '1'
        oros = '0'
        logosperatosis = '0'
        afm_proswpoy = '075987945'

        anaggelies_e7 = ET.Element("AnaggeliesE7")
        anaggelies_e7.set('xmlns', "http://www.yeka.gr/E7")
        anaggelies_e7.set('xmlns:xs', "http://www.w3.org/2001/XMLSchema")

        for entry in self.data[1:]:
            afm = entry[0]
            last_name = entry[1]
            first_name = entry[2]
            father_name = entry[3]
            mother_name = entry[4]

            xlsx_sex = entry[5]
            if xlsx_sex == 'ΑΝΤΡΑΣ':
                sex = '0'
            else:
                sex = '1'

            birthdate = entry[6]

            xlsx_marital_status = entry[7]
            if xlsx_marital_status == 'ΑΓΑΜΟΣ':
                marital_status = '0'
            elif xlsx_marital_status == 'ΕΓΓΑΜΟΣ':
                marital_status = '1'
            elif xlsx_marital_status == 'ΔΙΑΖΕΥΓΜΕΝΟΣ':
                marital_status = '2'
            else:
                marital_status = '3'

            children = entry[8]
            amka = entry[9]
            id = entry[10]
            specialty_code = entry[12]
            start_date = entry[13]
            end_date = entry[14]

            xlsx_salary = entry[15]
            if ',' in xlsx_salary:
                salary = xlsx_salary
            else:
                salary = xlsx_salary + ',00'

            int_part, dec_part = salary.split(',')
            if len(int_part) > 3:
                salary = f'{int_part[0]}.{int_part[1:]},{dec_part}'

            epipedo_morfosis = entry[16]
            if epipedo_morfosis == 'ΑΕΙ':
                epipedo_morfosis = '11'

            xlsx_work_status = entry[17]
            if xlsx_work_status == 'ΠΛΗΡΗΣ':
                work_status = '0'
            elif xlsx_work_status == 'ΜΕΡΙΚΗ':
                work_status = '1'
            else:
                work_status = '2'

            anaggelia_e7 = ET.SubElement(anaggelies_e7, "AnaggeliaE7")
            anaggelia_e7.set('xmlns', "")

            f_aa_pararthmatos = ET.SubElement(anaggelia_e7, "f_aa_pararthmatos")
            f_aa_pararthmatos.text = aa_pararthmatos
            f_rel_protocol = ET.SubElement(anaggelia_e7, "f_rel_protocol")
            f_rel_date = ET.SubElement(anaggelia_e7, "f_rel_date")
            f_ypiresia_sepe = ET.SubElement(anaggelia_e7, "f_ypiresia_sepe")
            f_ypiresia_sepe.text = ypiresia_sepe
            f_ypiresia_oaed = ET.SubElement(anaggelia_e7, "f_ypiresia_oaed")
            f_ypiresia_oaed.text = ypiresia_oaed
            f_ergodotikh_organwsh = ET.SubElement(anaggelia_e7, "f_ergodotikh_organwsh")
            f_kad_kyria = ET.SubElement(anaggelia_e7, "f_kad_kyria")
            f_kad_kyria.text = kad_kyria
            f_kad_deyt_1 = ET.SubElement(anaggelia_e7, "f_kad_deyt_1")
            f_kad_deyt_2 = ET.SubElement(anaggelia_e7, "f_kad_deyt_2")
            f_kad_deyt_3 = ET.SubElement(anaggelia_e7, "f_kad_deyt_3")
            f_kad_deyt_4 = ET.SubElement(anaggelia_e7, "f_kad_deyt_4")
            f_kad_pararthmatos = ET.SubElement(anaggelia_e7, "f_kad_pararthmatos")
            f_kad_pararthmatos.text = kad_pararthmatos
            f_kallikratis_pararthmatos = ET.SubElement(anaggelia_e7, "f_kallikratis_pararthmatos")
            f_kallikratis_pararthmatos.text = kallikratis_pararthmatos
            f_eponymo = ET.SubElement(anaggelia_e7, "f_eponymo")
            f_eponymo.text = last_name
            f_onoma = ET.SubElement(anaggelia_e7, "f_onoma")
            f_onoma.text = first_name
            f_eponymo_patros = ET.SubElement(anaggelia_e7, "f_eponymo_patros")
            f_onoma_patros = ET.SubElement(anaggelia_e7, "f_onoma_patros")
            f_onoma_patros.text = father_name
            f_eponymo_mitros = ET.SubElement(anaggelia_e7, "f_eponymo_mitros")
            f_onoma_mitros = ET.SubElement(anaggelia_e7, "f_onoma_mitros")
            f_onoma_mitros.text = mother_name
            f_topos_gennhshs = ET.SubElement(anaggelia_e7, "f_topos_gennhshs")
            f_birthdate = ET.SubElement(anaggelia_e7, "f_birthdate")
            f_birthdate.text = birthdate
            f_sex = ET.SubElement(anaggelia_e7, "f_sex")
            f_sex.text = sex
            f_yphkoothta = ET.SubElement(anaggelia_e7, "f_yphkoothta")
            f_yphkoothta.text = yphkoothta
            f_typos_taytothtas = ET.SubElement(anaggelia_e7, "f_typos_taytothtas")
            f_typos_taytothtas.text = typos_taytothtas
            f_ar_taytothtas = ET.SubElement(anaggelia_e7, "f_ar_taytothtas")
            f_ar_taytothtas.text = id
            f_ekdousa_arxh = ET.SubElement(anaggelia_e7, "f_ekdousa_arxh")
            f_date_ekdosis = ET.SubElement(anaggelia_e7, "f_date_ekdosis")
            f_date_ekdosis_lixi = ET.SubElement(anaggelia_e7, "f_date_ekdosis_lixi")
            f_res_permit_inst = ET.SubElement(anaggelia_e7, "f_res_permit_inst")
            f_res_permit_inst_type = ET.SubElement(anaggelia_e7, "f_res_permit_inst_type")
            f_res_permit_inst_ar = ET.SubElement(anaggelia_e7, "f_res_permit_inst_ar")
            f_res_permit_inst_lixi = ET.SubElement(anaggelia_e7, "f_res_permit_inst_lixi")
            f_res_permit_ap = ET.SubElement(anaggelia_e7, "f_res_permit_ap")
            f_res_permit_ap_type = ET.SubElement(anaggelia_e7, "f_res_permit_ap_type")
            f_res_permit_ap_ar = ET.SubElement(anaggelia_e7, "f_res_permit_ap_ar")
            f_res_permit_ap_lixi = ET.SubElement(anaggelia_e7, "f_res_permit_ap_lixi")
            f_res_permit_visa = ET.SubElement(anaggelia_e7, "f_res_permit_visa")
            f_res_permit_visa_ar = ET.SubElement(anaggelia_e7, "f_res_permit_visa_ar")
            f_res_permit_visa_from = ET.SubElement(anaggelia_e7, "f_res_permit_visa_from")
            f_res_permit_visa_to = ET.SubElement(anaggelia_e7, "f_res_permit_visa_to")
            f_marital_status = ET.SubElement(anaggelia_e7, "f_marital_status")
            f_marital_status.text = marital_status
            f_arithmos_teknon = ET.SubElement(anaggelia_e7, "f_arithmos_teknon")
            f_arithmos_teknon.text = children
            f_afm = ET.SubElement(anaggelia_e7, "f_afm")
            f_afm.text = afm
            f_doy = ET.SubElement(anaggelia_e7, "f_doy")
            f_amika = ET.SubElement(anaggelia_e7, "f_amika")
            f_amka = ET.SubElement(anaggelia_e7, "f_amka")
            f_amka.text = amka
            f_code_anergias = ET.SubElement(anaggelia_e7, "f_code_anergias")
            f_ar_vivliou_anilikou = ET.SubElement(anaggelia_e7, "f_ar_vivliou_anilikou")
            f_dieythinsi = ET.SubElement(anaggelia_e7, "f_dieythinsi")
            f_kallikratis = ET.SubElement(anaggelia_e7, "f_kallikratis")
            f_tk = ET.SubElement(anaggelia_e7, "f_tk")
            f_til = ET.SubElement(anaggelia_e7, "f_til")
            f_fax = ET.SubElement(anaggelia_e7, "f_fax")
            f_email = ET.SubElement(anaggelia_e7, "f_email")
            f_epipedo_morfosis = ET.SubElement(anaggelia_e7, "f_epipedo_morfosis")
            f_epipedo_morfosis.text = epipedo_morfosis
            f_professional_education = ET.SubElement(anaggelia_e7, "f_professional_education")
            f_expertise_field = ET.SubElement(anaggelia_e7, "f_expertise_field")
            f_subject_area = ET.SubElement(anaggelia_e7, "f_subject_area")
            f_subject_group = ET.SubElement(anaggelia_e7, "f_subject_group")
            f_education_agency = ET.SubElement(anaggelia_e7, "f_education_agency")
            f_education_date_from = ET.SubElement(anaggelia_e7, "f_education_date_from")
            f_education_date_to = ET.SubElement(anaggelia_e7, "f_education_date_to")
            f_duration = ET.SubElement(anaggelia_e7, "f_duration")
            f_education_year = ET.SubElement(anaggelia_e7, "f_education_year")
            f_fl1 = ET.SubElement(anaggelia_e7, "f_fl1")
            f_fl2 = ET.SubElement(anaggelia_e7, "f_fl2")
            f_fl3 = ET.SubElement(anaggelia_e7, "f_fl3")
            f_fl4 = ET.SubElement(anaggelia_e7, "f_fl4")
            f_pc = ET.SubElement(anaggelia_e7, "f_pc")
            f_pc_other = ET.SubElement(anaggelia_e7, "f_pc_other")
            f_xaraktirismos = ET.SubElement(anaggelia_e7, "f_xaraktirismos")
            f_xaraktirismos.text = xaraktirismos
            f_sxeshapasxolisis = ET.SubElement(anaggelia_e7, "f_sxeshapasxolisis")
            f_sxeshapasxolisis.text = sxeshapasxolisis
            f_kathestosapasxolisis = ET.SubElement(anaggelia_e7, "f_kathestosapasxolisis")
            f_kathestosapasxolisis.text = work_status
            f_oros = ET.SubElement(anaggelia_e7, "f_oros")
            f_oros.text = oros
            f_eidikothta = ET.SubElement(anaggelia_e7, "f_eidikothta")
            f_eidikothta.text = specialty_code
            f_apodoxes = ET.SubElement(anaggelia_e7, "f_apodoxes")
            f_apodoxes.text = salary
            f_proslipsidate = ET.SubElement(anaggelia_e7, "f_proslipsidate")
            f_proslipsidate.text = start_date
            f_lixisymbashdate = ET.SubElement(anaggelia_e7, "f_lixisymbashdate")
            f_lixisymbashdate.text = end_date
            f_apolysisdate = ET.SubElement(anaggelia_e7, "f_apolysisdate")
            f_apolysisdate.text = end_date
            f_comments = ET.SubElement(anaggelia_e7, "f_comments")
            f_logosperatosis = ET.SubElement(anaggelia_e7, "f_logosperatosis")
            f_logosperatosis.text = logosperatosis
            f_logosperatosiscomments = ET.SubElement(anaggelia_e7, "f_logosperatosiscomments")
            f_afm_proswpoy = ET.SubElement(anaggelia_e7, "f_afm_proswpoy")
            f_afm_proswpoy.text = afm_proswpoy
            f_file = ET.SubElement(anaggelia_e7, "f_file")
            f_foreign_file = ET.SubElement(anaggelia_e7, "f_foreign_file")
            f_young_file = ET.SubElement(anaggelia_e7, "f_young_file")

        xmlstr = minidom.parseString(ET.tostring(anaggelies_e7)).toprettyxml(indent="  ", encoding="utf-8")
        with open("ergani.xml", "wb") as f:
            f.write(xmlstr)

    def run(self):
        self.create_xml()
        showinfo(title="Ολοκλήρωση εκτέλεσης",
                 message="Η δημιουργία του αρχείου xml για τις απολύσεις στην Εργάνη ολοκληρώθηκε.")
        self.window.destroy()

    def create_widgets(self):
        self.f_data = Frame(self.window)

        self.l_data = Label(self.f_data, text="Αρχείο xlsx:")
        self.l_data.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.data_filename = StringVar()
        self.ntr_data_filename = Entry(self.f_data, width=128, state='readonly', textvariable=self.data_filename)
        self.ntr_data_filename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btn_get_data = Button(self.f_data, text="Επιλέξτε αρχείο...", command=self.get_data)
        self.btn_get_data.grid(column=2, row=0, padx=10, pady=10)

        self.btn_run = Button(self.f_data, text="Εκτέλεση", command=self.run, state='disabled')
        self.btn_run.grid(column=0, columnspan=3, row=10, padx=10, pady=10)

        self.f_data.pack()


gui = GUI()
gui.window.mainloop()
