from tkinter import *
from tkinter import filedialog
from tkinter.ttk import *
from tkinter.messagebox import showinfo, showwarning
from openpyxl import *
from openpyxl.utils import get_column_letter


class GUI:
    def __init__(self):
        self.window = Tk()

        self.window.title("Δημιουργία αρχείου dde3.xlsx για τις απολύσεις στην Εργάνη")
        self.window.resizable(False, False)
        self.create_widgets()

    def parse_xlsx_data(self):
        wb = load_workbook(filename=self.data_filename.get())
        sheet = wb.active

        self.data = list()

        for row in sheet.iter_rows():
            entry = list()
            for cell in row:
                if cell.value is None:
                    text = ""
                else:
                    text = str(cell.value).strip()

                entry.append(text)

            self.data.append(entry)

    def get_data(self):
        f_name = filedialog.askopenfilename(initialdir="./data/",
                                            title="Επιλέξτε το αρχείο xlsx",
                                            filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if f_name == "":
            return

        self.data_filename.set(f_name)
        self.parse_xlsx_data()
        self.btn_run.configure(state='normal')

    def create_xlsx(self):
        wb = Workbook()
        ws = wb.active

        header = self.data[0][:19]
        ws.append(header)

        for row, entry in enumerate(self.data[1:]):
            out_entry = entry[:19].copy()

            for i in [1, 2, 3, 4, 6]:
                # Row + 2 corresponds to excel row

                if i in [1, 2, 3, 4] and entry[i] != entry[i + 24]:
                    print(f'{row + 2}: {entry[i]} <> {entry[i + 24]}')
                    out_entry[i] = entry[i + 24]
                elif i == 6 and entry[6] != entry[29]:
                    print(f'{row + 2}: {entry[6]} <> {entry[29]}')
                    out_entry[6] = entry[29]

            ws.append(out_entry)

        self.set_cols_width(ws)

        output_file = "dde3.xlsx"
        self.safe_save(wb, output_file)

    def set_cols_width(self, ws):
        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

    def safe_save(self, wb, output_file):
        not_saved = True

        while not_saved:
            try:
                wb.save(output_file)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message=f"Παρακαλώ κλείστε το αρχείο '{output_file}' ώστε να ολοκληρωθεί η αποθήκευση.")
            else:
                not_saved = False

    def run(self):
        self.create_xlsx()
        showinfo(title="Ολοκλήρωση εκτέλεσης",
                 message="Η δημιουργία του αρχείου dde3.xlsx για τις απολύσεις στην Εργάνη ολοκληρώθηκε.")
        self.window.destroy()

    def create_widgets(self):
        self.f_data = Frame(self.window)

        self.l_data = Label(self.f_data, text="Αρχείο xlsx:")
        self.l_data.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.data_filename = StringVar()
        self.ntr_data_filename = Entry(self.f_data, width=128, state='readonly', textvariable=self.data_filename)
        self.ntr_data_filename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btn_get_data = Button(self.f_data, text="Επιλέξτε αρχείο...", command=self.get_data)
        self.btn_get_data.grid(column=2, row=0, padx=10, pady=10)

        self.btn_run = Button(self.f_data, text="Εκτέλεση", command=self.run, state='disabled')
        self.btn_run.grid(column=0, columnspan=3, row=10, padx=10, pady=10)

        self.f_data.pack()


gui = GUI()
gui.window.mainloop()
