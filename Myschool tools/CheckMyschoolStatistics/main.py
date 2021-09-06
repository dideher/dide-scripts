from tkinter import *
from tkinter import filedialog
from tkinter.ttk import *
from tkinter.messagebox import showinfo, showwarning
import os
from os import walk
from openpyxl import *
from openpyxl.utils import get_column_letter
import xlrd


class GUI():
    def __init__(self):
        self.window = Tk()

        self.window.title("Έλεγχος στατιστικών Myschool")
        self.window.resizable(False, False)
        self.create_widgets()

    def getInputDirName(self):
        dName = filedialog.askdirectory(initialdir="./data/", title="Επιλέξτε τον φάκελο που περιέχει τα αρχεία xls")

        if dName == "":
            return

        self.inputDirName.set(dName)
        self.btnOpenInputDir.configure(state='disabled')
        self.ntrYear.configure(state='normal')
        self.btnRun.configure(state='normal')

    def is_float(self, s):
        return s.replace('.', '', 1).isdecimal()

    def clean_spaces(self, text):
        while "  " in text:
            text = text.replace("  ", " ")

        return text

    def convertXls(self):
        inputDirectory = self.inputDirName.get()

        for path, dirs, files in walk(inputDirectory):
            for file in files:
                if file[-4:] == ".xls":
                    book_xls = xlrd.open_workbook(os.path.join(path, file))
                    book_xlsx = Workbook()

                    sheet_names = book_xls.sheet_names()
                    for sheet_index, sheet_name in enumerate(sheet_names):
                        sheet_xls = book_xls.sheet_by_name(sheet_name)
                        if sheet_index == 0:
                            sheet_xlsx = book_xlsx.active
                            sheet_xlsx.title = sheet_name
                        else:
                            sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

                        for row in range(0, sheet_xls.nrows):
                            for col in range(0, sheet_xls.ncols):
                                sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)

                    book_xlsx.save(os.path.join(path, file + 'x'))

    def parseXlsxData(self, file, data, row_offset=0, col_stop=0, skip_list=None):
        workbook = load_workbook(filename=file)
        sheet = workbook.active

        ro = 0
        for row in sheet.iter_rows():
            if ro < row_offset:
                ro += 1
                continue

            entry = list()
            for col, cell in enumerate(row):
                if col_stop != 0 and col >= col_stop:
                    break

                if skip_list != None and col in skip_list:
                    continue

                if cell.value is None:
                    text = ""
                else:
                    text = str(cell.value).strip()

                if self.is_float(text):
                    text = text.replace(".", ",")

                entry.append(text)

            data.append(entry)

    def verifyMainData(self):
        errors = ''

        # Έλεγχος στα επιμέρους αθροίσματα
        for entry in self.mainStatisticsData:
            for i in range(9):
                if int(entry[10 + i * 3]) != int(entry[10 + i * 3 - 2]) + int(entry[10 + i * 3 - 1]):
                    errors += f'{entry[6]} - {entry[7]}: Λάθος στο {i + 1} άθροισμα ({entry[10 + i * 3]} != {entry[10 + i * 3 - 2]} + {entry[10 + i * 3 - 1]})\n'

        # Έλεγχος της πρώτης τριάδας σε σχέση με το άθροισμα των υπολοίπων
        for entry in self.mainStatisticsData:
            boys = 0
            girls = 0
            sum = 0

            for i in range(8):
                boys += int(entry[13 + i * 3 - 2])
                girls += int(entry[13 + i * 3 - 1])
                sum += int(entry[13 + i * 3])

            if int(entry[8]) != boys or int(entry[9]) != girls or int(entry[10]) != sum:
                errors += f'{entry[6]} - {entry[7]}: Δεν συμφωνεί η πρώτη τριάδα ({entry[8]} + {entry[9]} = {entry[10]}) με το άθροισμα των υπολοίπων ({boys} + {girls} = {sum}).\n'

        print(20 * '-', 'Έλεγχος Συγκεντωτικού Πίνακα', 20 * '-')
        if errors == '':
            print('Δεν υπήρχαν λάθη.')
        else:
            print(errors)

        print(70 * '-')

    def checkUnique(self):
        errorsList = list()
        checkList = list()

        for entry in self.mainStatisticsData:
            if entry[6:8] not in checkList:
                checkList.append(entry[6:8])
            else:
                errorsList.append(f'Συγκεντρωτικός Πίνακας: H εγγραφή {entry[6:8]} υπάρχει ήδη.')

        for i in range(8):
            checkList.clear()

            for entry in self.specificLists[i]:
                if entry[6:8] not in checkList:
                    checkList.append(entry[6:8])
                else:
                    errorsList.append(f'Πίνακας {i + 2}: H εγγραφή {entry[6:8]} υπάρχει ήδη.')

        errorsList.sort()

        print(20 * '-', 'Έλεγχος Μοναδικότητας Εγγραφών', 20 * '-')
        if len(errorsList) == 0:
            print('Δεν υπήρχαν λάθη.')
        else:
            for entry in errorsList:
                print(entry)
        print(70 * '-')

    def splitMainDataToLists(self):
        self.mainStatisticsLists = list()

        for i in range(8):
            self.mainStatisticsLists.append(list())

        for entry in self.mainStatisticsData:
            common = entry[:8]
            for i in range(8):
                if (entry[11 + i * 3:14 + i * 3] != ['0', '0', '0']):
                    self.mainStatisticsLists[i].append(common + entry[11 + i * 3:14 + i * 3])

        for i in range(8):
            self.mainStatisticsLists[i] = self.cleanList(self.mainStatisticsLists[i])

    def createSpecificLists(self):
        self.specificLists = list()

        inputDirectory = self.inputDirName.get()

        for i in range(8):
            self.specificLists.append(list())

            xlsxFile = os.path.join(inputDirectory, f'10_{i + 2}.xlsx')

            if i == 0:
                self.parseXlsxData(xlsxFile, self.specificLists[i], 2, 11)
            elif i == 1:
                self.parseXlsxData(xlsxFile, self.specificLists[i], 3, 14, [8, 9, 10])
            elif i == 2:
                self.parseXlsxData(xlsxFile, self.specificLists[i], 2, 14, [8, 9, 10])
            else:
                self.parseXlsxData(xlsxFile, self.specificLists[i], 3, 11)

            self.specificLists[i] = self.cleanList(self.specificLists[i])

    def cleanList(self, l):
        clean_list = list()

        for item in l:
            if item[-3:] != ['0', '0', '0']:
                clean_list.append(item)

        return clean_list

    def checkSpecificLists(self):
        year = self.year.get()
        errorsDict = dict()

        for i in range(8):
            for entry in self.mainStatisticsLists[i]:
                if entry not in self.specificLists[i]:
                    found = False
                    for item in self.specificLists[i]:
                        if entry[:8] == item[:8]:
                            found = True
                            break

                    if entry[6] not in errorsDict:
                        errorsDict[entry[6]] = set()

                    if found:
                        errorsDict[entry[6]].add(
                            f'{entry[6:8]} Πίνακας {i + 2}: ασυμφωνία τιμών μεταξύ Συγκεντρωτικού και Επιμέρους Πίνακα.\n' +
                            f'\t- Συγκεντρωτικός Πίνακας: {entry[8:12]}\n' +
                            f'\t- Επιμέρους Πίνακας: {item[8:12]}')
                    else:
                        errorsDict[entry[6]].add(
                            f'{entry[6:8]} Πίνακας {i + 2}: εμφανίζεται μόνο στον Συγκεντρωτικό Πίνακα.\n' +
                            f'\t- Συγκεντρωτικός Πίνακας: {entry[8:12]}')

            for entry in self.specificLists[i]:
                if entry not in self.mainStatisticsLists[i]:
                    found = False
                    for item in self.mainStatisticsLists[i]:
                        if entry[:8] == item[:8]:
                            found = True
                            break

                    if entry[6] not in errorsDict:
                        errorsDict[entry[6]] = set()

                    if not found:
                        errorsDict[entry[6]].add(
                            f'{entry[6:8]} Πίνακας {i + 2}: εμφανίζεται μόνο στον Επιμέρους Πίνακα.\n' +
                            f'\t- Επιμέρους Πίνακας: {entry[8:12]}')

        printList = list()
        for err in errorsDict:
            for item in errorsDict[err]:
                printList.append([year, err, item])

        printList.sort()

        print(20 * '-', 'Λάθη ανά σχολείο', 20 * '-')
        for i, item in enumerate(printList):
            print(f'{i + 1:3}) {item[2]}')
        print(80 * '-')

        exportList = [['Σχ. Έτος', 'Σχολείο', 'Λάθος'], ]
        exportList += printList

        self.saveFile(exportList, f'{year}.xlsx')

    def setColsWidth(self, ws):
        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

    def saveFile(self, l, filename):
        wb = Workbook()
        ws = wb.active

        for row in l:
            ws.append(row)

        self.setColsWidth(ws)

        outputFile = os.path.join(self.inputDirName.get(), filename)

        notSaved = True

        while notSaved:
            try:
                wb.save(outputFile)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message=f"Παρακαλώ κλείστε το αρχείο '{filename}' ώστε να ολοκληρωθεί η αποθήκευση.")
            else:
                notSaved = False

    def run(self):
        if self.ntrYear.get() == '':
            showwarning(title='Μη καταχώριση έτους', message='Πρέπει να καταχωρίσετε το σχ. έτος.')
            return

        self.ntrYear.configure(state='readonly')
        self.btnRun.configure(state='disabled')

        self.convertXls()

        inputDirectory = self.inputDirName.get()
        mainStatisticsFile = os.path.join(inputDirectory, '10_1.xlsx')
        self.mainStatisticsData = list()
        self.parseXlsxData(mainStatisticsFile, self.mainStatisticsData, 2)
        self.verifyMainData()
        self.splitMainDataToLists()
        self.createSpecificLists()
        self.checkUnique()
        self.checkSpecificLists()

        showinfo(title="Ολοκλήρωση εκτέλεσης", message="Ο έλεγχος ολοκληρώθηκε.")

    def create_widgets(self):
        self.fData = Frame(self.window)

        self.lInputDirName = Label(self.fData, text="Φάκελος αρχείων xls:")
        self.lInputDirName.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.inputDirName = StringVar()
        self.ntrInputDirName = Entry(self.fData, width=128, state='readonly', textvariable=self.inputDirName)
        self.ntrInputDirName.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenInputDir = Button(self.fData, text="Επιλέξτε φάκελο...", command=self.getInputDirName)
        self.btnOpenInputDir.grid(column=2, row=0, padx=10, pady=10)

        self.lYear = Label(self.fData, text="Σχολικό Έτος:")
        self.lYear.grid(column=0, row=1, padx=10, pady=10, sticky=E)

        self.year = StringVar()
        self.ntrYear = Entry(self.fData, width=128, state='disabled', textvariable=self.year)
        self.ntrYear.grid(column=1, row=1, padx=10, pady=10, sticky=W)

        self.btnRun = Button(self.fData, text="Εκτέλεση", command=self.run, state='disabled')
        self.btnRun.grid(column=1, row=10, padx=10, pady=10)

        self.fData.pack()


gui = GUI()
gui.window.mainloop()
