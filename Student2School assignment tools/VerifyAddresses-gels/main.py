from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from openpyxl import *
from openpyxl.utils import get_column_letter
from os.path import exists
import sqlite3
import re


def parse_xlsx_data(filename):
    wb = load_workbook(filename=filename)
    sheet = wb.active

    students = list()

    for row in sheet.iter_rows():
        student = list()
        for cell in row:
            if cell.value is None:
                student.append("")
            else:
                text1 = (str(cell.value).upper().replace(" .", ". ").replace("Ά", "Α").replace("Έ", "Ε")
                         .replace("Ή", "Η").replace("Ί", "Ι").replace("Ϊ́", "Ϊ").replace("Ύ", "Υ").replace("Ϋ́", "Ϋ")
                         .replace("Ό", "Ο").replace("Ώ", "Ω").replace("'", " ").strip())

                text2 = re.sub(r'[ ]+', r' ', text1)
                student.append(re.sub(r'([0-9]+)Ο', r'\1ο', text2))

        students.append(student)

    return students[1:]


class GUI:
    def __init__(self):
        self.window = Tk()

        self.window.title("Διασταύρωση των διευθύνσεων των μαθητών")
        self.window.resizable(False, False)
        self.create_widgets()

    def check_db(self):
        if not exists('students.db'):
            self.btn_get_file.configure(state='normal')
        else:
            self.conn = sqlite3.connect('students.db')
            self.cursor = self.conn.cursor()

            self.cursor.execute("SELECT * FROM students")
            self.students = self.cursor.fetchall()
            self.entries_count.set(str(len(self.students)))

            self.cursor.execute("SELECT * FROM info")
            student_no = self.cursor.fetchone()[0]

            self.load_student(student_no + 1)

    def create_db(self):
        self.conn = sqlite3.connect('students.db')
        self.cursor = self.conn.cursor()

        self.cursor.execute("""CREATE TABLE students (
                            student_id text,
                            sch_year text,
                            sch_class text,
                            address text,
                            origin_sch text,
                            gm_sch text,
                            gm_address text,
                            gm_coords text,
                            gv3_sch text,
                            gv3_address text,
                            gv3_coords text,
                            bm_sch text,
                            bm_address text,
                            bm_coords text,
                            hm_sch text,
                            hm_address text,
                            hm_coords text,
                            assignment_sch text,
                            assignment_address text,
                            assignment_coords text
                            )""")

        self.cursor.execute("""CREATE TABLE info (
                            last_edit integer
                            )""")

        self.conn.commit()

    def initialize_cache(self):
        self.addresses = dict()

        # self.sch_points = [["1ο ΓΥΜΝΑΣΙΟ", "35.3370758,25.1394652"],
        #                    ["2ο ΓΥΜΝΑΣΙΟ", "35.3306017,25.1254976"],
        #                    ["3ο ΓΥΜΝΑΣΙΟ", "35.3251757,25.1417821"],
        #                    ["4ο ΓΥΜΝΑΣΙΟ", "35.331492,25.1219348"],
        #                    ["5ο ΓΥΜΝΑΣΙΟ", "35.3129978,25.1372143999999"],
        #                    ["6ο ΓΥΜΝΑΣΙΟ", "35.340316,25.1393276"],
        #                    ["7ο ΓΥΜΝΑΣΙΟ", "35.3074499,25.1483900000001"],
        #                    ["8ο ΓΥΜΝΑΣΙΟ", "35.3336673,25.1101404"],
        #                    ["9ο ΓΥΜΝΑΣΙΟ", "35.3302412,25.1156387999999"],
        #                    ["10ο ΓΥΜΝΑΣΙΟ", "35.3290121,25.1060864"],
        #                    ["11ο ΓΥΜΝΑΣΙΟ", "35.3397199,25.1177898999999"],
        #                    ["12ο ΓΥΜΝΑΣΙΟ", "35.3358555,25.1502662"],
        #                    ["13ο ΓΥΜΝΑΣΙΟ", "35.3129578,25.1526083"],
        #                    ["ΝΕΑΣ ΑΛΙΚΑΡΝΑΣΣΟΥ", "35.3341232,25.1631108"],
        #                    ["ΑΓΙΑΣ ΒΑΡΒΑΡΑΣ", "35.1376927,25.0015113"],
        #                    ["ΑΓΙΟΥ ΜΥΡΩΝΑ", "35.2368647,25.0300641"],
        #                    ["ΑΓΙΩΝ ΔΕΚΑ", "35.061388,24.9608543"],
        #                    ["ΑΡΚΑΛΟΧΩΡΙΟΥ", "35.1487553,25.2680617999999"],
        #                    ["ΑΡΧΑΝΩΝ", "35.2304856,25.1570982"],
        #                    ["ΑΣΗΜΙΟΥ", "35.0466868,25.0911212"],
        #                    ["ΒΑΓΙΟΝΙΑΣ", "35.0098447,24.9958625"],
        #                    ["ΒΕΝΕΡΑΤΟΥ", "35.1985125,25.0384399"],
        #                    ["ΒΙΑΝΝΟΥ", "35.0514492,25.4125072"],
        #                    ["ΓΑΖΙΟΥ", "35.322475,25.070012"],
        #                    ["ΓΕΡΓΕΡΗΣ", "35.1314938,24.9401852"],
        #                    ["ΓΟΥΒΩΝ", "35.3261815,25.282647"],
        #                    ["ΕΠΙΣΚΟΠΗΣ", "35.2588729,25.2376677"],
        #                    ["ΖΑΡΟΥ", "35.1287972,24.9015107"],
        #                    ["ΘΡΑΨΑΝΟΥ", "35.186414,25.2826108"],
        #                    ["ΚΑΣΤΕΛΛΙΟΥ", "35.2060004,25.3402019"],
        #                    ["ΚΡΟΥΣΩΝΑ", "35.2320542,24.9866706"],
        #                    ["ΛΙΜΕΝΟΣ ΧΕΡΣΟΝΗΣΟΥ", "35.3131466,25.3872639"],
        #                    ["ΜΑΛΙΩΝ", "35.2845288,25.4567733"],
        #                    ["ΜΕΛΕΣΩΝ", "35.1907214,25.1999411"],
        #                    ["ΜΟΙΡΩΝ", "35.0531029,24.8690718"],
        #                    ["ΜΟΧΟΥ", "35.2586352,25.4282749"],
        #                    ["ΠΟΜΠΙΑΣ", "35.0106093,24.8649278"],
        #                    ["ΠΡΟΦΗΤΗ ΗΛΙΑ", "35.2095025,25.10061350000001"],
        #                    ["ΠΥΡΓΟΥ", "35.0080059,25.1513958"],
        #                    ["ΤΕΦΕΛΙΟΥ", "35.0943148,25.15931640000001"],
        #                    ["ΤΥΛΙΣΟΥ", "35.3047481,25.0178673"],
        #                    ["ΤΥΜΠΑΚΙΟΥ", "35.0671367,24.7662306"],
        #                    ["ΧΑΡΑΚΑ", "35.0154883,25.1195821"]]

        self.sch_points = [["1ο ΓΕΛ", "0.0,0.0"],
                           ["2ο ΓΕΛ", "0.0,0.0"],
                           ["3ο ΓΕΛ", "0.0,0.0"],
                           ["4ο ΓΕΛ", "0.0,0.0"],
                           ["5ο ΓΕΛ", "0.0,0.0"],
                           ["6ο ΓΕΛ", "0.0,0.0"],
                           ["7ο ΓΕΛ", "0.0,0.0"],
                           ["8ο ΓΕΛ", "0.0,0.0"],
                           ["10ο ΓΕΛ", "0.0,0.0"],
                           ["11ο ΓΕΛ", "0.0,0.0"],
                           ["13ο ΓΕΛ", "0.0,0.0"],
                           ["ΝΕΑΣ ΑΛΙΚΑΡΝΑΣΣΟΥ", "0.0,0.0"],
                           ["ΑΓΙΑΣ ΒΑΡΒΑΡΑΣ", "0.0,0.0"],
                           ["ΑΓΙΟΥ ΜΥΡΩΝΑ", "0.0,0.0"],
                           ["ΑΓΙΩΝ ΔΕΚΑ", "0.0,0.0"],
                           ["ΑΡΚΑΛΟΧΩΡΙΟΥ", "0.0,0.0"],
                           ["ΑΡΧΑΝΩΝ", "0.0,0.0"],
                           ["ΑΣΗΜΙΟΥ", "0.0,0.0"],
                           ["ΒΙΑΝΝΟΥ", "0.0,0.0"],
                           ["ΓΑΖΙΟΥ", "0.0,0.0"],
                           ["ΓΟΥΒΩΝ", "0.0,0.0"],
                           ["ΕΠΙΣΚΟΠΗΣ", "0.0,0.0"],
                           ["ΚΑΣΤΕΛΛΙΟΥ", "0.0,0.0"],
                           ["ΚΡΟΥΣΩΝΑ", "0.0,0.0"],
                           ["ΛΙΜ. ΧΕΡΣΟΝΗΣΟΥ", "0.0,0.0"],
                           ["ΜΑΛΙΩΝ", "0.0,0.0"],
                           ["ΜΕΛΕΣΩΝ", "0.0,0.0"],
                           ["ΜΟΙΡΩΝ", "0.0,0.0"],
                           ["ΜΟΧΟΥ", "0.0,0.0"],
                           ["ΠΟΜΠΙΑΣ", "0.0,0.0"],
                           ["ΤΥΜΠΑΚΙΟΥ", "0.0,0.0"],
                           ["ΧΑΡΑΚΑ", "0.0,0.0"]]

        self.schools = [sch[0] for sch in self.sch_points]
        self.cb_manual_assignment['values'] = self.schools

    def get_filename(self):
        f_name = filedialog.askopenfilename(initialdir=".", title="Επιλέξτε το αρχείο των μαθητών",
                                            filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if f_name == "":
            return

        self.students_filename.set(f_name)
        self.students = parse_xlsx_data(f_name)
        self.entries_count.set(str(len(self.students)))

        if len(self.students) == 0:
            showwarning(title='Αρχείο χωρίς εγγραφές', message='Παρακαλώ δοκιμάστε με νέο αρχείο.')
            return

        self.btn_get_file.configure(state='disabled')

        self.create_db()

        for student in self.students:
            student_id = student[0]
            sch_year = student[1]
            sch_class = student[2]
            address = student[3]
            origin_sch = student[4]
            gm_sch = student[5]
            gm_address = student[6]
            gm_coords = student[7]
            gv3_sch = student[8]
            gv3_address = student[9]
            gv3_coords = student[10]
            bm_sch = student[11]
            bm_address = student[12]
            bm_coords = student[13]
            hm_sch = student[14]
            hm_address = student[15]
            hm_coords = student[16]
            assignment_sch = ''
            assignment_address = ''
            assignment_coords = ''

            self.cursor.execute(
                f"INSERT INTO students VALUES ('{student_id}', '{sch_year}', '{sch_class}', "
                f"'{address}', '{origin_sch}', "
                f"'{gm_sch}', '{gm_address}', '{gm_coords}', '{gv3_sch}', '{gv3_address}', '{gv3_coords}', "
                f"'{bm_sch}', '{bm_address}', '{bm_coords}', '{hm_sch}', '{hm_address}', '{hm_coords}', "
                f"'{assignment_sch}', '{assignment_address}', '{assignment_coords}')")

        self.cursor.execute("INSERT INTO info VALUES ('0')")

        self.conn.commit()

        self.load_student(1)

    def load_student(self, student_no):
        self.btn_address_not_found.configure(state='normal')
        self.btn_gm_accept.configure(state='normal')
        self.btn_gv3_accept.configure(state='normal')
        self.btn_bm_accept.configure(state='normal')
        self.btn_hm_accept.configure(state='normal')
        self.cb_manual_assignment['values'] = self.schools
        self.cb_manual_assignment.configure(state='normal')
        self.manual_assignment.set('')
        self.btn_ma_accept.configure(state='disabled')
        self.btn_save_file.configure(state='normal')

        if student_no > len(self.students):
            student_no = len(self.students)

        if student_no > 1:
            self.btn_prev.configure(state='normal')
        else:
            self.btn_prev.configure(state='disabled')

        self.cursor.execute("SELECT * FROM info")
        last_edit = self.cursor.fetchone()[0]

        if student_no <= last_edit and student_no < len(self.students):
            self.btn_next.configure(state='normal')
        else:
            self.btn_next.configure(state='disabled')

        current_id = self.students[student_no - 1][0]

        self.entry_num.set(str(student_no))
        self.cursor.execute(f"SELECT * FROM students WHERE student_id == '{current_id}'")
        self.student = self.cursor.fetchone()

        student = self.student
        student_id = student[0]
        sch_year = student[1]
        sch_class = student[2]
        address = student[3]
        origin_sch = student[4]
        gm_sch = student[5]
        gm_address = student[6]
        gm_coords = student[7]
        gv3_sch = student[8]
        gv3_address = student[9]
        gv3_coords = student[10]
        bm_sch = student[11]
        bm_address = student[12]
        bm_coords = student[13]
        hm_sch = student[14]
        hm_address = student[15]
        hm_coords = student[16]
        assignment_sch = student[17]
        assignment_address = student[18]
        assignment_coords = student[19]

        self.student_school.set(origin_sch)
        self.student_address.set(address)
        self.assignment.set(assignment_sch)
        self.gm_address.set(gm_address)
        self.gm_school.set(gm_sch)
        self.gv3_address.set(gv3_address)
        self.gv3_school.set(gv3_sch)
        self.bm_address.set(bm_address)
        self.bm_school.set(bm_sch)
        self.hm_address.set(hm_address)
        self.hm_school.set(hm_sch)

        self.l_gm_status.configure(background='')
        self.l_gv3_status.configure(background='')
        self.l_bm_status.configure(background='')
        self.l_hm_status.configure(background='')

        if assignment_sch == gm_sch:
            self.l_gm_status.configure(background='Green')
        elif assignment_sch != '':
            self.l_gm_status.configure(background='Red')

        if assignment_sch == gv3_sch:
            self.l_gv3_status.configure(background='Green')
        elif assignment_sch != '':
            self.l_gv3_status.configure(background='Red')

        if assignment_sch == bm_sch:
            self.l_bm_status.configure(background='Green')
        elif assignment_sch != '':
            self.l_bm_status.configure(background='Red')

        if assignment_sch == hm_sch:
            self.l_hm_status.configure(background='Green')
        elif assignment_sch != '':
            self.l_hm_status.configure(background='Red')

        if student_no > last_edit:
            self.check_address(student_no)

    def finalize(self, assignment_sch):
        self.assignment.set(assignment_sch)

        gm_sch = self.gm_school.get()
        gv3_sch = self.gv3_school.get()
        bm_sch = self.bm_school.get()
        hm_sch = self.hm_school.get()

        if assignment_sch == gm_sch:
            self.l_gm_status.configure(background='Green')
        elif assignment_sch != '':
            self.l_gm_status.configure(background='Red')

        if assignment_sch == gv3_sch:
            self.l_gv3_status.configure(background='Green')
        elif assignment_sch != '':
            self.l_gv3_status.configure(background='Red')

        if assignment_sch == bm_sch:
            self.l_bm_status.configure(background='Green')
        elif assignment_sch != '':
            self.l_bm_status.configure(background='Red')

        if assignment_sch == hm_sch:
            self.l_hm_status.configure(background='Green')
        elif assignment_sch != '':
            self.l_hm_status.configure(background='Red')

        showinfo(title="Ολοκλήρωση διασταύρωσης",
                 message="Η διαδικασία της διασταύρωσης των διευθύνσεων ολοκληρώθηκε.")
        self.save_file()

    def prev_student(self):
        student_no = int(self.entry_num.get())
        self.load_student(student_no - 1)

    def next_student(self):
        student_no = int(self.entry_num.get())
        self.load_student(student_no + 1)

    def update_after_accept(self, assignment):
        student_no = int(self.entry_num.get())

        self.cursor.execute("SELECT * FROM info")
        last_edit = self.cursor.fetchone()[0]

        if last_edit < student_no:
            self.register_address(assignment)
            self.cursor.execute(f"UPDATE info "
                                f"SET last_edit = '{student_no}'")
            self.conn.commit()
        else:
            self.update_address(assignment)

        if student_no < len(self.students):
            self.load_student(student_no + 1)
        else:
            self.finalize(assignment[0])

    def update_address(self, assignment):
        student = self.student
        address = student[3]

        self.addresses[address] = assignment

        self.cursor.execute(f"UPDATE students "
                            f"SET assignment_sch = '{assignment[0]}', assignment_address = '{assignment[1]}', "
                            f"assignment_coords = '{assignment[2]}' "
                            f"WHERE address == '{address}'")
        self.conn.commit()

    def address_not_found(self):
        student = self.student
        student_id = student[0]

        self.cursor.execute(f"UPDATE students "
                            f"SET assignment_sch = 'N/A', assignment_address = 'N/A', assignment_coords = '0.0,0.0' "
                            f"WHERE student_id == '{student_id}'")
        self.conn.commit()

        self.update_after_accept(['N/A', 'N/A', '0.0,0.0'])

    def gm_accept(self):
        student = self.student
        student_id = student[0]
        gm_sch = student[5]
        gm_address = student[6]
        gm_coords = student[7]

        self.cursor.execute(f"UPDATE students "
                            f"SET assignment_sch = '{gm_sch}', "
                            f"assignment_address = '{gm_address}', "
                            f"assignment_coords = '{gm_coords}' "
                            f"WHERE student_id == '{student_id}'")

        self.update_after_accept([gm_sch, gm_address, gm_coords])

    def gv3_accept(self):
        student = self.student
        student_id = student[0]
        gv3_sch = student[8]
        gv3_address = student[9]
        gv3_coords = student[10]

        self.cursor.execute(f"UPDATE students "
                            f"SET assignment_sch = '{gv3_sch}', "
                            f"assignment_address = '{gv3_address}', "
                            f"assignment_coords = '{gv3_coords}' "
                            f"WHERE student_id == '{student_id}'")

        self.update_after_accept([gv3_sch, gv3_address, gv3_coords])

    def bm_accept(self):
        student = self.student
        student_id = student[0]
        bm_sch = student[11]
        bm_address = student[12]
        bm_coords = student[13]

        self.cursor.execute(f"UPDATE students "
                            f"SET assignment_sch = '{bm_sch}', "
                            f"assignment_address = '{bm_address}', "
                            f"assignment_coords = '{bm_coords}' "
                            f"WHERE student_id == '{student_id}'")

        self.update_after_accept([bm_sch, bm_address, bm_coords])

    def hm_accept(self):
        student = self.student
        student_id = student[0]
        hm_sch = student[14]
        hm_address = student[15]
        hm_coords = student[16]

        self.cursor.execute(f"UPDATE students "
                            f"SET assignment_sch = '{hm_sch}', "
                            f"assignment_address = '{hm_address}', "
                            f"assignment_coords = '{hm_coords}' "
                            f"WHERE student_id == '{student_id}'")

        self.update_after_accept([hm_sch, hm_address, hm_coords])

    def ma_accept(self):
        student_no = int(self.entry_num.get())
        ma_text = self.manual_assignment.get()

        mac = -1
        for i, item in enumerate(self.schools):
            if item == ma_text:
                mac = i

        if mac == -1:
            showwarning(title='Λάθος ονομασία σχολείου', message='Παρακαλώ επιλέξτε σχολείο από τη λίστα.')
            self.manual_assignment.set('')
            self.cb_manual_assignment['values'] = self.schools
            return

        student = self.student
        student_id = student[0]

        self.cursor.execute(f"UPDATE students "
                            f"SET assignment_sch = '{self.sch_points[mac][0]}', "
                            f"assignment_address = 'Manual assignment', "
                            f"assignment_coords = '{self.sch_points[mac][1]}' "
                            f"WHERE student_id == '{student_id}'")

        self.update_after_accept([self.sch_points[mac][0],
                                  'Manual assignment',
                                  self.sch_points[mac][1]])

    def manual_assignment_select(self, event_object):
        self.btn_ma_accept.configure(state='normal')

    def manual_assignment_filter_list(self, e):
        typed = self.manual_assignment.get()

        if typed == '':
            self.cb_manual_assignment['values'] = self.schools
        else:
            filtered_list = list()
            for item in self.schools:
                if typed.upper() in item.upper():
                    filtered_list.append(item)

            self.cb_manual_assignment['values'] = filtered_list

    def check_address(self, student_no):
        student = self.student
        student_id = student[0]
        address = student[3]

        if address not in self.addresses:
            return

        result = self.addresses[address]

        print(f"[{address}] --> {result}")

        self.cursor.execute(f"UPDATE students "
                            f"SET assignment_sch = '{result[0]}', "
                            f"assignment_address = '{result[1]}', "
                            f"assignment_coords = '{result[2]}' "
                            f"WHERE student_id == '{student_id}'")

        self.cursor.execute("SELECT * FROM info")
        last_edit = self.cursor.fetchone()[0]

        if last_edit < student_no:
            self.cursor.execute(f"UPDATE info "
                                f"SET last_edit = '{student_no}'")
        self.conn.commit()

        if student_no < len(self.students):
            self.load_student(student_no + 1)
        else:
            self.finalize(result[0])

    def register_address(self, result):
        student = self.student
        address = student[3]

        self.addresses[address] = result

    def save_file(self):
        wb = Workbook()
        ws = wb.active
        header = ["Α. Μ. /Α. Α. ΑΙΤΗΣΗΣ",
                  "ΤΕΛΕΥΤΑΙΑ ΧΡΟΝΙΑ ΦΟΙΤΗΣΗΣ",
                  "ΤΑΞΗ ΠΡΟΟΡΙΣΜΟΥ",
                  "ΔΙΕΥΘΥΝΣΗ ΚΑΤΟΙΚΙΑΣ",
                  "ΣΧΟΛΕΙΟ ΠΡΟΕΛΕΥΣΗΣ",
                  "Σχολείο (Google Maps)",
                  "Διεύθυνση (Google Maps)",
                  "Συντεταγμένες (Google Maps)",
                  "Σχολείο (Google V3)",
                  "Διεύθυνση (Google V3)",
                  "Συντεταγμένες (Google V3)",
                  "Σχολείο (Bing Maps)",
                  "Διεύθυνση (Bing Maps)",
                  "Συντεταγμένες (Bing Maps)",
                  "Σχολείο (Here Maps)",
                  "Διεύθυνση (Here Maps)",
                  "Συντεταγμένες (Here Maps)",
                  "Σχολείο Κατανομής",
                  "Διεύθυνση Κατανομής",
                  "Συντεταγμένες Κατανομής"
                  ]
        ws.append(header)

        self.cursor.execute(f"SELECT * FROM students")
        q = self.cursor.fetchall()

        for student in q:
            ws.append(student)

        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

        output_file = "db_output.xlsx"
        showinfo(title="Αρχείο εξόδου",
                 message="Η κατανομή θα αποθηκευτεί στο αρχείο: " + output_file)

        not_saved = True

        while not_saved:
            try:
                wb.save(output_file)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message=f"Παρακαλώ κλείστε το αρχείο '{output_file}' "
                                    f"ώστε να ολοκληρωθεί η αποθήκευση της νέας κατανομής.")
            else:
                not_saved = False

    def create_widgets(self):
        self.f_main = Frame(self.window)

        self.l_students_file = Label(self.f_main, text="Αρχείο:")
        self.l_students_file.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.students_filename = StringVar()
        self.students_filename.set('')
        self.ntr_students_filename = Entry(self.f_main, width=128, state='readonly',
                                           textvariable=self.students_filename)
        self.ntr_students_filename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btn_get_file = Button(self.f_main, text="Επιλέξτε αρχείο...", state='disabled', command=self.get_filename)
        self.btn_get_file.grid(column=2, row=0, padx=10, pady=10)

        self.lf_info = LabelFrame(self.f_main, text="Πληροφορίες / Εξαγωγή Βάσης")
        self.lf_info.grid(column=0, row=1, columnspan=3, padx=10, pady=10)
        self.l_entry_num = Label(self.lf_info, text="Α/Α:")
        self.l_entry_num.grid(column=0, row=0, padx=10, pady=5, sticky=E)
        self.entry_num = StringVar()
        self.ntr_entry_num = Entry(self.lf_info, width=10, state='readonly', textvariable=self.entry_num)
        self.ntr_entry_num.grid(column=1, row=0, padx=10, pady=5, sticky=W)
        self.btn_prev = Button(self.lf_info, text="<", state='disabled', command=self.prev_student)
        self.btn_prev.grid(column=2, row=0, padx=10, pady=10)
        self.btn_next = Button(self.lf_info, text=">", state='disabled', command=self.next_student)
        self.btn_next.grid(column=3, row=0, padx=10, pady=10)
        self.l_entries_count = Label(self.lf_info, text="Πλήθος εγγραφών:")
        self.l_entries_count.grid(column=4, row=0, padx=10, pady=5, sticky=E)
        self.entries_count = StringVar()
        self.ntr_entries_count = Entry(self.lf_info, width=10, state='readonly', textvariable=self.entries_count)
        self.ntr_entries_count.grid(column=5, row=0, padx=10, pady=5, sticky=W)
        self.btn_save_file = Button(self.lf_info, text="Εξαγωγή Βάσης", state='disabled', command=self.save_file)
        self.btn_save_file.grid(column=6, row=0, padx=10, pady=10)

        self.lf_student_info = LabelFrame(self.f_main, text="Στοιχεία μαθητή")
        self.lf_student_info.grid(column=0, row=2, columnspan=3, padx=10, pady=10, sticky=EW)
        self.l_student_school = Label(self.lf_student_info, text="Σχολείο μαθητή:")
        self.l_student_school.grid(column=0, row=0, padx=10, pady=10, sticky=E)
        self.student_school = StringVar()
        self.ntr_student_school = Entry(self.lf_student_info, width=140, state='readonly',
                                        textvariable=self.student_school)
        self.ntr_student_school.grid(column=1, row=0, padx=10, pady=5)
        self.l_assignment = Label(self.lf_student_info, text="Σχολείο κατανομής:")
        self.l_assignment.grid(column=0, row=1, padx=10, pady=10, sticky=E)
        self.assignment = StringVar()
        self.ntr_assignment = Entry(self.lf_student_info, width=140, state='readonly',
                                    textvariable=self.assignment)
        self.ntr_assignment.grid(column=1, row=1, padx=10, pady=5)
        self.l_student_address = Label(self.lf_student_info, text="Διεύθυνση μαθητή:")
        self.l_student_address.grid(column=0, row=2, padx=10, pady=10, sticky=E)
        self.student_address = StringVar()
        self.ntr_student_address = Entry(self.lf_student_info, width=140, state='readonly',
                                         textvariable=self.student_address)
        self.ntr_student_address.grid(column=1, row=2, padx=10, pady=5)

        self.lf_gecoding_info = LabelFrame(self.f_main, text="Διεύθυνση / Κατανομή")
        self.lf_gecoding_info.grid(column=0, row=3, columnspan=3, padx=10, pady=10, sticky=EW)

        self.gm_address = StringVar()
        self.ntr_gm_address = Entry(self.lf_gecoding_info, width=100, state='readonly', textvariable=self.gm_address)
        self.ntr_gm_address.grid(column=0, row=0, padx=10, pady=5, sticky=W)
        self.gm_school = StringVar()
        self.ntr_gm_school = Entry(self.lf_gecoding_info, width=40, state='readonly', textvariable=self.gm_school)
        self.ntr_gm_school.grid(column=1, row=0, padx=10, pady=5)
        self.l_gm_status = Label(self.lf_gecoding_info, text=" ")
        self.l_gm_status.grid(column=2, row=0, padx=10, pady=10, sticky=E)
        self.btn_gm_accept = Button(self.lf_gecoding_info, text="Αποδοχή", state='disabled', command=self.gm_accept)
        self.btn_gm_accept.grid(column=3, row=0, padx=10, pady=5)

        self.gv3_address = StringVar()
        self.ntr_gv3_address = Entry(self.lf_gecoding_info, width=100, state='readonly', textvariable=self.gv3_address)
        self.ntr_gv3_address.grid(column=0, row=1, padx=10, pady=5, sticky=W)
        self.gv3_school = StringVar()
        self.ntr_gv3_school = Entry(self.lf_gecoding_info, width=40, state='readonly', textvariable=self.gv3_school)
        self.ntr_gv3_school.grid(column=1, row=1, padx=10, pady=5)
        self.l_gv3_status = Label(self.lf_gecoding_info, text=" ")
        self.l_gv3_status.grid(column=2, row=1, padx=10, pady=10, sticky=E)
        self.btn_gv3_accept = Button(self.lf_gecoding_info, text="Αποδοχή", state='disabled', command=self.gv3_accept)
        self.btn_gv3_accept.grid(column=3, row=1, padx=10, pady=5)

        self.bm_address = StringVar()
        self.ntr_bm_address = Entry(self.lf_gecoding_info, width=100, state='readonly', textvariable=self.bm_address)
        self.ntr_bm_address.grid(column=0, row=2, padx=10, pady=5, sticky=W)
        self.bm_school = StringVar()
        self.ntr_bm_school = Entry(self.lf_gecoding_info, width=40, state='readonly', textvariable=self.bm_school)
        self.ntr_bm_school.grid(column=1, row=2, padx=10, pady=5)
        self.l_bm_status = Label(self.lf_gecoding_info, text=" ")
        self.l_bm_status.grid(column=2, row=2, padx=10, pady=10, sticky=E)
        self.btn_bm_accept = Button(self.lf_gecoding_info, text="Αποδοχή", state='disabled', command=self.bm_accept)
        self.btn_bm_accept.grid(column=3, row=2, padx=10, pady=5)

        self.hm_address = StringVar()
        self.ntr_hm_address = Entry(self.lf_gecoding_info, width=100, state='readonly', textvariable=self.hm_address)
        self.ntr_hm_address.grid(column=0, row=3, padx=10, pady=5, sticky=W)
        self.hm_school = StringVar()
        self.ntr_hm_school = Entry(self.lf_gecoding_info, width=40, state='readonly', textvariable=self.hm_school)
        self.ntr_hm_school.grid(column=1, row=3, padx=10, pady=5)
        self.l_hm_status = Label(self.lf_gecoding_info, text=" ")
        self.l_hm_status.grid(column=2, row=3, padx=10, pady=10, sticky=E)
        self.btn_hm_accept = Button(self.lf_gecoding_info, text="Αποδοχή", state='disabled', command=self.hm_accept)
        self.btn_hm_accept.grid(column=3, row=3, padx=10, pady=5)

        self.lf_manual_assign = LabelFrame(self.f_main, text="Επιλογή σχολείου")
        self.lf_manual_assign.grid(column=0, row=4, columnspan=3, padx=10, pady=10, sticky=EW)

        self.manual_assignment = StringVar()
        self.cb_manual_assignment = Combobox(self.lf_manual_assign, width=142, textvariable=self.manual_assignment,
                                             state='disabled')
        self.cb_manual_assignment.bind("<KeyRelease>", self.manual_assignment_filter_list)
        self.cb_manual_assignment.bind("<<ComboboxSelected>>", self.manual_assignment_select)
        self.cb_manual_assignment.grid(column=0, row=0, columnspan=2, padx=10, pady=10)
        self.btn_ma_accept = Button(self.lf_manual_assign, text="Αποδοχή", state='disabled', command=self.ma_accept)
        self.btn_ma_accept.grid(column=2, row=0, padx=10, pady=10)

        self.btn_address_not_found = Button(self.f_main, text="Η διεύθυνση δεν βρέθηκε", state='disabled',
                                            command=self.address_not_found)
        self.btn_address_not_found.grid(column=0, row=6, columnspan=3, padx=10, pady=5, sticky=EW)

        self.initialize_cache()
        self.check_db()

        self.f_main.pack()


gui = GUI()
gui.window.mainloop()
