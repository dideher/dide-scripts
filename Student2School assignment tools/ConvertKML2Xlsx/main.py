from tkinter import *
from tkinter import filedialog
from tkinter.ttk import *
from tkinter.messagebox import showwarning, showinfo
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import os
from pykml import parser


class GUI():
    def __init__(self):
        self.window = Tk()

        self.window.title("Μετατροπή αρχείου σημείων kml σε xlsx")
        self.window.resizable(False, False)

        self.data = list()

        self.create_widgets()

    def setColsWidth(self, ws):
        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

    def saveFile(self):
        wb = Workbook()
        ws = wb.active

        for row in self.data:
            ws.append(row)

        self.setColsWidth(ws)

        outputFile = "output.xlsx"

        notSaved = True

        while notSaved:
            try:
                wb.save(outputFile)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message="Παρακαλώ κλείστε το αρχείο '{}' ώστε να ολοκληρωθεί η αποθήκευση.".format(
                                outputFile))
            else:
                notSaved = False

    def getDataFilename(self):
        fName = filedialog.askopenfilename(initialdir="./data/",
                                           title="Επιλέξτε το αρχείο kml",
                                           filetypes=(("kml files", "*.kml"), ("all files", "*.*")))

        if fName == "":
            return

        self.dataFilename.set(fName)
        self.ntrDataFilename.configure(state='disabled')
        self.btnOpenData.configure(state='disabled')

        with open(fName, encoding='utf-8') as f:
            doc = parser.parse(f).getroot()

        self.data.append(('Σχολείο', 'Longitude', 'Latitude'))
        for e in doc.Document.Folder.findall('.//{http://www.opengis.net/kml/2.2}Placemark'):
            coor = e.Point.coordinates.text.split(',')
            longitude = coor[0].replace('\n', '').strip()
            latitude = coor[1].replace('\n', '').strip()
            self.data.append((str(e.name), str(longitude), str(latitude)))

        self.btnRun.configure(state='normal')

    def run(self):
        self.btnRun.configure(state='disabled')
        self.saveFile()
        os.startfile("output.xlsx")
        showinfo(title="Ολοκλήρωση εκτέλεσης", message="Ο υπολογισμός ολοκληρώθηκε.")

    def create_widgets(self):
        self.fData = Frame(self.window)

        self.lData = Label(self.fData, text="Αρχείο kml:")
        self.lData.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.dataFilename = StringVar()
        self.ntrDataFilename = Entry(self.fData, width=128, state='readonly', textvariable=self.dataFilename)
        self.ntrDataFilename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenData = Button(self.fData, text="Επιλέξτε αρχείο...", command=self.getDataFilename)
        self.btnOpenData.grid(column=2, row=0, padx=10, pady=10)

        self.btnRun = Button(self.fData, text="Εκτέλεση", command=self.run, state='disabled')
        self.btnRun.grid(column=1, row=10, padx=10, pady=10)

        self.fData.pack()


gui = GUI()
gui.window.mainloop()
