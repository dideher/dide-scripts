from openpyxl import load_workbook, Workbook


def prepare():
    workbook = load_workbook(filename="./data/data.xlsx")
    sheet = workbook.active

    data = list()
    data.append(["Σχολείο", "Σύνολο 19-20", "Β-Γ Τάξη 20-21"])

    skipRow = True
    i = 1
    for cellSchool, cellClass, cellCount in zip(sheet['A'], sheet['B'], sheet['C']):
        if skipRow:
            skipRow = False
            continue

        if i == 1:
            school = cellSchool.value.strip()
            aClass = cellCount.value
            i = 2
        elif i == 2:
            bClass = cellCount.value
            i = 3
        elif i == 3:
            cClass = cellCount.value
            i = 1

            data.append([school, aClass + bClass + cClass, aClass + bClass])

    wb = Workbook()
    ws = wb.active

    for row in data:
        ws.append(row)

    wb.save("./data/prepared.xlsx")


def compare():
    workbook = load_workbook(filename="./data/prepared.xlsx")
    sheet = workbook.active

    data = list()
    data.append(["Σχολείο", "Σύνολο 19-20", "Β-Γ Τάξη 20-21", "Α Τάξη 20-21", "Σύνολο 20-21", "Διαφορά"])

    skipRow = True
    i = 1
    for cellSchool, cellSum19_20, cellBC20_21, cellA20_21 in zip(sheet['A'], sheet['B'], sheet['C'], sheet['D']):
        if skipRow:
            skipRow = False
            continue

        school = cellSchool.value
        sum19_20 = cellSum19_20.value
        bc20_21 = cellBC20_21.value
        a20_21 = cellA20_21.value
        sum20_21 = bc20_21 + a20_21
        diff = sum20_21 - sum19_20
        data.append([school, sum19_20, bc20_21, a20_21, sum20_21, diff])

    wb = Workbook()
    ws = wb.active

    for row in data:
        ws.append(row)

    wb.save("./data/compared.xlsx")


def main():
    #prepare()
    compare()

if __name__ == "__main__":
    main()