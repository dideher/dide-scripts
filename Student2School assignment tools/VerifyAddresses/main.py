from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from parseXlsxData import *
from openpyxl import *
from openpyxl.utils import get_column_letter


class GUI():
    def __init__(self):
        self.window = Tk()

        self.window.title("Διασταύρωση των διευθύνσεων των μαθητών")
        self.window.resizable(False, False)
        self.create_widgets()

    def getFilename(self):
        fName = filedialog.askopenfilename(initialdir=".", title="Επιλέξτε το αρχείο των μαθητών",
                                           filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if fName == "":
            return

        self.btnOpenFile.configure(state='disabled')
        self.studentsFilename.set(fName)
        self.students = parseXlsxData(fName)
        self.data = list()
        self.data.append(self.students[0].copy())
        self.data[0].append("Σχολείο Κατανομής")
        self.data[0].append("Διεύθυνση Κατανομής")
        self.data[0].append("Συντεταγμένες Κατανομής")
        self.currentStudent = 0
        self.cbStudentAddressCol['values'] = self.students[0]
        self.cbGoogleAddressCol['values'] = self.students[0]
        self.cbStudentAddressCol.configure(state='normal')

    def loadStudent(self):
        self.currentStudent += 1
        cs = self.currentStudent
        if cs < len(self.students):
            self.data.append(self.students[cs].copy())
            self.entryNum.set(cs)
            self.studentAddress.set(self.students[cs][self.sac])
            self.gmSchool.set(self.students[cs][self.gac - 1])
            self.gmAddress.set(self.students[cs][self.gac + 0])
            self.gv3School.set(self.students[cs][self.gac + 2])
            self.gv3Address.set(self.students[cs][self.gac + 3])
            self.bmSchool.set(self.students[cs][self.gac + 5])
            self.bmAddress.set(self.students[cs][self.gac + 6])
            self.hmSchool.set(self.students[cs][self.gac + 8])
            self.hmAddress.set(self.students[cs][self.gac + 9])
        else:
            showinfo(title="Ολοκλήρωση διασταύρωσης",
                     message="Η διαδικασία της διασταύρωσης των διευθύνσεων ολοκληρώθηκε.")
            self.entryNum.set("")
            self.entriesCount.set("")
            self.studentAddress.set("")
            self.gmAddress.set("")
            self.gmSchool.set("")
            self.gv3Address.set("")
            self.gv3School.set("")
            self.bmAddress.set("")
            self.bmSchool.set("")
            self.hmAddress.set("")
            self.hmSchool.set("")
            self.btnSaveFile.configure(state='disabled')
            self.btnGMAccept.configure(state='disabled')
            self.btnGV3Accept.configure(state='disabled')
            self.btnBMAccept.configure(state='disabled')
            self.btnHMAccept.configure(state='disabled')
            self.btnAddressNotFound.configure(state='disabled')
            self.saveFile(excludeLast=False)

    def addressNotFound(self):
        cs = self.currentStudent
        self.data[cs].append("N/A")
        self.data[cs].append("N/A")
        self.data[cs].append("N/A")

        self.loadStudent()

    def gmAccept(self):
        cs = self.currentStudent
        self.data[cs].append(self.students[cs][self.gac - 1])
        self.data[cs].append(self.students[cs][self.gac + 0])
        self.data[cs].append(self.students[cs][self.gac + 1])

        self.loadStudent()

    def gv3Accept(self):
        cs = self.currentStudent
        self.data[cs].append(self.students[cs][self.gac + 2])
        self.data[cs].append(self.students[cs][self.gac + 3])
        self.data[cs].append(self.students[cs][self.gac + 4])

        self.loadStudent()

    def bmAccept(self):
        cs = self.currentStudent
        self.data[cs].append(self.students[cs][self.gac + 5])
        self.data[cs].append(self.students[cs][self.gac + 6])
        self.data[cs].append(self.students[cs][self.gac + 7])

        self.loadStudent()

    def hmAccept(self):
        cs = self.currentStudent
        self.data[cs].append(self.students[cs][self.gac + 8])
        self.data[cs].append(self.students[cs][self.gac + 9])
        self.data[cs].append(self.students[cs][self.gac + 10])

        self.loadStudent()

    def saveFile(self, excludeLast=True):
        wb = Workbook()
        ws = wb.active

        if excludeLast:
            for student in self.data[:-1]:
                ws.append(student)
        else:
            for student in self.data:
                ws.append(student)

        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

        outputFile = self.studentsFilename.get().replace(".xlsx", "_output.xlsx")
        showinfo(title="Αρχείο εξόδου",
                 message="Η κατανομή θα αποθηκευτεί στο αρχείο: " + outputFile)

        notSaved = True

        while notSaved:
            try:
                wb.save(outputFile)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message="Παρακαλώ κλείστε το αρχείο '{}' ώστε να ολοκληρωθεί η αποθήκευση της νέας κατανομής.".format(
                                outputFile))
            else:
                notSaved = False

    def cbStudentAddressColSelect(self, eventObject):
        self.cbGoogleAddressCol.configure(state='normal')
        self.cbStudentAddressCol.configure(state='disabled')

    def cbGoogleAddressColSelect(self, eventObject):
        self.cbGoogleAddressCol.configure(state='disabled')
        self.entriesCount.set(len(self.students) - 1)
        self.sac = self.cbStudentAddressCol.current()
        self.gac = self.cbGoogleAddressCol.current()

        self.loadStudent()

    def create_widgets(self):
        self.fMain = Frame(self.window)

        self.lStudentsFile = Label(self.fMain, text="Αρχείο:")
        self.lStudentsFile.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.studentsFilename = StringVar()
        self.studentsFilename.set('')
        self.ntrStudentsFilename = Entry(self.fMain, width=128, state='readonly', textvariable=self.studentsFilename)
        self.ntrStudentsFilename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenFile = Button(self.fMain, text="Επιλέξτε αρχείο...", command=self.getFilename)
        self.btnOpenFile.grid(column=2, row=0, padx=10, pady=10)

        self.lStudentAddressCol = Label(self.fMain, text="Στήλη για διεύθυνση μαθητή:")
        self.lStudentAddressCol.grid(column=0, row=1, padx=10, pady=10, sticky=E)

        self.studentAddressCol = StringVar()
        self.cbStudentAddressCol = Combobox(self.fMain, width=40, textvariable=self.studentAddressCol, state='disabled')
        self.cbStudentAddressCol.bind("<<ComboboxSelected>>", self.cbStudentAddressColSelect)
        self.cbStudentAddressCol.grid(column=1, row=1, padx=10, pady=10, sticky='NSEW')

        self.lGoogleAddressCol = Label(self.fMain, text="Στήλη για διεύθυνση Google:")
        self.lGoogleAddressCol.grid(column=0, row=2, padx=10, pady=10, sticky=E)

        self.googleAddressCol = StringVar()
        self.cbGoogleAddressCol = Combobox(self.fMain, width=40, textvariable=self.googleAddressCol, state='disabled')
        self.cbGoogleAddressCol.bind("<<ComboboxSelected>>", self.cbGoogleAddressColSelect)
        self.cbGoogleAddressCol.grid(column=1, row=2, padx=10, pady=10, sticky='NSEW')

        self.lfFileInfo = LabelFrame(self.fMain, text="Πληροφορίες αρχείου")
        self.lfFileInfo.grid(column=0, row=3, columnspan=3, padx=10, pady=10)
        self.lEntryNum = Label(self.lfFileInfo, text="Α/Α:")
        self.lEntryNum.grid(column=0, row=0, padx=10, pady=5, sticky=E)
        self.entryNum = StringVar()
        self.ntrEntryNum = Entry(self.lfFileInfo, width=10, state='readonly', textvariable=self.entryNum)
        self.ntrEntryNum.grid(column=1, row=0, padx=10, pady=5, sticky=W)
        self.lEntriesCount = Label(self.lfFileInfo, text="Πλήθος εγγραφών:")
        self.lEntriesCount.grid(column=2, row=0, padx=10, pady=5, sticky=E)
        self.entriesCount = StringVar()
        self.ntrEntriesCount = Entry(self.lfFileInfo, width=10, state='readonly', textvariable=self.entriesCount)
        self.ntrEntriesCount.grid(column=3, row=0, padx=10, pady=5, sticky=W)
        self.btnSaveFile = Button(self.lfFileInfo, text="Αποθήκευση", command=self.saveFile)
        self.btnSaveFile.grid(column=4, row=0, padx=10, pady=10)

        self.lfStudentInfo = LabelFrame(self.fMain, text="Διεύθυνση μαθητή")
        self.lfStudentInfo.grid(column=0, row=4, columnspan=3, padx=10, pady=10, sticky=EW)
        self.studentAddress = StringVar()
        self.ntrStudentAddress = Entry(self.lfStudentInfo, width=160, state='readonly',
                                       textvariable=self.studentAddress)
        self.ntrStudentAddress.grid(column=0, row=0, padx=10, pady=5)

        self.lfGecodingInfo = LabelFrame(self.fMain, text="Διεύθυνση / κατανομή")
        self.lfGecodingInfo.grid(column=0, row=5, columnspan=3, padx=10, pady=10, sticky=EW)

        self.gmAddress = StringVar()
        self.ntrGMAddress = Entry(self.lfGecodingInfo, width=100, state='readonly', textvariable=self.gmAddress)
        self.ntrGMAddress.grid(column=0, row=0, padx=10, pady=5, sticky=W)
        self.gmSchool = StringVar()
        self.ntrGMSchool = Entry(self.lfGecodingInfo, width=40, state='readonly', textvariable=self.gmSchool)
        self.ntrGMSchool.grid(column=1, row=0, padx=10, pady=5)
        self.btnGMAccept = Button(self.lfGecodingInfo, text="Αποδοχή", command=self.gmAccept)
        self.btnGMAccept.grid(column=2, row=0, padx=10, pady=5)

        self.gv3Address = StringVar()
        self.ntrGV3Address = Entry(self.lfGecodingInfo, width=100, state='readonly', textvariable=self.gv3Address)
        self.ntrGV3Address.grid(column=0, row=1, padx=10, pady=5, sticky=W)
        self.gv3School = StringVar()
        self.ntrGV3School = Entry(self.lfGecodingInfo, width=40, state='readonly', textvariable=self.gv3School)
        self.ntrGV3School.grid(column=1, row=1, padx=10, pady=5)
        self.btnGV3Accept = Button(self.lfGecodingInfo, text="Αποδοχή", command=self.gv3Accept)
        self.btnGV3Accept.grid(column=2, row=1, padx=10, pady=5)

        self.bmAddress = StringVar()
        self.ntrBMAddress = Entry(self.lfGecodingInfo, width=100, state='readonly', textvariable=self.bmAddress)
        self.ntrBMAddress.grid(column=0, row=2, padx=10, pady=5, sticky=W)
        self.bmSchool = StringVar()
        self.ntrBMSchool = Entry(self.lfGecodingInfo, width=40, state='readonly', textvariable=self.bmSchool)
        self.ntrBMSchool.grid(column=1, row=2, padx=10, pady=5)
        self.btnBMAccept = Button(self.lfGecodingInfo, text="Αποδοχή", command=self.bmAccept)
        self.btnBMAccept.grid(column=2, row=2, padx=10, pady=5)

        self.hmAddress = StringVar()
        self.ntrHMAddress = Entry(self.lfGecodingInfo, width=100, state='readonly', textvariable=self.hmAddress)
        self.ntrHMAddress.grid(column=0, row=3, padx=10, pady=5, sticky=W)
        self.hmSchool = StringVar()
        self.ntrHMSchool = Entry(self.lfGecodingInfo, width=40, state='readonly', textvariable=self.hmSchool)
        self.ntrHMSchool.grid(column=1, row=3, padx=10, pady=5)
        self.btnHMAccept = Button(self.lfGecodingInfo, text="Αποδοχή", command=self.hmAccept)
        self.btnHMAccept.grid(column=2, row=3, padx=10, pady=5)

        self.btnAddressNotFound = Button(self.fMain, text="Η διεύθυνση δεν βρέθηκε", command=self.addressNotFound)
        self.btnAddressNotFound.grid(column=0, row=6, columnspan=3, padx=10, pady=5, sticky=EW)

        self.fMain.pack()


gui = GUI()
gui.window.mainloop()
