from tkinter import *
from tkinter import filedialog
from tkinter.ttk import *
from tkinter.messagebox import showwarning, showinfo
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import os
import geopy.distance
from pykml import parser


class GUI():
    def __init__(self):
        self.window = Tk()

        self.window.title("Υπολογισμός αποστάσεων")
        self.window.resizable(False, False)

        self.data1 = list()
        self.data2 = list()
        self.data3 = list()

        self.create_widgets()


    def setColsWidth(self, ws):
        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23


    def saveFile(self):
        wb = Workbook()
        ws = wb.active

        for row in self.data3:
            ws.append(row)

        self.setColsWidth(ws)

        outputFile = "output.xlsx"

        notSaved = True

        while notSaved:
            try:
                wb.save(outputFile)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message="Παρακαλώ κλείστε το αρχείο '{}' ώστε να ολοκληρωθεί η αποθήκευση.".format(
                                outputFile))
            else:
                notSaved = False


    def calc(self):
        title1 = self.dataFilename1.get().split('/')[-1:][0].replace(".kml", "").upper()
        title2 = self.dataFilename2.get().split('/')[-1:][0].replace(".kml", "").upper()

        self.data3.append((title1, title2, 'ΑΠΟΣΤΑΣΗ (ΚΜ)'))

        for item1 in self.data1:
            for item2 in self.data2:
                coords_1 = (item1[1], item1[2])
                coords_2 = (item2[1], item2[2])
                self.data3.append((str(item1[0]), str(item2[0]), str(geopy.distance.distance(coords_1, coords_2).km)))

        self.saveFile()
        os.startfile("output.xlsx")


    def getDataFilename1(self):
        fName = filedialog.askopenfilename(initialdir="./data/",
                                           title="Επιλέξτε το αρχείο kml",
                                           filetypes=(("kml files", "*.kml"), ("all files", "*.*")))

        if fName == "":
            return

        self.dataFilename1.set(fName)
        self.ntrDataFilename1.configure(state='disabled')
        self.btnOpenData1.configure(state='disabled')

        with open(fName, encoding='utf-8') as f:
            doc = parser.parse(f).getroot()

        for e in doc.Document.Folder.findall('.//{http://www.opengis.net/kml/2.2}Placemark'):
            coor = e.Point.coordinates.text.split(',')
            longitude = coor[0].replace('\n', '').strip()
            latitude = coor[1].replace('\n', '').strip()
            self.data1.append((e.name, longitude, latitude))

        self.ntrDataFilename2.configure(state='readonly')
        self.btnOpenData2.configure(state='normal')


    def getDataFilename2(self):
        fName = filedialog.askopenfilename(initialdir="./data/",
                                           title="Επιλέξτε το αρχείο kml",
                                           filetypes=(("kml files", "*.kml"), ("all files", "*.*")))

        if fName == "":
            return

        self.dataFilename2.set(fName)
        self.ntrDataFilename2.configure(state='disabled')
        self.btnOpenData2.configure(state='disabled')

        with open(fName, encoding='utf-8') as f:
            doc = parser.parse(f).getroot()

        for e in doc.Document.Folder.findall('.//{http://www.opengis.net/kml/2.2}Placemark'):
            coor = e.Point.coordinates.text.split(',')
            longitude = coor[0].replace('\n', '').strip()
            latitude = coor[1].replace('\n', '').strip()
            self.data2.append((e.name, longitude, latitude))

        self.btnRun.configure(state='normal')


    def run(self):
        self.btnRun.configure(state='disabled')
        self.calc()
        showinfo(title="Ολοκλήρωση εκτέλεσης", message="Ο υπολογισμός ολοκληρώθηκε.")


    def create_widgets(self):
        self.fData = Frame(self.window)

        self.lData1 = Label(self.fData, text="1ο αρχείο kml:")
        self.lData1.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.dataFilename1 = StringVar()
        self.ntrDataFilename1 = Entry(self.fData, width=128, state='readonly', textvariable=self.dataFilename1)
        self.ntrDataFilename1.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenData1 = Button(self.fData, text="Επιλέξτε αρχείο...", command=self.getDataFilename1)
        self.btnOpenData1.grid(column=2, row=0, padx=10, pady=10)

        self.lData2 = Label(self.fData, text="2ο αρχείο kml:")
        self.lData2.grid(column=0, row=1, padx=10, pady=10, sticky=E)

        self.dataFilename2 = StringVar()
        self.ntrDataFilename2 = Entry(self.fData, width=128, state='disabled', textvariable=self.dataFilename2)
        self.ntrDataFilename2.grid(column=1, row=1, padx=10, pady=10, sticky=W)

        self.btnOpenData2 = Button(self.fData, text="Επιλέξτε αρχείο...", command=self.getDataFilename2, state='disabled')
        self.btnOpenData2.grid(column=2, row=1, padx=10, pady=10)

        self.btnRun = Button(self.fData, text="Εκτέλεση", command=self.run, state='disabled')
        self.btnRun.grid(column=1, row=10, padx=10, pady=10)

        self.fData.pack()


gui = GUI()
gui.window.mainloop()
