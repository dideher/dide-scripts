from tkinter import *
from tkinter import filedialog
from tkinter.ttk import *
from tkinter.messagebox import showwarning, showinfo
from openpyxl.utils import get_column_letter
from openpyxl import *
import os
import geopy.distance


def sch_distance(item):
    return float(item[1])


def set_cols_width(ws):
    column_widths = []
    for row in ws.iter_rows():
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], len(str(cell.value)))
            except IndexError:
                column_widths.append(len(str(cell.value)))

    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23


class GUI:
    def __init__(self):
        self.window = Tk()

        self.window.title("Υπολογισμός αποστάσεων")
        self.window.resizable(False, False)

        self.schools_data = list()
        self.students_data = list()
        self.data3 = list()

        self.create_widgets()

    def save_file(self):
        wb = Workbook()
        ws = wb.active

        for row in self.data3:
            ws.append(row)

        set_cols_width(ws)

        output_file = "output.xlsx"

        not_saved = True

        while not_saved:
            try:
                wb.save(output_file)
            except:
                showwarning(title="Αρχείο σε χρήση...",
                            message=f"Παρακαλώ κλείστε το αρχείο '{output_file}' ώστε να ολοκληρωθεί η αποθήκευση.")
            else:
                not_saved = False

    def calc(self):
        self.data3.append(self.students_data[0])
        header = list()
        for i in range(3):
            header.append(f'Επιλογή {i + 1}')

        self.data3[0] += header

        for item2 in self.students_data[1:]:
            coords_2 = (item2[-2], item2[-1])
            schools = list()

            for item1 in self.schools_data[1:]:
                coords_1 = (item1[1], item1[2])

                schools.append([item1[0], str(geopy.distance.distance(coords_1, coords_2).km)])
            schools.sort(key=sch_distance)

            entry = item2
            for item in schools[:3]:
                entry += [f'{item[0]} ({item[1]})']

            self.data3.append(entry)

        self.save_file()
        os.startfile("output.xlsx")

    def parse_schools(self):
        wb = load_workbook(filename=self.schools_filename.get())
        ws = wb.active

        for row in ws.iter_rows():
            entry = list()
            for cell in row:
                if cell.value is None:
                    text = ""
                else:
                    text = str(cell.value).strip()

                entry.append(text)

            self.schools_data.append(entry)

    def parse_students(self):
        wb = load_workbook(filename=self.students_filename.get())
        sheet = wb.active

        for r, row in enumerate(sheet.iter_rows()):
            entry = list()
            for i, cell in enumerate(row):
                if cell.value is None:
                    text = ""
                else:
                    text = str(cell.value).strip()

                if i == len(row) - 1:
                    if r == 0:
                        entry.append('Longitude')
                        entry.append('Latitude')
                    else:
                        lat, lng = text.split(',')
                        entry.append(lng)
                        entry.append(lat)
                else:
                    entry.append(text)

            self.students_data.append(entry)

    def get_schools_data(self):
        f_name = filedialog.askopenfilename(initialdir="./data/",
                                            title="Επιλέξτε το αρχείο των σχολείων",
                                            filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if f_name == "":
            return

        self.schools_filename.set(f_name)
        self.ntr_schools_data.configure(state='disabled')
        self.btn_get_schools_data.configure(state='disabled')

        self.parse_schools()

        self.ntr_students_data.configure(state='readonly')
        self.btn_get_students_data.configure(state='normal')

    def get_students_data(self):
        f_name = filedialog.askopenfilename(initialdir="./data/",
                                            title="Επιλέξτε το αρχείο των μαθητών",
                                            filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if f_name == "":
            return

        self.students_filename.set(f_name)
        self.ntr_students_data.configure(state='disabled')
        self.btn_get_students_data.configure(state='disabled')

        self.parse_students()

        self.btn_run.configure(state='normal')

    def run(self):
        self.btn_run.configure(state='disabled')
        self.calc()
        showinfo(title="Ολοκλήρωση εκτέλεσης", message="Ο υπολογισμός ολοκληρώθηκε.")
        self.window.destroy()

    def create_widgets(self):
        self.f_data = Frame(self.window)

        self.l_schools_data = Label(self.f_data, text="Αρχείο σχολείων:")
        self.l_schools_data.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.schools_filename = StringVar()
        self.ntr_schools_data = Entry(self.f_data, width=128, state='readonly', textvariable=self.schools_filename)
        self.ntr_schools_data.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btn_get_schools_data = Button(self.f_data, text="Επιλέξτε αρχείο...", command=self.get_schools_data)
        self.btn_get_schools_data.grid(column=2, row=0, padx=10, pady=10)

        self.l_students_data = Label(self.f_data, text="Αρχείο μαθητών:")
        self.l_students_data.grid(column=0, row=1, padx=10, pady=10, sticky=E)

        self.students_filename = StringVar()
        self.ntr_students_data = Entry(self.f_data, width=128, state='disabled', textvariable=self.students_filename)
        self.ntr_students_data.grid(column=1, row=1, padx=10, pady=10, sticky=W)

        self.btn_get_students_data = Button(self.f_data, text="Επιλέξτε αρχείο...", command=self.get_students_data,
                                            state='disabled')
        self.btn_get_students_data.grid(column=2, row=1, padx=10, pady=10)

        self.btn_run = Button(self.f_data, text="Εκτέλεση", command=self.run, state='disabled')
        self.btn_run.grid(column=1, row=10, padx=10, pady=10)

        self.f_data.pack()


gui = GUI()
gui.window.mainloop()
