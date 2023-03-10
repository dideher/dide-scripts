from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from openpyxl import *
from openpyxl.utils import get_column_letter
from pathlib import Path
import shutil
from openpyxl import load_workbook
import re


def fix_cell(cell):
    text1 = str(cell.value).upper().replace(" .", ". ").replace("Ά", "Α").replace("Έ", "Ε") \
        .replace("Ή", "Η").replace("Ί", "Ι").replace("Ϊ́", "Ϊ").replace("Ύ", "Υ").replace("Ϋ́", "Ϋ") \
        .replace("Ό", "Ο").replace("Ώ", "Ω").strip()

    text2 = re.sub(r'([ ]+)', r' ', text1)
    text3 = re.sub(r'([0-9]+)Ο', r'\1ο', text2)

    return text3


def parse_students(file_name):
    wb = load_workbook(filename=file_name)
    ws = wb.active

    students = list()

    for row in ws.iter_rows():
        student = list()
        for cell in row:
            if cell.value is None:
                student.append("")
            else:
                text = fix_cell(cell)
                student.append(text)

        students.append(student)

    return students


def parse_schools(file_name):
    wb = load_workbook(filename=file_name)
    ws = wb.active

    schools = dict()

    rows = ws.iter_rows()
    next(rows)

    for row in rows:
        if row[0].value is not None and row[1].value is not None:
            origin_school = fix_cell(row[0])
            destination_school = fix_cell(row[1])

            schools[origin_school] = destination_school

    return schools


def save_file(data, output_file):
    wb = Workbook()
    ws = wb.active

    for entry in data:
        ws.append(entry)

    column_widths = []
    for row in ws.iter_rows():
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], len(str(cell.value)))
            except IndexError:
                column_widths.append(len(str(cell.value)))

    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

    not_saved = True

    while not_saved:
        try:
            wb.save(output_file)
        except:
            showwarning(title="Αρχείο σε χρήση...",
                        message=f"Παρακαλώ κλείστε το αρχείο '{output_file}' ώστε να ολοκληρωθεί η αποθήκευση.")
        else:
            not_saved = False


def create_folders():
    dirpath = Path("./reports")
    if dirpath.exists() and dirpath.is_dir():
        shutil.rmtree(dirpath)

    dirpath = Path("./reports/xtras")
    dirpath.mkdir(parents=True, exist_ok=True)

    dirpath = Path("./reports/_ok")
    dirpath.mkdir(parents=True, exist_ok=True)

    dirpath = Path("./reports/_check")
    dirpath.mkdir(parents=True, exist_ok=True)


class GUI:
    def __init__(self):
        self.window = Tk()

        self.window.title("Φιλτράρισμα της κατανομής των μαθητών")
        self.window.resizable(False, False)
        self.create_widgets()

    def get_students_filename(self):
        f_name = filedialog.askopenfilename(initialdir=".", title="Επιλέξτε το αρχείο της κατανομής των μαθητών",
                                            filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if f_name == "":
            return

        self.btn_get_students_file.configure(state='disabled')
        self.students_filename.set(f_name)
        self.students = parse_students(f_name)
        self.cb_school_col['values'] = self.students[0]
        self.cb_school_col.configure(state='readonly')
        self.cb_google_col['values'] = self.students[0]
        self.urban = list()
        self.rural = list()

    def get_schools_filename(self):
        f_name = filedialog.askopenfilename(initialdir=".", title="Επιλέξτε το αρχείο των σχολείων",
                                            filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if f_name == "":
            return

        self.btn_get_schools_file.configure(state='disabled')
        self.btn_run.configure(state='normal')
        self.schools_filename.set(f_name)
        self.schools = parse_schools(f_name)

    def clean_parentheses(self):
        gmc = self.cb_google_col.current()

        for student in self.students[1:]:
            for i in [0, 3, 6, 9]:
                if '(' in student[gmc + i]:
                    student[gmc + i] = re.sub("\(.*", "", student[gmc + i]).strip()

    def split_urban_rural(self):
        sc = self.cb_school_col.current()

        self.urban.append(self.students[0])
        self.rural.append(self.students[0] + ['ΣΧΟΛΕΙΟ ΚΑΤΑΝΟΜΗΣ', ])

        for student in self.students[1:]:
            if student[sc] in self.schools:
                self.rural.append(student + [self.schools[student[sc]], ])
            else:
                self.urban.append(student)

        urban_count = len(self.urban[1:])
        rural_count = len(self.rural[1:])
        save_file(self.urban, f"./reports/xtras/0-urban ({urban_count}).xlsx")
        save_file(self.rural, f"./reports/xtras/0-rural ({rural_count}).xlsx")

        print(20 * '-', ' Διαχωρισμός σε μαθητές Πόλης και Περιφέρειας ', 20 * '-')
        print('Πλήθος μαθητών Πόλης: ', urban_count)
        print('Πλήθος μαθητών Περιφέρειας: ', rural_count)
        print('Σύνολο: ', urban_count + rural_count)

    def check_rural(self):
        gmc = self.cb_google_col.current()

        self.rural_ok = list()
        self.rural_check = list()
        rural_check_1_or_2_matches = list()
        rural_check_0_matches = list()

        self.rural_ok.append(self.rural[0] + ['ΣΥΝΤΕΤΑΓΜΕΝΕΣ ΚΑΤΑΝΟΜΗΣ', ])
        self.rural_check.append(self.rural[0][:-1])
        rural_check_1_or_2_matches.append(self.rural[0] + ['ΠΑΡΑΤΗΡΗΣΕΙΣ', ])
        rural_check_0_matches.append(self.rural[0])

        for student in self.rural[1:]:
            count = 0
            coords = ''
            descr = ''

            for i in [0, 3, 6, 9]:
                if i == 0:
                    tag = '-G1-'
                elif i == 3:
                    tag = '-G2-'
                elif i == 6:
                    tag = '-B-'
                else:
                    tag = '-H-'

                if student[gmc + i] == student[gmc + 12]:
                    count += 1
                    coords = student[gmc + i + 2]
                    descr += tag

            if count >= 3:
                self.rural_ok.append(student + [coords, ])
            elif count > 0:
                self.rural_check.append(student[:-1])
                rural_check_1_or_2_matches.append(student + [descr, ])
            else:
                self.rural_check.append(student[:-1])
                rural_check_0_matches.append(student)

        ok_count = len(self.rural_ok[1:])
        check_count = len(self.rural_check[1:])
        check_count_1_or_2 = len(rural_check_1_or_2_matches[1:])
        check_count_0 = len(rural_check_0_matches[1:])
        save_file(self.rural_ok, f"./reports/_ok/1-rural_ok ({ok_count}).xlsx")
        save_file(self.rural_check, f"./reports/_check/1-rural_check ({check_count}).xlsx")
        save_file(rural_check_1_or_2_matches,
                  f"./reports/xtras/1-1-rural_check_1_or_2_matches ({check_count_1_or_2}).xlsx")
        save_file(rural_check_0_matches, f"./reports/xtras/1-2-rural_check_0_matches ({check_count_0}).xlsx")

        print(20 * '-', ' Έλεγχος σε μαθητές Περιφέρειας ', 20 * '-')
        print('Πλήθος μαθητών ΟΚ: ', ok_count)
        print('Πλήθος μαθητών με 2 ή λιγότερες αντιστοιχίσεις: ', check_count_1_or_2)
        print('Πλήθος μαθητών με 0 αντιστοιχίσεις: ', check_count_0)
        print('Πλήθος μαθητών για νέο έλεγχο: ', check_count)
        print('Σύνολο: ', ok_count + check_count)

    def check_urban(self):
        sc = self.cb_school_col.current()
        gmc = self.cb_google_col.current()

        self.urban_ok = list()
        self.urban_ok.append(self.urban[0] + ['ΣΧΟΛΕΙΟ ΚΑΤΑΝΟΜΗΣ', 'ΣΥΝΤΕΤΑΓΜΕΝΕΣ ΚΑΤΑΝΟΜΗΣ'])
        self.urban_check = list()
        self.urban_check.append(self.urban[0])

        count_4 = 0
        count_3 = 0

        only_postal_code = list()
        only_postal_code.append(self.urban[0])
        not_found = list()
        not_found.append(self.urban[0])
        gazi_drops = list()
        gazi_drops.append(self.urban[0])

        for student in self.urban[1:]:
            school = ''
            coords = ''
            success = False

            if student[gmc] == student[gmc + 3] == student[gmc + 6] == student[gmc + 9]:
                if student[gmc] != 'N/A':
                    success = True
                    count_4 += 1
                    school = student[gmc]
                    coords = student[gmc + 2]
                else:
                    not_found.append(student)
            elif student[gmc] == student[gmc + 3] == student[gmc + 6]:
                if student[gmc] != 'N/A':
                    if not student[gmc + 1].startswith(('ΗΡΑΚΛΕΙΟ', 'ΜΑΛΕΒΙΖΙ')) or not student[gmc + 4].startswith(
                            ('ΗΡΑΚΛΕΙΟ', 'ΜΑΛΕΒΙΖΙ')):
                        if student[sc] not in ['1ο ΟΛΟΗΜΕΡΟ ΔΗΜΟΤΙΚΟ ΣΧΟΛΕΙΟ ΑΓΙΑΣ ΜΑΡΙΝΑΣ-TEΜΠΟΝΕΡΑ',
                                               '2ο ΟΛΟΗΜΕΡΟ ΔΗΜΟΤΙΚΟ ΣΧΟΛΕΙΟ ΑΓΙΑΣ ΜΑΡΙΝΑΣ',
                                               '1ο ΟΛΟΗΜΕΡΟ ΔΗΜΟΤΙΚΟ ΣΧΟΛΕΙΟ ΓΑΖΙΟΥ',
                                               '2ο ΔΗΜΟΤΙΚΟ ΣΧΟΛΕΙΟ ΓΑΖΙΟΥ',
                                               '3ο ΔΗΜΟΤΙΚΟ ΣΧΟΛΕΙΟ ΓΑΖΙΟΥ']:
                            success = True
                            count_3 += 1
                            school = student[gmc]
                            coords = student[gmc + 2]
                        else:
                            if student[gmc] in ['ΓΑΖΙΟΥ', '8ο ΓΥΜΝΑΣΙΟ', '10ο ΓΥΜΝΑΣΙΟ', 'ΤΥΛΙΣΟΥ']:
                                success = True
                                count_3 += 1
                                school = student[gmc]
                                coords = student[gmc + 2]
                            else:
                                gazi_drops.append(student)
                    else:
                        only_postal_code.append(student)
                else:
                    not_found.append(student)
            elif student[gmc] == student[gmc + 3] == student[gmc + 9]:
                if student[gmc] != 'N/A':
                    if not student[gmc + 1].startswith(('ΗΡΑΚΛΕΙΟ', 'ΜΑΛΕΒΙΖΙ')) or not student[gmc + 4].startswith(
                            ('ΗΡΑΚΛΕΙΟ', 'ΜΑΛΕΒΙΖΙ')):
                        if student[sc] not in ['1ο ΟΛΟΗΜΕΡΟ ΔΗΜΟΤΙΚΟ ΣΧΟΛΕΙΟ ΑΓΙΑΣ ΜΑΡΙΝΑΣ-TEΜΠΟΝΕΡΑ',
                                               '2ο ΟΛΟΗΜΕΡΟ ΔΗΜΟΤΙΚΟ ΣΧΟΛΕΙΟ ΑΓΙΑΣ ΜΑΡΙΝΑΣ',
                                               '1ο ΟΛΟΗΜΕΡΟ ΔΗΜΟΤΙΚΟ ΣΧΟΛΕΙΟ ΓΑΖΙΟΥ',
                                               '2ο ΔΗΜΟΤΙΚΟ ΣΧΟΛΕΙΟ ΓΑΖΙΟΥ',
                                               '3ο ΔΗΜΟΤΙΚΟ ΣΧΟΛΕΙΟ ΓΑΖΙΟΥ']:
                            success = True
                            count_3 += 1
                            school = student[gmc]
                            coords = student[gmc + 2]
                        else:
                            if student[gmc] in ['ΓΑΖΙΟΥ', '8ο ΓΥΜΝΑΣΙΟ', '10ο ΓΥΜΝΑΣΙΟ', 'ΤΥΛΙΣΟΥ']:
                                success = True
                                count_3 += 1
                                school = student[gmc]
                                coords = student[gmc + 2]
                            else:
                                gazi_drops.append(student)
                    else:
                        only_postal_code.append(student)
                else:
                    not_found.append(student)
            elif student[gmc] == student[gmc + 6] == student[gmc + 9]:
                if student[gmc] != 'N/A':
                    if not student[gmc + 1].startswith(('ΗΡΑΚΛΕΙΟ', 'ΜΑΛΕΒΙΖΙ')):
                        if student[sc] not in ['1ο ΟΛΟΗΜΕΡΟ ΔΗΜΟΤΙΚΟ ΣΧΟΛΕΙΟ ΑΓΙΑΣ ΜΑΡΙΝΑΣ-TEΜΠΟΝΕΡΑ',
                                               '2ο ΟΛΟΗΜΕΡΟ ΔΗΜΟΤΙΚΟ ΣΧΟΛΕΙΟ ΑΓΙΑΣ ΜΑΡΙΝΑΣ',
                                               '1ο ΟΛΟΗΜΕΡΟ ΔΗΜΟΤΙΚΟ ΣΧΟΛΕΙΟ ΓΑΖΙΟΥ',
                                               '2ο ΔΗΜΟΤΙΚΟ ΣΧΟΛΕΙΟ ΓΑΖΙΟΥ',
                                               '3ο ΔΗΜΟΤΙΚΟ ΣΧΟΛΕΙΟ ΓΑΖΙΟΥ']:
                            success = True
                            count_3 += 1
                            school = student[gmc]
                            coords = student[gmc + 2]
                        else:
                            if student[gmc] in ['ΓΑΖΙΟΥ', '8ο ΓΥΜΝΑΣΙΟ', '10ο ΓΥΜΝΑΣΙΟ', 'ΤΥΛΙΣΟΥ']:
                                success = True
                                count_3 += 1
                                school = student[gmc]
                                coords = student[gmc + 2]
                            else:
                                gazi_drops.append(student)
                    else:
                        only_postal_code.append(student)
                else:
                    not_found.append(student)
            elif student[gmc + 3] == student[gmc + 6] == student[gmc + 9]:
                if student[gmc + 3] != 'N/A':
                    if not student[gmc + 4].startswith(('ΗΡΑΚΛΕΙΟ', 'ΜΑΛΕΒΙΖΙ')):
                        if student[sc] not in ['1ο ΟΛΟΗΜΕΡΟ ΔΗΜΟΤΙΚΟ ΣΧΟΛΕΙΟ ΑΓΙΑΣ ΜΑΡΙΝΑΣ-TEΜΠΟΝΕΡΑ',
                                               '2ο ΟΛΟΗΜΕΡΟ ΔΗΜΟΤΙΚΟ ΣΧΟΛΕΙΟ ΑΓΙΑΣ ΜΑΡΙΝΑΣ',
                                               '1ο ΟΛΟΗΜΕΡΟ ΔΗΜΟΤΙΚΟ ΣΧΟΛΕΙΟ ΓΑΖΙΟΥ',
                                               '2ο ΔΗΜΟΤΙΚΟ ΣΧΟΛΕΙΟ ΓΑΖΙΟΥ',
                                               '3ο ΔΗΜΟΤΙΚΟ ΣΧΟΛΕΙΟ ΓΑΖΙΟΥ']:
                            success = True
                            count_3 += 1
                            school = student[gmc + 3]
                            coords = student[gmc + 5]
                        else:
                            if student[gmc + 3] in ['ΓΑΖΙΟΥ', '8ο ΓΥΜΝΑΣΙΟ', '10ο ΓΥΜΝΑΣΙΟ', 'ΤΥΛΙΣΟΥ']:
                                success = True
                                count_3 += 1
                                school = student[gmc + 3]
                                coords = student[gmc + 5]
                            else:
                                gazi_drops.append(student)
                    else:
                        only_postal_code.append(student)
                else:
                    not_found.append(student)

            if success:
                self.urban_ok.append(student + [school, coords])
            else:
                self.urban_check.append(student)

        ok_count = len(self.urban_ok[1:])
        check_count = len(self.urban_check[1:])
        count_only_postal_code = len(only_postal_code[1:])
        count_not_found = len(not_found[1:])
        count_gazi_drops = len(gazi_drops[1:])
        save_file(self.urban_ok, f"./reports/_ok/2-urban_ok ({ok_count}).xlsx")
        save_file(self.urban_check, f"./reports/_check/2-urban_check ({check_count}).xlsx")
        save_file(only_postal_code,
                  f"./reports/xtras/2-1-urban_check_only_postal_code ({count_only_postal_code}).xlsx")
        save_file(not_found, f"./reports/xtras/2-2-urban_check_not_found ({count_not_found}).xlsx")
        save_file(gazi_drops, f"./reports/xtras/2-3-urban_check_gazi_drops ({count_gazi_drops}).xlsx")

        print(20 * '-', ' Έλεγχος σε μαθητές Πόλης ', 20 * '-')
        print('Πλήθος μαθητών ΟΚ: ', ok_count)
        print('Πλήθος μαθητών για νέο έλεγχο: ', check_count)
        print('Διευθύνσεις που εντοπίστηκαν μόνο με το ΤΚ: ', count_only_postal_code)
        print('Πλήθος σχολείων που δεν εντοπίστηκαν: ', count_not_found)
        print('Ασυμφωνία με Γυμνάσια περιοχής Γαζίου: ', count_gazi_drops)
        print('Σύνολο: ', ok_count + check_count)

    def merge(self):
        # ok = list()
        # check = list()

        ok = self.rural_ok[:] + self.urban_ok[1:]
        check = self.rural_check[:] + self.urban_check[1:]

        ok_count = len(ok[1:])
        check_count = len(check[1:])

        save_file(ok, f"./reports/ok ({ok_count}).xlsx")
        save_file(check, f"./reports/check ({check_count}).xlsx")

        print(20 * '-', ' Συνενώσεις ', 20 * '-')
        print('Πλήθος μαθητών ΟΚ: ', ok_count)
        print('Πλήθος μαθητών για νέο έλεγχο: ', check_count)
        print('Σύνολο: ', ok_count + check_count)

    def run(self):
        self.btn_run.configure(state='disabled')

        create_folders()

        self.clean_parentheses()
        self.split_urban_rural()
        self.check_rural()
        self.check_urban()
        self.merge()

        showinfo(title="Αρχεία εξόδου",
                 message=f'Τα αποτελέσματα έχουν αποθηκευτεί στον φάκελο "reports".')

        self.window.destroy()

    def cb_school_col_select(self, event_object):
        if self.cb_school_col.current() != -1:
            self.cb_school_col.configure(state='disabled')
            self.cb_google_col.configure(state='readonly')

    def cb_google_col_select(self, event_object):
        if self.cb_google_col.current() != -1:
            self.cb_google_col.configure(state='disabled')
            self.ntr_schools_filename.configure(state='readonly')
            self.btn_get_schools_file.configure(state='normal')

    def create_widgets(self):
        self.f_data = Frame(self.window)

        self.l_students_file = Label(self.f_data, text="Αρχείο κατανομής μαθητών:")
        self.l_students_file.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.students_filename = StringVar()
        self.students_filename.set('')
        self.ntr_students_filename = Entry(self.f_data, width=128, state='readonly',
                                           textvariable=self.students_filename)
        self.ntr_students_filename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btn_get_students_file = Button(self.f_data, text="Επιλέξτε αρχείο...", command=self.get_students_filename)
        self.btn_get_students_file.grid(column=2, row=0, padx=10, pady=10)

        self.l_school_col = Label(self.f_data, text="Στήλη Σχολείου Προέλευσης:")
        self.l_school_col.grid(column=0, row=1, padx=10, pady=10, sticky=E)

        self.school_col = StringVar()
        self.cb_school_col = Combobox(self.f_data, width=125, textvariable=self.school_col, state='disabled')
        self.cb_school_col.grid(column=1, row=1, padx=10, pady=10, sticky=W)
        self.cb_school_col.bind("<<ComboboxSelected>>", self.cb_school_col_select)

        self.l_google_col = Label(self.f_data, text="Στήλη Σχολείου κατανομής Google Maps:")
        self.l_google_col.grid(column=0, row=2, padx=10, pady=10, sticky=E)

        self.google_col = StringVar()
        self.cb_google_col = Combobox(self.f_data, width=125, textvariable=self.google_col, state='disabled')
        self.cb_google_col.bind("<<ComboboxSelected>>", self.cb_google_col_select)
        self.cb_google_col.grid(column=1, row=2, padx=10, pady=10, sticky='W')

        self.l_schools_file = Label(self.f_data, text="Αρχείο Σχολείων Περιφέρειας:")
        self.l_schools_file.grid(column=0, row=3, padx=10, pady=10, sticky=E)

        self.schools_filename = StringVar()
        self.schools_filename.set('')
        self.ntr_schools_filename = Entry(self.f_data, width=128, state='disabled', textvariable=self.schools_filename)
        self.ntr_schools_filename.grid(column=1, row=3, padx=10, pady=10, sticky=W)

        self.btn_get_schools_file = Button(self.f_data, text="Επιλέξτε αρχείο...", command=self.get_schools_filename,
                                           state='disabled')
        self.btn_get_schools_file.grid(column=2, row=3, padx=10, pady=10)

        self.btn_run = Button(self.f_data, text="Εκτέλεση", command=self.run, state='disabled')
        self.btn_run.grid(column=1, row=10, padx=10, pady=10)

        self.f_data.pack()


gui = GUI()
gui.window.mainloop()
