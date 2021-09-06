from tkinter import *
from tkinter import filedialog
from tkinter.scrolledtext import ScrolledText
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from parseMapData import createLists
from schoolsPolygons import *
from parseXlsxData import *
from openpyxl import *
from openpyxl.utils import get_column_letter


class GUI():
    def __init__(self):
        self.window = Tk()

        self.window.title("Κατανομή μαθητών σε Σχολικές Μονάδες")
        self.window.resizable(False, False)
        self.create_widgets()


    def getMapDataFilename(self):
        fName = filedialog.askopenfilename(initialdir="./data/", title="Επιλέξτε το αρχείο με τις περιοχές των Σχολικών Μονάδων",
                                           filetypes=(("kml files", "*.kml"), ("all files", "*.*")))

        if fName == "":
            return

        self.mapDataFilename.set(fName)
        self.namesList, self.polygonsList = createLists(fName)

        sameSize, debugText = checkListsLengths(self.namesList, self.polygonsList)
        self.stDebug.insert(INSERT, debugText)
        polygonsAreValid, debugText = checkPolygonsValidity(self.namesList, self.polygonsList)
        self.stDebug.insert(INSERT, debugText)
        noIntersections, debugText = checkPolygonsIntersections(self.namesList, self.polygonsList)
        self.stDebug.insert(INSERT, debugText)

        if sameSize:
            self.lCount.configure(background='green')
        else:
            self.lCount.configure(background='red')

        if polygonsAreValid:
            self.lPolygonsValidity.configure(background='green')
        else:
            self.lPolygonsValidity.configure(background='red')

        if noIntersections:
            self.lPolygonsIntersections.configure(background='green')
        else:
            self.lPolygonsIntersections.configure(background='red')

        self.btnNextGoToStudents.configure(state='normal')


    def nextGoToTabStudents(self):
        self.tabControl.tab(0, state='disabled')
        self.tabControl.tab(1, state='normal')
        self.tabControl.select(1)


    def getStudentsFilename(self):
        fName = filedialog.askopenfilename(initialdir="./data/", title="Επιλέξτε το αρχείο με τους μαθητές",
                                           filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if fName == "":
            return

        self.studentsFilename.set(fName)
        self.students = parseXlsxData(fName)
        self.cbStudentsCols['values'] = self.students[0]
        self.studentsCount.set(len(self.students) - 1)
        self.pbProgress['maximum'] = self.studentsCount.get()


    def cbStudentsColsSelect(self, eventObject):
        self.studentPointCol = self.cbStudentsCols.current()

        if self.cbStudentsCols.current() != -1:
            self.btnNextGoToRun.configure(state='normal')


    def prevGoToTabMapData(self):
        self.tabControl.tab(1, state='disabled')
        self.tabControl.tab(0, state='normal')
        self.tabControl.select(0)


    def prevGoToTabStudents(self):
        self.tabControl.tab(2, state='disabled')
        self.tabControl.tab(1, state='normal')
        self.tabControl.select(1)


    def nextGoToTabRun(self):
        self.tabControl.tab(1, state='disabled')
        self.tabControl.tab(2, state='normal')
        self.tabControl.select(2)


    def run(self):
        self.lProgress.configure(text="Εκτέλεση κατανομής σε εξέλιξη...")

        data = list()

        for student in self.students:
            data.append(student.copy())

        for i in range(len(data)):
            if i == 0:
                data[i].append("Σχολείο")
                continue

            self.lProgress.configure(text="Εκτέλεση κατανομής σε εξέλιξη... ({}/{})".format(i, self.studentsCount.get()))

            x, y = self.students[i][self.studentPointCol].replace(' ', '').split(',')
            point = Point(float(x), float(y))
            school = searchPointInsidePolygons(point, self.namesList, self.polygonsList)
            if not school:
                schoolText = "Η διεύθυνση εντοπίζεται σε όριο περιοχών."
            else:
                schoolText = str(school).replace("[\'", "").replace("\']", "")

            data[i].append(schoolText)

            self.pbProgress['value'] = i
            self.pbProgress.update()

        self.lProgress.configure(text="Ολοκλήρωση κατανομής.\nΑποθήκευση σε αρχείο...")

        wb = Workbook()
        ws = wb.active

        for student in data:
            ws.append(student)

        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

        outputFile = self.studentsFilename.get().replace(".xlsx", "_output.xlsx")
        showinfo(title="Αρχείο εξόδου",
                    message="Η κατανομή θα αποθηκευτεί στο αρχείο: " + outputFile)

        notSaved = True

        while notSaved:
            try:
                wb.save(outputFile)
            except:
                showwarning(title="Αρχείο σε χρήση...", message="Παρακαλώ κλείστε το αρχείο '{}' ώστε να ολοκληρωθεί η αποθήκευση της νέας κατανομής.".format(outputFile))
            else:
                notSaved = False

        self.lProgress.configure(text="Ολοκλήρωση όλων των εργασιών.")
        self.pbProgress['value'] = 0
        self.pbProgress.update()


    def create_widgets(self):
        # Tabs
        self.tabControl = Notebook(self.window)
        self.tabMapData = Frame(self.tabControl)
        self.tabControl.add(self.tabMapData, text="Περιοχές Σχολικών Μονάδων")
        self.tabStudents = Frame(self.tabControl)
        self.tabControl.add(self.tabStudents, text="Μαθητές", state='disabled')
        self.tabRun = Frame(self.tabControl)
        self.tabControl.add(self.tabRun, text="Εκτέλεση κατανομής", state='disabled')
        self.tabControl.pack(expand=1, fill="both")

        # Tab: MapData
        self.fMapData = Frame(self.tabMapData)

        self.lMapData = Label(self.fMapData, text="Αρχείο:")
        self.lMapData.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.mapDataFilename = StringVar()
        self.ntrMapDataFilename = Entry(self.fMapData, width=128, state='readonly', textvariable=self.mapDataFilename)
        self.ntrMapDataFilename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenMapData = Button(self.fMapData, text="Επιλέξτε αρχείο...", command=self.getMapDataFilename)
        self.btnOpenMapData.grid(column=2, row=0, padx=10, pady=10)

        self.lfVerifyFrame = LabelFrame(self.fMapData, text="Έλεγχοι")
        self.lfVerifyFrame.grid(column=0, row=1, columnspan=3, padx=10, pady=10)

        self.lCount = Label(self.lfVerifyFrame, text="Πλήθος ονομάτων == Πλήθος πολυγώνων")
        self.lCount.grid(column=0, row=0, padx=10, pady=10)

        self.lPolygonsValidity = Label(self.lfVerifyFrame, text="Ορθότητα πολυγώνων")
        self.lPolygonsValidity.grid(column=0, row=1, padx=10, pady=10)

        self.lPolygonsIntersections = Label(self.lfVerifyFrame, text="Επικαλύψεις πολυγώνων")
        self.lPolygonsIntersections.grid(column=0, row=2, padx=10, pady=10)

        self.stDebug = ScrolledText(self.lfVerifyFrame, width=80, height=10, wrap=WORD)
        self.stDebug.grid(column=1, row=0, columnspan=2, rowspan=3, padx=10, pady=10)

        self.btnNextGoToStudents = Button(self.fMapData, text="Επόμενο", command=self.nextGoToTabStudents, state='disabled')
        self.btnNextGoToStudents.grid(column=2, row=10, padx=10, pady=10)

        self.fMapData.pack()

        # Tab: Students
        self.fStudents = Frame(self.tabStudents)

        self.lStudents = Label(self.fStudents, text="Αρχείο:")
        self.lStudents.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.studentsFilename = StringVar()
        self.ntrStudentsFilename = Entry(self.fStudents, width=128, state='readonly', textvariable=self.studentsFilename)
        self.ntrStudentsFilename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenStudents = Button(self.fStudents, text="Επιλέξτε αρχείο...", command=self.getStudentsFilename)
        self.btnOpenStudents.grid(column=2, row=0, padx=10, pady=10)

        self.lPointCol = Label(self.fStudents, text="Στήλη Συντεταγμένων:")
        self.lPointCol.grid(column=0, row=1, padx=10, pady=10, sticky=E)

        self.studentsCols = StringVar()
        self.cbStudentsCols = Combobox(self.fStudents, width=40, textvariable=self.studentsCols, state='readonly')
        self.cbStudentsCols.grid(column=1, row=1, padx=10, pady=10, sticky=W)
        self.cbStudentsCols.bind("<<ComboboxSelected>>", self.cbStudentsColsSelect)

        self.lStudentsCount = Label(self.fStudents, text="Πλήθος μαθητών:")
        self.lStudentsCount.grid(column=0, row=2, padx=10, pady=10, sticky=E)

        self.studentsCount = StringVar()
        self.ntrStudentsCount = Entry(self.fStudents, width=12, state='readonly', textvariable=self.studentsCount)
        self.ntrStudentsCount.grid(column=1, row=2, padx=10, pady=10, sticky=W)

        self.btnPrevGoToMapData = Button(self.fStudents, text="Προηγούμενο", command=self.prevGoToTabMapData)
        self.btnPrevGoToMapData.grid(column=0, row=10, padx=10, pady=10)

        self.btnNextGoToRun = Button(self.fStudents, text="Επόμενο", command=self.nextGoToTabRun, state='disabled')
        self.btnNextGoToRun.grid(column=2, row=10, padx=10, pady=10)

        self.fStudents.pack()

        # Tab: Run
        self.fRun = Frame(self.tabRun)

        self.lProgress = Label(self.fRun, text="Αναμονή για εκτέλεση κατανομής ...")
        self.lProgress.grid(column=0, row=0, columnspan=3, padx=10, pady=10)

        self.pbProgress = Progressbar(self.fRun, orient='horizontal', length=300, mode='determinate')
        self.pbProgress.grid(column=0, row=1, columnspan=3, padx=10, pady=10)

        self.btnPrevGoToStudents = Button(self.fRun, text="Προηγούμενο", command=self.prevGoToTabStudents)
        self.btnPrevGoToStudents.grid(column=0, row=10, padx=10, pady=10)

        self.btnRun = Button(self.fRun, text="Εκτέλεση κατανομής", command=self.run)
        self.btnRun.grid(column=2, row=10, padx=10, pady=10)

        self.fRun.pack()


gui = GUI()
gui.window.mainloop()
