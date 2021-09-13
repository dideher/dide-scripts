from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import csv


class GUI():
    def __init__(self):
        self.window = Tk()

        self.window.title("Επαλήθευση καταχωρίσεων Λυκείων πόλης στο e-eggrafes")
        self.window.resizable(False, False)
        self.create_widgets()

    def getCsvFilename(self):
        fName = filedialog.askopenfilename(initialdir=".", title="Επιλέξτε το αρχείο csv",
                                           filetypes=(("csv files", "*.csv"), ("all files", "*.*")))

        if fName == "":
            return

        self.btnOpenCsvFile.configure(state='disabled')
        self.csvFilename.set(fName)

        self.csvData = self.parseCsvData(fName)
        self.ntrXlsxFilename.configure(state='readonly')
        self.btnOpenXlsxFile.configure(state='normal')

    def getXlsxFilename(self):
        fName = filedialog.askopenfilename(initialdir=".", title="Επιλέξτε το αρχείο xlsx",
                                           filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if fName == "":
            return

        self.btnOpenXlsxFile.configure(state='disabled')
        self.xlsxFilename.set(fName)

        self.xlsxData = self.parseXlsxData(fName)
        self.ntrShort2FullNameFile.configure(state='readonly')
        self.btnOpenShort2FullNameFile.configure(state='normal')

    def getShort2FullNameFile(self):
        fName = filedialog.askopenfilename(initialdir=".", title="Επιλέξτε το αρχείο xlsx",
                                           filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if fName == "":
            return

        self.btnOpenShort2FullNameFile.configure(state='disabled')
        self.short2fullNameFilename.set(fName)

        self.short2fullNameData = self.parseXlsxData(fName)
        self.btnRun.configure(state='normal')

    def parseXlsxData(self, fileName):
        workbook = load_workbook(filename=fileName)
        sheet = workbook.active

        data = list()

        for row in sheet.iter_rows():
            entry = list()
            for cell in row:
                if cell.value is None:
                    entry.append("")
                else:
                    text1 = (
                        str(cell.value).upper().replace(".", ". ").replace(" .", ". ").replace("Ά", "Α").
                            replace("Έ", "Ε").replace("Ή", "Η").replace("Ί", "Ι").replace("Ϊ́", "Ϊ").replace("Ύ", "Υ").
                            replace("Ϋ́", "Ϋ").replace("Ό", "Ο").replace("Ώ", "Ω").strip())

                    text2 = re.sub(r'([ ]+)', r' ', text1)
                    entry.append(re.sub(r'([0-9]+)Ο', r'\1ο', text2))

            data.append(entry)

        return data[1:]

    def parseCsvData(self, inputFile):
        data = list()

        with open(inputFile, 'rt', encoding='utf-8-sig') as f:
            reader = csv.reader(f, quotechar="'")

            for row in reader:
                data.append(row)

        return data[1:]

    def check(self):
        notFoundErrors = ''

        shortToFullLectic = { entry[0] : entry[1] for entry in self.short2fullNameData}

        for csvEntry in self.csvData:
            found = False
            for xlsxEntry in self.xlsxData:
                if csvEntry[0] == xlsxEntry[0]:
                    found = True
                    if csvEntry[5] == 'Μουσικό Σχολείο ΗρακλείουΓυμνάσιο με Λυκειακές Τάξεις':
                        csvEntry[5] = 'ΜΟΥΣΙΚΟ ΣΧΟΛΕΙΟ ΗΡΑΚΛΕΙΟΥΓΥΜΝΑΣΙΟ ΜΕ ΛΥΚΕΙΑΚΕΣ ΤΑΞΕΙΣ'

                    if xlsxEntry[-1] not in shortToFullLectic:
                        print(f'Ασυμφωνία: E-eggrafes {csvEntry} <--> Κατανομή {xlsxEntry}')
                    elif csvEntry[5] != shortToFullLectic[xlsxEntry[-1]]:
                        print(f'Ασυμφωνία: E-eggrafes {csvEntry} <--> Κατανομή {xlsxEntry}')
                    break

            if not found:
                notFoundErrors += f'Δεν βρέθηκε: {csvEntry}\n'

        if notFoundErrors != '':
            print(32 * '-', ' Δεν βρέθηκαν ', 32 * '-')
            print(notFoundErrors)

    def run(self):
        self.btnRun.configure(state='disabled')

        self.check()

        showinfo(title='Ολοκλήρωση Εκτέλεσης',
                 message=f'Ο έλεγχος ολοκληρώθηκε.')

    def create_widgets(self):
        self.fData = Frame(self.window)

        self.lCsvFile = Label(self.fData, text="Αρχείο αναφοράς e-eggrafes (csv):")
        self.lCsvFile.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.csvFilename = StringVar()
        self.csvFilename.set('')
        self.ntrCsvFilename = Entry(self.fData, width=128, state='readonly', textvariable=self.csvFilename)
        self.ntrCsvFilename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenCsvFile = Button(self.fData, text="Επιλέξτε αρχείο...", command=self.getCsvFilename)
        self.btnOpenCsvFile.grid(column=2, row=0, padx=10, pady=10)

        self.lXlsxFile = Label(self.fData, text="Αρχείο κατανομής (xlsx):")
        self.lXlsxFile.grid(column=0, row=1, padx=10, pady=10, sticky=E)

        self.xlsxFilename = StringVar()
        self.xlsxFilename.set('')
        self.ntrXlsxFilename = Entry(self.fData, width=128, state='disabled', textvariable=self.xlsxFilename)
        self.ntrXlsxFilename.grid(column=1, row=1, padx=10, pady=10, sticky=W)

        self.btnOpenXlsxFile = Button(self.fData, text="Επιλέξτε αρχείο...", state='disabled',
                                      command=self.getXlsxFilename)
        self.btnOpenXlsxFile.grid(column=2, row=1, padx=10, pady=10)

        self.lShort2FullNameFile = Label(self.fData,
                                         text="Αρχείο αντιστοίχισης σύντομου σε πλήρες\nονόματος σχολείου (xlsx):")
        self.lShort2FullNameFile.grid(column=0, row=2, padx=10, pady=10, sticky=E)

        self.short2fullNameFilename = StringVar()
        self.short2fullNameFilename.set('')
        self.ntrShort2FullNameFile = Entry(self.fData, width=128, state='disabled',
                                           textvariable=self.short2fullNameFilename)
        self.ntrShort2FullNameFile.grid(column=1, row=2, padx=10, pady=10, sticky=W)

        self.btnOpenShort2FullNameFile = Button(self.fData, text="Επιλέξτε αρχείο...", state='disabled',
                                                command=self.getShort2FullNameFile)
        self.btnOpenShort2FullNameFile.grid(column=2, row=2, padx=10, pady=10)

        self.btnRun = Button(self.fData, text="Εκτέλεση", command=self.run, state='disabled')
        self.btnRun.grid(column=1, row=10, padx=10, pady=10)

        self.fData.pack()


gui = GUI()
gui.window.mainloop()
