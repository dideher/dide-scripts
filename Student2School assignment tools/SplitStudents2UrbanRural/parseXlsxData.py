from openpyxl import load_workbook
import re


def fixCell(cell):
    text1 = (str(cell.value).upper().replace(" .", ". ").replace("Ά", "Α").replace("Έ", "Ε")
             .replace("Ή", "Η").replace("Ί", "Ι").replace("Ϊ́", "Ϊ").replace("Ύ", "Υ").replace("Ϋ́", "Ϋ")
             .replace("Ό", "Ο").replace("Ώ", "Ω").strip())

    text2 = re.sub(r'([ ]+)', r' ', text1)
    text3 = re.sub(r'([0-9]+)Ο', r'\1ο', text2)

    return text3


def parseStudents(fileName):
    workbook = load_workbook(filename=fileName)
    sheet = workbook.active

    students = list()

    for row in sheet.iter_rows():
        student = list()
        for cell in row:
            if cell.value is None:
                student.append("")
            else:
                text = fixCell(cell)
                student.append(text)

        students.append(student)

    return students


def parseSchools(fileName):
    workbook = load_workbook(filename=fileName)
    sheet = workbook.active

    schools = dict()

    rows = sheet.iter_rows()
    next(rows)

    for row in rows:
        if row[0].value is not None and row[1].value is not None:
            origin_school = fixCell(row[0])
            destination_school = fixCell(row[1])

            schools[origin_school] = destination_school

    return schools