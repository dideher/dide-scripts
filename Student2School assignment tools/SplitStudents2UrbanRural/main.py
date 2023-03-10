from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from openpyxl import *
from openpyxl.utils import get_column_letter
import re


def fix_cell(cell):
    text1 = (str(cell.value).upper().replace(" .", ". ").replace("Ά", "Α").replace("Έ", "Ε")
             .replace("Ή", "Η").replace("Ί", "Ι").replace("Ϊ́", "Ϊ").replace("Ύ", "Υ").replace("Ϋ́", "Ϋ")
             .replace("Ό", "Ο").replace("Ώ", "Ω").strip())

    text2 = re.sub(r'([ ]+)', r' ', text1)
    text3 = re.sub(r'([0-9]+)Ο', r'\1ο', text2)

    return text3


def parse_students(filename):
    wb = load_workbook(filename=filename)
    sheet = wb.active

    students = list()

    for row in sheet.iter_rows():
        student = list()
        for cell in row:
            if cell.value is None:
                student.append("")
            else:
                text = fix_cell(cell)
                student.append(text)

        students.append(student)

    return students


def parse_schools(filename):
    wb = load_workbook(filename=filename)
    sheet = wb.active

    schools = dict()

    rows = sheet.iter_rows()
    next(rows)

    for row in rows:
        if row[0].value is not None and row[1].value is not None:
            origin_school = fix_cell(row[0])
            destination_school = fix_cell(row[1])

            schools[origin_school] = destination_school

    return schools


def save_file(data, output_file):
    wb = Workbook()
    ws = wb.active

    for entry in data:
        ws.append(entry)

    column_widths = []
    for row in ws.iter_rows():
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], len(str(cell.value)))
            except IndexError:
                column_widths.append(len(str(cell.value)))

    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

    not_saved = True

    while not_saved:
        try:
            wb.save(output_file)
        except:
            showwarning(title="Αρχείο σε χρήση...",
                        message=f"Παρακαλώ κλείστε το αρχείο '{output_file}' ώστε να ολοκληρωθεί η αποθήκευση.")
        else:
            not_saved = False


class GUI:
    def __init__(self):
        self.window = Tk()

        self.window.title("Διαχωρισμός μαθητών σε πόλης και περιφέρειας")
        self.window.resizable(False, False)
        self.create_widgets()

    def get_students_filename(self):
        f_name = filedialog.askopenfilename(initialdir=".", title="Επιλέξτε το αρχείο των μαθητών",
                                            filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if f_name == "":
            return

        self.btn_get_students_file.configure(state='disabled')
        self.students_filename.set(f_name)
        self.students = parse_students(f_name)
        self.cb_school_col['values'] = self.students[0]
        self.cb_school_col.configure(state='readonly')
        self.urban = list()
        self.rural = list()

    def get_schools_filename(self):
        f_name = filedialog.askopenfilename(initialdir=".", title="Επιλέξτε το αρχείο των σχολείων",
                                            filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if f_name == "":
            return

        self.btn_get_schools_file.configure(state='disabled')
        self.btn_run.configure(state='normal')
        self.schools_filename.set(f_name)
        self.schools = parse_schools(f_name)

    def run(self):
        self.cb_school_col.configure(state='disabled')
        self.btn_run.configure(state='disabled')

        sc = self.cb_school_col.current()

        self.urban.append(self.students[0])
        self.rural.append(self.students[0] + ['ΣΧΟΛΕΙΟ ΚΑΤΑΝΟΜΗΣ', ])

        for student in self.students[1:]:
            if student[sc] in self.schools:
                self.rural.append(student + [self.schools[student[sc]], ])
            else:
                self.urban.append(student)

        save_file(self.urban, "urban.xlsx")
        save_file(self.rural, "rural.xlsx")

        showinfo(title="Αρχεία εξόδου",
                 message=f'Τα αποτελέσματα έχουν αποθηκευτεί στον φάκελο εκτέλεσης του προγράμματος.')

        self.window.destroy()

    def cb_school_col_select(self, event_object):
        if self.cb_school_col.current() != -1:
            self.ntr_schools_filename.configure(state='readonly')
            self.btn_get_schools_file.configure(state='normal')

    def create_widgets(self):
        self.f_data = Frame(self.window)

        self.l_students_file = Label(self.f_data, text="Αρχείο μαθητών:")
        self.l_students_file.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.students_filename = StringVar()
        self.students_filename.set('')
        self.ntr_students_filename = Entry(self.f_data, width=128, state='readonly',
                                           textvariable=self.students_filename)
        self.ntr_students_filename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btn_get_students_file = Button(self.f_data, text="Επιλέξτε αρχείο...", command=self.get_students_filename)
        self.btn_get_students_file.grid(column=2, row=0, padx=10, pady=10)

        self.l_school_col = Label(self.f_data, text="Στήλη Σχολείου Προέλευσης:")
        self.l_school_col.grid(column=0, row=1, padx=10, pady=10, sticky=E)

        self.school_col = StringVar()
        self.cb_school_col = Combobox(self.f_data, width=125, textvariable=self.school_col, state='disabled')
        self.cb_school_col.grid(column=1, row=1, padx=10, pady=10, sticky=W)
        self.cb_school_col.bind("<<ComboboxSelected>>", self.cb_school_col_select)

        self.l_schools_file = Label(self.f_data, text="Αρχείο σχολείων\nπεριφέρειας:")
        self.l_schools_file.grid(column=0, row=2, padx=10, pady=10, sticky=E)

        self.schools_filename = StringVar()
        self.schools_filename.set('')
        self.ntr_schools_filename = Entry(self.f_data, width=128, state='disabled', textvariable=self.schools_filename)
        self.ntr_schools_filename.grid(column=1, row=2, padx=10, pady=10, sticky=W)

        self.btn_get_schools_file = Button(self.f_data, text="Επιλέξτε αρχείο...", command=self.get_schools_filename,
                                           state='disabled')
        self.btn_get_schools_file.grid(column=2, row=2, padx=10, pady=10)

        self.btn_run = Button(self.f_data, text="Εκτέλεση διαχωρισμού", command=self.run, state='disabled')
        self.btn_run.grid(column=1, row=10, padx=10, pady=10)

        self.f_data.pack()


gui = GUI()
gui.window.mainloop()
