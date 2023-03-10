from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from openpyxl import *
from openpyxl.utils import get_column_letter
from pathlib import Path
import shutil
import re


def fix_cell(cell):
    text1 = str(cell.value).upper().replace(" .", ". ").replace("Ά", "Α").replace("Έ", "Ε") \
        .replace("Ή", "Η").replace("Ί", "Ι").replace("Ϊ́", "Ϊ").replace("Ύ", "Υ").replace("Ϋ́", "Ϋ") \
        .replace("Ό", "Ο").replace("Ώ", "Ω").strip()

    text2 = re.sub(r'([ ]+)', r' ', text1)
    text3 = re.sub(r'([0-9]+)Ο', r'\1ο', text2)

    return text3


def parse_students(file_name):
    wb = load_workbook(filename=file_name)
    ws = wb.active

    students = list()

    for row in ws.iter_rows():
        student = list()
        for cell in row:
            if cell.value is None:
                student.append("")
            else:
                text = fix_cell(cell)
                student.append(text)

        students.append(student)

    return students


def create_folders():
    dirpath = Path("./reports")
    if dirpath.exists() and dirpath.is_dir():
        shutil.rmtree(dirpath)

    dirpath = Path("./reports/xtras")
    dirpath.mkdir(parents=True, exist_ok=True)

    dirpath = Path("./reports/_ok")
    dirpath.mkdir(parents=True, exist_ok=True)

    dirpath = Path("./reports/_check")
    dirpath.mkdir(parents=True, exist_ok=True)


def save_file(data, output_file):
    wb = Workbook()
    ws = wb.active

    for entry in data:
        ws.append(entry)

    column_widths = []
    for row in ws.iter_rows():
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], len(str(cell.value)))
            except IndexError:
                column_widths.append(len(str(cell.value)))

    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

    not_saved = True

    while not_saved:
        try:
            wb.save(output_file)
        except:
            showwarning(title="Αρχείο σε χρήση...",
                        message=f"Παρακαλώ κλείστε το αρχείο '{output_file}' ώστε να ολοκληρωθεί η αποθήκευση.")
        else:
            not_saved = False


class GUI:
    def __init__(self):
        self.window = Tk()

        self.window.title("Φιλτράρισμα της κατανομής των μαθητών")
        self.window.resizable(False, False)
        self.create_widgets()

    def get_students_filename(self):
        f_name = filedialog.askopenfilename(initialdir=".", title="Επιλέξτε το αρχείο της κατανομής των μαθητών",
                                            filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if f_name == "":
            return

        self.btn_get_students_file.configure(state='disabled')
        self.students_filename.set(f_name)
        self.students = parse_students(f_name)
        self.cb_google_col['values'] = self.students[0]
        self.cb_google_col.configure(state='readonly')

    def clean_parentheses(self):
        gmc = self.cb_google_col.current()

        for student in self.students[1:]:
            for i in [0, 3, 6, 9]:
                if '(' in student[gmc + i]:
                    student[gmc + i] = re.sub("\(.*", "", student[gmc + i]).strip()

    def check_urban(self):
        gmc = self.cb_google_col.current()

        self.urban_ok = list()
        self.urban_ok.append(self.students[0] + ['ΣΧΟΛΕΙΟ ΚΑΤΑΝΟΜΗΣ', 'ΣΥΝΤΕΤΑΓΜΕΝΕΣ ΚΑΤΑΝΟΜΗΣ'])
        self.urban_check = list()
        self.urban_check.append(self.students[0])

        count_4 = 0
        count_3 = 0

        self.only_postal_code = list()
        self.only_postal_code.append(self.students[0])
        self.not_found = list()
        self.not_found.append(self.students[0])

        for student in self.students[1:]:
            school = ''
            coords = ''
            success = False

            if student[gmc] == student[gmc + 3] == student[gmc + 6] == student[gmc + 9]:
                if student[gmc] != 'N/A':
                    success = True
                    count_4 += 1
                    school = student[gmc]
                    coords = student[gmc + 2]
                else:
                    self.not_found.append(student)
            elif student[gmc] == student[gmc + 3] == student[gmc + 6]:
                if student[gmc] != 'N/A':
                    if not student[gmc + 1].startswith(('ΗΡΑΚΛΕΙΟ', 'ΜΑΛΕΒΙΖΙ')) or not student[gmc + 4].startswith(
                            ('ΗΡΑΚΛΕΙΟ', 'ΜΑΛΕΒΙΖΙ')):
                        success = True
                        count_3 += 1
                        school = student[gmc]
                        coords = student[gmc + 2]
                    else:
                        self.only_postal_code.append(student)
                else:
                    self.not_found.append(student)
            elif student[gmc] == student[gmc + 3] == student[gmc + 9]:
                if student[gmc] != 'N/A':
                    if not student[gmc + 1].startswith(('ΗΡΑΚΛΕΙΟ', 'ΜΑΛΕΒΙΖΙ')) or not student[gmc + 4].startswith(
                            ('ΗΡΑΚΛΕΙΟ', 'ΜΑΛΕΒΙΖΙ')):
                        success = True
                        count_3 += 1
                        school = student[gmc]
                        coords = student[gmc + 2]
                    else:
                        self.only_postal_code.append(student)
                else:
                    self.not_found.append(student)
            elif student[gmc] == student[gmc + 6] == student[gmc + 9]:
                if student[gmc] != 'N/A':
                    if not student[gmc + 1].startswith(('ΗΡΑΚΛΕΙΟ', 'ΜΑΛΕΒΙΖΙ')):
                        success = True
                        count_3 += 1
                        school = student[gmc]
                        coords = student[gmc + 2]
                    else:
                        self.only_postal_code.append(student)
                else:
                    self.not_found.append(student)
            elif student[gmc + 3] == student[gmc + 6] == student[gmc + 9]:
                if student[gmc + 3] != 'N/A':
                    if not student[gmc + 4].startswith(('ΗΡΑΚΛΕΙΟ', 'ΜΑΛΕΒΙΖΙ')):
                        success = True
                        count_3 += 1
                        school = student[gmc + 3]
                        coords = student[gmc + 5]
                    else:
                        self.only_postal_code.append(student)
                else:
                    self.not_found.append(student)

            if success:
                self.urban_ok.append(student + [school, coords])
            else:
                self.urban_check.append(student)

    def export_data(self):
        ok_count = len(self.urban_ok[1:])
        check_count = len(self.urban_check[1:])
        count_only_postal_code = len(self.only_postal_code[1:])
        count_not_found = len(self.not_found[1:])
        save_file(self.urban_ok, f"./reports/_ok/urban_ok ({ok_count}).xlsx")
        save_file(self.urban_check, f"./reports/_check/urban_check ({check_count}).xlsx")
        save_file(self.only_postal_code,
                  f"./reports/xtras/urban_check_only_postal_code ({count_only_postal_code}).xlsx")
        save_file(self.not_found, f"./reports/xtras/urban_check_not_found ({count_not_found}).xlsx")

        print(20 * '-', ' Έλεγχος σε μαθητές Πόλης ', 20 * '-')
        print('Πλήθος μαθητών ΟΚ: ', ok_count)
        print('Πλήθος μαθητών για νέο έλεγχο: ', check_count)
        print('Διευθύνσεις που εντοπίστηκαν μόνο με το ΤΚ: ', count_only_postal_code)
        print('Πλήθος σχολείων που δεν εντοπίστηκαν: ', count_not_found)
        print('Σύνολο: ', ok_count + check_count)

    def run(self):
        self.btn_run.configure(state='disabled')

        create_folders()

        self.clean_parentheses()
        self.check_urban()
        self.export_data()

        showinfo(title="Αρχεία εξόδου",
                 message=f'Τα αποτελέσματα έχουν αποθηκευτεί στον φάκελο "reports".')

        self.window.destroy()

    def cb_google_col_select(self, event_object):
        if self.cb_google_col.current() != -1:
            self.cb_google_col.configure(state='disabled')
            self.btn_run.configure(state='normal')

    def create_widgets(self):
        self.f_data = Frame(self.window)

        self.l_students_file = Label(self.f_data, text="Αρχείο κατανομής μαθητών:")
        self.l_students_file.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.students_filename = StringVar()
        self.students_filename.set('')
        self.ntr_students_filename = Entry(self.f_data, width=128, state='readonly',
                                           textvariable=self.students_filename)
        self.ntr_students_filename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btn_get_students_file = Button(self.f_data, text="Επιλέξτε αρχείο...", command=self.get_students_filename)
        self.btn_get_students_file.grid(column=2, row=0, padx=10, pady=10)

        self.l_google_col = Label(self.f_data, text="Στήλη Σχολείου κατανομής Google Maps:")
        self.l_google_col.grid(column=0, row=1, padx=10, pady=10, sticky=E)

        self.google_col = StringVar()
        self.cb_google_col = Combobox(self.f_data, width=125, textvariable=self.google_col, state='disabled')
        self.cb_google_col.bind("<<ComboboxSelected>>", self.cb_google_col_select)
        self.cb_google_col.grid(column=1, row=1, padx=10, pady=10, sticky='W')

        self.btn_run = Button(self.f_data, text="Εκτέλεση", command=self.run, state='disabled')
        self.btn_run.grid(column=1, row=10, padx=10, pady=10)

        self.f_data.pack()


gui = GUI()
gui.window.mainloop()
