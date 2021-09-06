from tkinter import *
from tkinter import filedialog
from tkinter.scrolledtext import ScrolledText
from tkinter.messagebox import showwarning, showinfo
from tkinter.ttk import *
from parseMapData import createLists
from schoolsPolygons import *
from parseXlsxData import *
from addressCoordinates import *
from geopy.geocoders import Here, Bing
import googlemaps
from openpyxl import *
from openpyxl.utils import get_column_letter


class GUI():
    def __init__(self):
        self.window = Tk()

        self.window.title("Κατανομή μαθητών σε Σχολικές Μονάδες")
        self.window.resizable(False, False)
        self.create_widgets()


    def getMapDataFilename(self):
        fName = filedialog.askopenfilename(initialdir="./data/", title="Επιλέξτε το αρχείο με τις περιοχές των Σχολικών Μονάδων",
                                           filetypes=(("kml files", "*.kml"), ("all files", "*.*")))

        if fName == "":
            return

        self.mapDataFilename.set(fName)
        self.namesList, self.polygonsList = createLists(fName)

        sameSize, debugText = checkListsLengths(self.namesList, self.polygonsList)
        self.stDebug.insert(INSERT, debugText)
        polygonsAreValid, debugText = checkPolygonsValidity(self.namesList, self.polygonsList)
        self.stDebug.insert(INSERT, debugText)
        noIntersections, debugText = checkPolygonsIntersections(self.namesList, self.polygonsList)
        self.stDebug.insert(INSERT, debugText)

        if sameSize:
            self.lCount.configure(background='green')
        else:
            self.lCount.configure(background='red')

        if polygonsAreValid:
            self.lPolygonsValidity.configure(background='green')
        else:
            self.lPolygonsValidity.configure(background='red')

        if noIntersections:
            self.lPolygonsIntersections.configure(background='green')
        else:
            self.lPolygonsIntersections.configure(background='red')

        self.btnNextGoToStudents.configure(state='normal')


    def nextGoToTabStudents(self):
        self.tabControl.tab(0, state='disabled')
        self.tabControl.tab(1, state='normal')
        self.tabControl.select(1)


    def getStudentsFilename(self):
        fName = filedialog.askopenfilename(initialdir="./data/", title="Επιλέξτε το αρχείο με τους μαθητές",
                                           filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))

        if fName == "":
            return

        self.studentsFilename.set(fName)
        self.students = parseXlsxData(fName)
        self.cbStudentsCols['values'] = self.students[0]
        self.studentsCount.set(len(self.students) - 1)
        self.pbProgress['maximum'] = self.studentsCount.get()


    def cbStudentsColsSelect(self, eventObject):
        self.studentAddrCol = self.cbStudentsCols.current()

        if self.cbStudentsCols.current() != -1:
            self.btnNextGoToKeysAPIs.configure(state='normal')


    def checkGKey(self):
        self.myGoogleKey = self.googleKey.get()
        if self.myGoogleKey != "":
            try:
                gmaps = googlemaps.Client(key=self.myGoogleKey)
            except:
                self.lGoogleKey.configure(background='red')
            else:
                self.ntrGoogleKey.configure(state='disabled')
                self.btnCheckGKey.configure(state='disabled')
                self.lGoogleKey.configure(background='green')
                #self.ckbGoogleMaps.configure(state='normal')
                self.GMchecked.set(True)
                #self.ckbGoogleV3.configure(state='normal')
                self.GV3checked.set(True)
                self.btnNextGoToRun.configure(state='normal')


    def checkBKey(self):
        self.myBingKey = self.bingKey.get()
        if self.myBingKey != "":
            try:
                bing = Bing(api_key=self.myBingKey, timeout=1000)
                result = bing.geocode("test", exactly_one=True)
            except:
                self.lBingKey.configure(background='red')
            else:
                self.ntrBingKey.configure(state='disabled')
                self.btnCheckBKey.configure(state='disabled')
                self.lBingKey.configure(background='green')
                #self.ckbBingMaps.configure(state='normal')
                self.BMchecked.set(True)
                self.btnNextGoToRun.configure(state='normal')


    def checkHKey(self):
        self.myHereKey = self.hereKey.get()
        if self.myHereKey != "":
            try:
                here = Here(apikey=self.myHereKey, timeout=1000)
                result = here.geocode("test", exactly_one=True, language="el")
            except:
                self.lHereKey.configure(background='red')
            else:
                self.ntrHereKey.configure(state='disabled')
                self.btnCheckHKey.configure(state='disabled')
                self.lHereKey.configure(background='green')
                #self.ckbHereMaps.configure(state='normal')
                self.HMchecked.set(True)
                self.btnNextGoToRun.configure(state='normal')


    def prevGoToTabMapData(self):
        self.tabControl.tab(1, state='disabled')
        self.tabControl.tab(0, state='normal')
        self.tabControl.select(0)


    def nextGoToTabKeysAPIs(self):
        self.tabControl.tab(1, state='disabled')
        self.tabControl.tab(2, state='normal')
        self.tabControl.select(2)


    def prevGoToTabStudents(self):
        self.tabControl.tab(2, state='disabled')
        self.tabControl.tab(1, state='normal')
        self.tabControl.select(1)


    def nextGoToTabRun(self):
        self.tabControl.tab(2, state='disabled')
        self.tabControl.tab(3, state='normal')
        self.tabControl.select(3)


    def prevGoToTabKeysAPIs(self):
        self.tabControl.tab(3, state='disabled')
        self.tabControl.tab(2, state='normal')
        self.tabControl.select(2)


    def run(self):
        self.lProgress.configure(text="Εκτέλεση κατανομής σε εξέλιξη...")

        data = list()

        for student in self.students:
            data.append(student.copy())

        for i in range(len(data)):
            if i == 0:
                if self.GMchecked.get():
                    data[i].append("Σχολείο (Google Maps)")
                    data[i].append("Διεύθυνση (Google Maps)")
                    data[i].append("Συντεταγμένες (Google Maps)")
                if self.GV3checked.get():
                    data[i].append("Σχολείο (Google V3)")
                    data[i].append("Διεύθυνση (Google V3)")
                    data[i].append("Συντεταγμένες (Google V3)")
                if self.BMchecked.get():
                    data[i].append("Σχολείο (Bing Maps)")
                    data[i].append("Διεύθυνση (Bing Maps)")
                    data[i].append("Συντεταγμένες (Bing Maps)")
                if self.HMchecked.get():
                    data[i].append("Σχολείο (Here Maps)")
                    data[i].append("Διεύθυνση (Here Maps)")
                    data[i].append("Συντεταγμένες (Here Maps)")

                continue

            self.lProgress.configure(text="Εκτέλεση κατανομής σε εξέλιξη... ({}/{})".format(i, self.studentsCount.get()))

            if self.GMchecked.get():
                point, address = searchAddressInGoogleMaps(self.myGoogleKey, self.students[i][self.studentAddrCol])

                if address != "Can't find the address." and address != "Exception.":
                    school = searchPointInsidePolygons(point, self.namesList, self.polygonsList)
                    if not school:
                        schoolText = "Η διεύθυνση εντοπίζεται σε όριο περιοχών."
                    else:
                        schoolText = str(school).replace("[\'", "").replace("\']", "")
                else:
                    schoolText = "N/A"

                data[i].append(schoolText)
                data[i].append(address)
                data[i].append(f'{point.x},{point.y}')

            if self.GV3checked.get():
                point, address = searchAddressInGoogleV3(self.myGoogleKey, self.students[i][self.studentAddrCol])

                if address != "Can't find the address." and address != "Exception.":
                    school = searchPointInsidePolygons(point, self.namesList, self.polygonsList)
                    if not school:
                        schoolText = "Η διεύθυνση εντοπίζεται σε όριο περιοχών."
                    else:
                        schoolText = str(school).replace("[\'", "").replace("\']", "")
                else:
                    schoolText = "N/A"

                data[i].append(schoolText)
                data[i].append(address)
                data[i].append(f'{point.x},{point.y}')

            if self.BMchecked.get():
                point, address = searchAddressInBingMaps(self.myBingKey, self.students[i][self.studentAddrCol])

                if address != "Can't find the address." and address != "Exception.":
                    school = searchPointInsidePolygons(point, self.namesList, self.polygonsList)
                    if not school:
                        schoolText = "Η διεύθυνση εντοπίζεται σε όριο περιοχών."
                    else:
                        schoolText = str(school).replace("[\'", "").replace("\']", "")
                else:
                    schoolText = "N/A"

                data[i].append(schoolText)
                data[i].append(address)
                data[i].append(f'{point.x},{point.y}')

            if self.HMchecked.get():
                point, address = searchAddressInHereMaps(self.myHereKey, self.students[i][self.studentAddrCol])

                if address != "Can't find the address." and address != "Exception.":
                    school = searchPointInsidePolygons(point, self.namesList, self.polygonsList)
                    if not school:
                        schoolText = "Η διεύθυνση εντοπίζεται σε όριο περιοχών."
                    else:
                        schoolText = str(school).replace("[\'", "").replace("\']", "")
                else:
                    schoolText = "N/A"

                data[i].append(schoolText)
                data[i].append(address)
                data[i].append(f'{point.x},{point.y}')

            self.pbProgress['value'] = i
            self.pbProgress.update()

        self.lProgress.configure(text="Ολοκλήρωση κατανομής.\nΑποθήκευση σε αρχείο...")

        wb = Workbook()
        ws = wb.active

        for student in data:
            ws.append(student)

        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.23

        outputFile = self.studentsFilename.get().replace(".xlsx", "_output.xlsx")
        showinfo(title="Αρχείο εξόδου",
                    message="Η κατανομή θα αποθηκευτεί στο αρχείο: " + outputFile)

        notSaved = True

        while notSaved:
            try:
                wb.save(outputFile)
            except:
                showwarning(title="Αρχείο σε χρήση...", message="Παρακαλώ κλείστε το αρχείο '{}' ώστε να ολοκληρωθεί η αποθήκευση της νέας κατανομής.".format(outputFile))
            else:
                notSaved = False

        self.lProgress.configure(text="Ολοκλήρωση όλων των εργασιών.")
        self.pbProgress['value'] = 0
        self.pbProgress.update()


    def create_widgets(self):
        # Tabs
        self.tabControl = Notebook(self.window)
        self.tabMapData = Frame(self.tabControl)
        self.tabControl.add(self.tabMapData, text="Περιοχές Σχολικών Μονάδων")
        self.tabStudents = Frame(self.tabControl)
        self.tabControl.add(self.tabStudents, text="Μαθητές", state='disabled')
        self.tabKeysAPIs = Frame(self.tabControl)
        self.tabControl.add(self.tabKeysAPIs, text="Κλειδιά και Αναζήτηση", state='disabled')
        self.tabRun = Frame(self.tabControl)
        self.tabControl.add(self.tabRun, text="Εκτέλεση κατανομής", state='disabled')
        self.tabControl.pack(expand=1, fill="both")

        # Tab: MapData
        self.fMapData = Frame(self.tabMapData)

        self.lMapData = Label(self.fMapData, text="Αρχείο:")
        self.lMapData.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.mapDataFilename = StringVar()
        self.ntrMapDataFilename = Entry(self.fMapData, width=128, state='readonly', textvariable=self.mapDataFilename)
        self.ntrMapDataFilename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenMapData = Button(self.fMapData, text="Επιλέξτε αρχείο...", command=self.getMapDataFilename)
        self.btnOpenMapData.grid(column=2, row=0, padx=10, pady=10)

        self.lfVerifyFrame = LabelFrame(self.fMapData, text="Έλεγχοι")
        self.lfVerifyFrame.grid(column=0, row=1, columnspan=3, padx=10, pady=10)

        self.lCount = Label(self.lfVerifyFrame, text="Πλήθος ονομάτων == Πλήθος πολυγώνων")
        self.lCount.grid(column=0, row=0, padx=10, pady=10)

        self.lPolygonsValidity = Label(self.lfVerifyFrame, text="Ορθότητα πολυγώνων")
        self.lPolygonsValidity.grid(column=0, row=1, padx=10, pady=10)

        self.lPolygonsIntersections = Label(self.lfVerifyFrame, text="Επικαλύψεις πολυγώνων")
        self.lPolygonsIntersections.grid(column=0, row=2, padx=10, pady=10)

        self.stDebug = ScrolledText(self.lfVerifyFrame, width=80, height=10, wrap=WORD)
        self.stDebug.grid(column=1, row=0, columnspan=2, rowspan=3, padx=10, pady=10)

        self.btnNextGoToStudents = Button(self.fMapData, text="Επόμενο", command=self.nextGoToTabStudents, state='disabled')
        self.btnNextGoToStudents.grid(column=2, row=10, padx=10, pady=10)

        self.fMapData.pack()

        # Tab: Students
        self.fStudents = Frame(self.tabStudents)

        self.lStudents = Label(self.fStudents, text="Αρχείο:")
        self.lStudents.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.studentsFilename = StringVar()
        self.ntrStudentsFilename = Entry(self.fStudents, width=128, state='readonly', textvariable=self.studentsFilename)
        self.ntrStudentsFilename.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnOpenStudents = Button(self.fStudents, text="Επιλέξτε αρχείο...", command=self.getStudentsFilename)
        self.btnOpenStudents.grid(column=2, row=0, padx=10, pady=10)

        self.lAddressCol = Label(self.fStudents, text="Στήλη Διεύθυνσης:")
        self.lAddressCol.grid(column=0, row=1, padx=10, pady=10, sticky=E)

        self.studentsCols = StringVar()
        self.cbStudentsCols = Combobox(self.fStudents, width=40, textvariable=self.studentsCols, state='readonly')
        self.cbStudentsCols.grid(column=1, row=1, padx=10, pady=10, sticky=W)
        self.cbStudentsCols.bind("<<ComboboxSelected>>", self.cbStudentsColsSelect)

        self.lStudentsCount = Label(self.fStudents, text="Πλήθος μαθητών:")
        self.lStudentsCount.grid(column=0, row=2, padx=10, pady=10, sticky=E)

        self.studentsCount = StringVar()
        self.ntrStudentsCount = Entry(self.fStudents, width=12, state='readonly', textvariable=self.studentsCount)
        self.ntrStudentsCount.grid(column=1, row=2, padx=10, pady=10, sticky=W)

        self.btnPrevGoToMapData = Button(self.fStudents, text="Προηγούμενο", command=self.prevGoToTabMapData)
        self.btnPrevGoToMapData.grid(column=0, row=10, padx=10, pady=10)

        self.btnNextGoToKeysAPIs = Button(self.fStudents, text="Επόμενο", command=self.nextGoToTabKeysAPIs, state='disabled')
        self.btnNextGoToKeysAPIs.grid(column=2, row=10, padx=10, pady=10)

        self.fStudents.pack()

        # Tab: KeysAPIs
        self.fKeysAPIs = Frame(self.tabKeysAPIs)

        self.lfKeysFrame = LabelFrame(self.fKeysAPIs, text="Κλειδιά")
        self.lfKeysFrame.grid(column=0, row=0, columnspan=3, padx=10, pady=10)

        self.lGoogleKey = Label(self.lfKeysFrame, text="Google:")
        self.lGoogleKey.grid(column=0, row=0, padx=10, pady=10, sticky=E)

        self.googleKey = StringVar()
        self.ntrGoogleKey = Entry(self.lfKeysFrame, width=128, textvariable=self.googleKey)
        self.ntrGoogleKey.grid(column=1, row=0, padx=10, pady=10, sticky=W)

        self.btnCheckGKey = Button(self.lfKeysFrame, text="Έλεγξε το κλειδί", command=self.checkGKey)
        self.btnCheckGKey.grid(column=2, row=0, padx=10, pady=10)

        self.lBingKey = Label(self.lfKeysFrame, text="Bing:")
        self.lBingKey.grid(column=0, row=1, padx=10, pady=10, sticky=E)

        self.bingKey = StringVar()
        self.ntrBingKey = Entry(self.lfKeysFrame, width=128, textvariable=self.bingKey)
        self.ntrBingKey.grid(column=1, row=1, padx=10, pady=10, sticky=W)

        self.btnCheckBKey = Button(self.lfKeysFrame, text="Έλεγξε το κλειδί", command=self.checkBKey)
        self.btnCheckBKey.grid(column=2, row=1, padx=10, pady=10)

        self.lHereKey = Label(self.lfKeysFrame, text="Here:")
        self.lHereKey.grid(column=0, row=2, padx=10, pady=10, sticky=E)

        self.hereKey = StringVar()
        self.ntrHereKey = Entry(self.lfKeysFrame, width=128, textvariable=self.hereKey)
        self.ntrHereKey.grid(column=1, row=2, padx=10, pady=10, sticky=W)

        self.btnCheckHKey = Button(self.lfKeysFrame, text="Έλεγξε το κλειδί", command=self.checkHKey)
        self.btnCheckHKey.grid(column=2, row=2, padx=10, pady=10)

        self.lfAPIsFrame = LabelFrame(self.fKeysAPIs, text="Αναζήτηση με")
        self.lfAPIsFrame.grid(column=0, row=1, columnspan=3, padx=10, pady=10)

        self.GMchecked = BooleanVar()
        self.ckbGoogleMaps = Checkbutton(self.lfAPIsFrame, text="Google Maps", variable=self.GMchecked, state='disabled')
        self.ckbGoogleMaps.grid(column=0, row=0, padx=10, pady=10)

        self.GV3checked = BooleanVar()
        self.ckbGoogleV3 = Checkbutton(self.lfAPIsFrame, text="Google V3", variable=self.GV3checked, state='disabled')
        self.ckbGoogleV3.grid(column=1, row=0, padx=10, pady=10)

        self.BMchecked = BooleanVar()
        self.ckbBingMaps = Checkbutton(self.lfAPIsFrame, text="Bing Maps", variable=self.BMchecked, state='disabled')
        self.ckbBingMaps.grid(column=2, row=0, padx=10, pady=10)

        self.HMchecked = BooleanVar()
        self.ckbHereMaps = Checkbutton(self.lfAPIsFrame, text="Here Maps", variable=self.HMchecked, state='disabled')
        self.ckbHereMaps.grid(column=3, row=0, padx=10, pady=10)

        self.btnPrevGoToStudents = Button(self.fKeysAPIs, text="Προηγούμενο", command=self.prevGoToTabStudents)
        self.btnPrevGoToStudents.grid(column=0, row=10, padx=10, pady=10)

        self.btnNextGoToRun = Button(self.fKeysAPIs, text="Επόμενο", command=self.nextGoToTabRun, state='disabled')
        self.btnNextGoToRun.grid(column=2, row=10, padx=10, pady=10)

        self.fKeysAPIs.pack()

        # Tab: Run
        self.fRun = Frame(self.tabRun)

        self.lProgress = Label(self.fRun, text="Αναμονή για εκτέλεση κατανομής ...")
        self.lProgress.grid(column=0, row=0, columnspan=3, padx=10, pady=10)

        self.pbProgress = Progressbar(self.fRun, orient='horizontal', length=300, mode='determinate')
        self.pbProgress.grid(column=0, row=1, columnspan=3, padx=10, pady=10)

        self.btnPrevGoToKeysAPIs = Button(self.fRun, text="Προηγούμενο", command=self.prevGoToTabKeysAPIs)
        self.btnPrevGoToKeysAPIs.grid(column=0, row=10, padx=10, pady=10)

        self.btnRun = Button(self.fRun, text="Εκτέλεση κατανομής", command=self.run)
        self.btnRun.grid(column=2, row=10, padx=10, pady=10)

        self.fRun.pack()


gui = GUI()
gui.window.mainloop()
