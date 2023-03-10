from openpyxl import load_workbook
import re


def parseXlsxData(fileName):
    workbook = load_workbook(filename=fileName)
    sheet = workbook.active

    students = list()

    for row in sheet.iter_rows():
        student = list()
        for cell in row:
            if cell.value is None:
                student.append("")
            else:
                text1 = (
                    str(cell.value).upper().replace(".", ". ").replace(" .", ". ").replace("Ά", "Α").replace("Έ", "Ε")
                    .replace("Ή", "Η").replace("Ί", "Ι").replace("Ϊ́", "Ϊ").replace("Ύ", "Υ").replace("Ϋ́", "Ϋ")
                    .replace("Ό", "Ο").replace("Ώ", "Ω").strip())

                text2 = re.sub(r'([ ]+)', r' ', text1)
                student.append(re.sub(r'([0-9]+)Ο', r'\1ο', text2))

        students.append(student)

    return students
